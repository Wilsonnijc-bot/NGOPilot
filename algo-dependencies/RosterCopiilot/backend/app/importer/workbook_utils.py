"""Small openpyxl helpers used by importer smoke tests and future parsers."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from .errors import MissingSheetError, WorkbookLoadError


@dataclass(frozen=True, slots=True)
class UsedRange:
    min_row: int | None
    min_column: int | None
    max_row: int | None
    max_column: int | None
    value_count: int

    @property
    def empty(self) -> bool:
        return self.value_count == 0

    @property
    def height(self) -> int:
        if self.empty or self.min_row is None or self.max_row is None:
            return 0
        return self.max_row - self.min_row + 1

    @property
    def width(self) -> int:
        if self.empty or self.min_column is None or self.max_column is None:
            return 0
        return self.max_column - self.min_column + 1

    @property
    def max_column_letter(self) -> str | None:
        if self.max_column is None:
            return None
        return get_column_letter(self.max_column)


def load_workbook(
    path: Path | str,
    *,
    data_only: bool = False,
    read_only: bool = False,
) -> Workbook:
    """Open an Excel workbook with importer-specific error wrapping."""

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise WorkbookLoadError(f"workbook does not exist: {workbook_path}")
    try:
        return openpyxl_load_workbook(
            workbook_path,
            data_only=data_only,
            read_only=read_only,
        )
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise WorkbookLoadError(f"failed to load workbook: {workbook_path}") from exc


def require_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    """Return a worksheet or raise a typed importer error."""

    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise MissingSheetError(
            f"missing required sheet {sheet_name!r}; available sheets: {available}"
        )
    return workbook[sheet_name]


def effective_used_range(worksheet: Worksheet) -> UsedRange:
    """Return the bounds of cells with non-empty values.

    ``Worksheet.max_row`` can include formatting-only rows in the NGO files.
    The importer contract needs the value-backed range for shape checks.
    """

    min_row: int | None = None
    min_column: int | None = None
    max_row: int | None = None
    max_column: int | None = None
    value_count = 0

    for row in worksheet.iter_rows():
        for cell in row:
            if _is_blank(cell.value):
                continue
            value_count += 1
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            min_column = (
                cell.column if min_column is None else min(min_column, cell.column)
            )
            max_row = cell.row if max_row is None else max(max_row, cell.row)
            max_column = (
                cell.column if max_column is None else max(max_column, cell.column)
            )

    return UsedRange(
        min_row=min_row,
        min_column=min_column,
        max_row=max_row,
        max_column=max_column,
        value_count=value_count,
    )


def merged_cell_range(
    worksheet: Worksheet,
    cell_or_coordinate: Cell | MergedCell | str,
) -> CellRange | None:
    """Return the merged range containing a coordinate, if any."""

    coordinate = _coordinate(cell_or_coordinate)
    for cell_range in worksheet.merged_cells.ranges:
        if coordinate in cell_range:
            return cell_range
    return None


def resolve_merged_cell(
    worksheet: Worksheet,
    cell_or_coordinate: Cell | MergedCell | str,
) -> Cell:
    """Return the anchor cell for a merged coordinate, else the cell itself."""

    cell_range = merged_cell_range(worksheet, cell_or_coordinate)
    if cell_range is None:
        row, column = coordinate_to_tuple(_coordinate(cell_or_coordinate))
        return worksheet.cell(row=row, column=column)
    return worksheet.cell(row=cell_range.min_row, column=cell_range.min_col)


def read_cell_value(
    worksheet: Worksheet,
    cell_or_coordinate: Cell | MergedCell | str,
    *,
    resolve_merged: bool = True,
) -> Any:
    """Read a cell value, optionally resolving merged cells to their anchor."""

    if resolve_merged:
        return resolve_merged_cell(worksheet, cell_or_coordinate).value
    if isinstance(cell_or_coordinate, str):
        return worksheet[cell_or_coordinate].value
    return cell_or_coordinate.value


def read_fill_color(
    cell: Cell | MergedCell,
    *,
    worksheet: Worksheet | None = None,
    resolve_merged: bool = True,
) -> str | None:
    """Return a stable foreground-fill token for a cell.

    RGB fills return uppercase ARGB strings such as ``FFF4CCCC``. Theme and
    indexed fills are returned as explicit tokens because they depend on the
    workbook palette.
    """

    resolved_cell = cell
    if resolve_merged and isinstance(cell, MergedCell):
        if worksheet is None:
            worksheet = cell.parent
        resolved_cell = resolve_merged_cell(worksheet, cell)

    fill = resolved_cell.fill
    if fill.fill_type is None:
        return None

    color = fill.fgColor
    if color.type == "rgb":
        return str(color.rgb).upper()
    if color.type == "theme":
        return f"theme:{color.theme}:{color.tint}"
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.value is None:
        return None
    return str(color.value)


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _coordinate(cell_or_coordinate: Cell | MergedCell | str) -> str:
    if isinstance(cell_or_coordinate, str):
        return cell_or_coordinate
    return cell_or_coordinate.coordinate
