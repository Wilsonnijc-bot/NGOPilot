"""Importer for the escort case master workbook (護送個案總表)."""
from __future__ import annotations

import re
import unicodedata
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .base import ImportResult
from .models import CellRef, ImportAmbiguity, ImportBatchSummary, ParsedRecord, SourceRef
from .workbook_utils import load_workbook, require_sheet

DOC_REF = "docs/spec/excel_semantics.md §3; docs/spec/excel_io_contract.md §1"
MONTH_SHEET = "1月"
HEADER_ROW = 5
DATA_START_ROW = 6
CHANGE_SECTION_ROW = 148

PERIOD_MAP = {"上午": "AM", "下午": "PM"}
PREFERENCE_PATTERNS = (
    ("must", re.compile(r"只要\s*([一-鿿A-Za-z]{1,6})")),
    ("prefer", re.compile(r"建議安排(?:照顧員)?\s*([一-鿿A-Za-z]{1,8}?)(?:陪診|$|[),，）])")),
    ("prefer", re.compile(r"盡量安排(?:照顧員)?\s*([一-鿿A-Za-z]{1,8}?)(?:陪診|$|[,(，）])")),
    ("prefer", re.compile(r"Req\s*([一-鿿A-Za-z]{1,8}?)(?:陪診|$)")),
    ("prefer", re.compile(r"^安排啊?\s*([一-鿿A-Za-z]{1,6})$")),
)


def _normalize(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("：", ":").replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", " ", text).strip()


def _source(path: Path | None, ws: Worksheet, row: int, col: int,
            value: object = None) -> SourceRef:
    return SourceRef(
        workbook_path=path,
        sheet_name=ws.title,
        cell=CellRef(sheet_name=ws.title, row=row, column=col),
        raw_value=value,
        doc_ref=DOC_REF,
    )


def _date_iso(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def _parse_time(value: object, period: str | None) -> tuple[str | None, str | None]:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        hour, minute = value.hour, value.minute
        if period == "PM" and 1 <= hour < 12:
            hour += 12
        return f"{hour:02d}:{minute:02d}", None
    text = _normalize(value)
    if not text:
        return None, "missing appointment time"
    m = re.search(r"(\d{1,2})[:：](\d{2})", text)
    if not m:
        return None, f"unparsed appointment time: {text!r}"
    hour, minute = int(m.group(1)), int(m.group(2))
    if period == "PM" and 1 <= hour < 12:
        hour += 12
    return f"{hour:02d}:{minute:02d}", None


def _preference(note: str | None) -> tuple[str | None, str | None]:
    if not note:
        return None, None
    text = _normalize(note)
    for strength, pattern in PREFERENCE_PATTERNS:
        match = pattern.search(text)
        if match:
            alias = match.group(1).strip()
            alias = re.sub(r"(姐|姑娘|照顧員)$", "", alias).strip()
            if alias:
                return alias, strength
    return None, None


def parse_workbook(workbook: Workbook | Path) -> ImportResult[dict[str, Any]]:
    """Parse the month sheet into escort request candidates.

    Path input is preferred because formulas in the date column need the cached
    data-only values. Workbook input is accepted for tests but cannot recover
    formulas that openpyxl has not evaluated.
    """

    workbook_path = workbook if isinstance(workbook, Path) else None
    wb = load_workbook(workbook_path, data_only=True) if workbook_path else workbook
    return parse_month_sheet(require_sheet(wb, MONTH_SHEET), workbook_path=workbook_path)


def parse_month_sheet(
    worksheet: Worksheet,
    *,
    workbook_path: Path | None = None,
) -> ImportResult[dict[str, Any]]:
    records: list[ParsedRecord[dict[str, Any]]] = []
    ambiguities: list[ImportAmbiguity] = []
    current_date: object | None = None
    current_weekday: str | None = None
    histogram: Counter[str] = Counter()
    time_parsed = 0
    time_seen = 0
    preference_count = 0

    for row in range(DATA_START_ROW, CHANGE_SECTION_ROW):
        date_cell = worksheet.cell(row, 1).value
        weekday_cell = worksheet.cell(row, 2).value
        if date_cell not in (None, ""):
            current_date = date_cell
        if weekday_cell not in (None, ""):
            current_weekday = _normalize(weekday_cell)

        name = _normalize(worksheet.cell(row, 4).value)
        if not name:
            continue

        period_raw = _normalize(worksheet.cell(row, 3).value)
        period = PERIOD_MAP.get(period_raw)
        service_date = _date_iso(current_date)
        if period is None or service_date is None:
            ambiguities.append(ImportAmbiguity(
                code="MISSING_PERIOD" if period is None else "MISSING_DATE",
                message=(
                    f"escort row {row} cannot be scheduled because date/period "
                    f"is incomplete (date={service_date!r}, period={period_raw!r})"
                ),
                source=_source(workbook_path, worksheet, row, 4,
                               worksheet.cell(row, 4).value),
                severity="blocking",
                raw_value=name,
                resolution_hint="confirm the intended date and AM/PM slot",
            ))
            continue

        appointment_raw = worksheet.cell(row, 6).value
        appointment_time, time_warning = _parse_time(appointment_raw, period)
        if appointment_raw not in (None, ""):
            time_seen += 1
        if appointment_time:
            time_parsed += 1
        elif time_warning:
            ambiguities.append(ImportAmbiguity(
                code="UNPARSED_APPOINTMENT_TIME",
                message=time_warning,
                source=_source(workbook_path, worksheet, row, 6, appointment_raw),
                severity="warning",
                raw_value=_normalize(appointment_raw),
            ))

        note = _normalize(worksheet.cell(row, 10).value) or None
        preferred_worker_alias, preference_strength = _preference(note)
        if preferred_worker_alias:
            preference_count += 1

        record = {
            "source": "escort_workbook",
            "source_ref": f"{worksheet.title}!D{row}",
            "row": row,
            "service_date": service_date,
            "weekday_raw": current_weekday,
            "period": period,
            "period_raw": period_raw,
            "elder_alias": name,
            "unit": _normalize(worksheet.cell(row, 5).value) or None,
            "appointment_time": appointment_time,
            "appointment_time_raw": _normalize(appointment_raw) or None,
            "destination": _normalize(worksheet.cell(row, 7).value) or None,
            "subject": _normalize(worksheet.cell(row, 8).value) or None,
            "transport": _normalize(worksheet.cell(row, 9).value) or None,
            "raw_notes": note,
            "handler_raw": _normalize(worksheet.cell(row, 11).value) or None,
            "filled_at_raw": _normalize(worksheet.cell(row, 12).value) or None,
            "preferred_worker_alias": preferred_worker_alias,
            "preference_strength": preference_strength,
            "status": "requested",
        }
        histogram[f"{service_date}|{period}"] += 1
        records.append(ParsedRecord(
            record=record,
            source=_source(workbook_path, worksheet, row, 4,
                           worksheet.cell(row, 4).value),
            raw={str(i): worksheet.cell(row, i).value for i in range(1, 13)},
            parse_confidence="high" if appointment_time else "medium",
        ))

    changes = _parse_change_section(worksheet, workbook_path)
    records.extend(changes[0])
    ambiguities.extend(changes[1])

    time_rate = (time_parsed / time_seen) if time_seen else 1.0
    summary = ImportBatchSummary(
        parser_name="escort_workbook",
        status="ok",
        parsed_count=sum(1 for r in records
                         if r.record and r.record.get("status") == "requested"),
        inferred_count=preference_count,
        flagged_count=len(ambiguities),
        silently_dropped_cells=0,
        notes=(
            f"half-day demand histogram values: {sorted(set(histogram.values()))}",
            f"time parse rate: {time_rate:.3f}",
        ),
        doc_ref=DOC_REF,
    )
    return ImportResult(
        summary=summary,
        records=tuple(records),
        ambiguities=tuple(ambiguities),
        source_workbook=workbook_path,
    )


def _parse_change_section(
    worksheet: Worksheet,
    workbook_path: Path | None,
) -> tuple[list[ParsedRecord[dict[str, Any]]], list[ImportAmbiguity]]:
    records: list[ParsedRecord[dict[str, Any]]] = []
    ambiguities: list[ImportAmbiguity] = []
    for row in range(151, worksheet.max_row + 1):
        values = [_normalize(worksheet.cell(row, col).value) for col in range(3, 12)]
        if not any(values):
            continue
        if values[0].lower() == "e.g.":
            continue
        record = {
            "source": "escort_change_section",
            "source_ref": f"{worksheet.title}!C{row}:L{row}",
            "row": row,
            "service_date_raw": values[0] or None,
            "elder_alias": values[1] or None,
            "unit": values[2] or None,
            "appointment_time_raw": values[3] or None,
            "destination": values[4] or None,
            "subject": values[5] or None,
            "transport": values[6] or None,
            "change_raw": values[7] or None,
            "handler_raw": values[8] or None,
            "status": "change_or_cancellation",
        }
        records.append(ParsedRecord(
            record=record,
            source=_source(workbook_path, worksheet, row, 3,
                           worksheet.cell(row, 3).value),
            raw={str(i): worksheet.cell(row, i).value for i in range(3, 12)},
            parse_confidence="medium",
        ))
        ambiguities.append(ImportAmbiguity(
            code="ESCORT_CHANGE_SECTION",
            message="bottom cancellation/change section row requires human review",
            source=_source(workbook_path, worksheet, row, 3,
                           worksheet.cell(row, 3).value),
            severity="warning",
            raw_value=" | ".join(v for v in values if v),
        ))
    return records, ambiguities
