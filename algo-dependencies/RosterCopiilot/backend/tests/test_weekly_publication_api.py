"""Phase 1B package D: separate fail-closed final publication."""
from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from copy import deepcopy
from datetime import date
import hashlib
from io import BytesIO
from pathlib import Path
import subprocess
from urllib.parse import unquote

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest

from app.api import demo as demo_api
from app.domain import AuditItem, AuditKind, Severity
from app.exporter import prepare_generated_division_roster_export
from app.main import app
from app.scheduler import GeneratedDemands, run_scheduler
from app.services import state as state_service
from app.services.state import AppState
from app.services.weekly_demo import WeeklyRosterDemoBuilder

from fixtures.paths import ESCORT_WORKBOOK_PATH, HC_TIMETABLE_WORKBOOK_PATH


@pytest.fixture(scope="module")
def publication_demo_run():
    build = WeeklyRosterDemoBuilder().build(
        hc_workbook_path=HC_TIMETABLE_WORKBOOK_PATH,
        escort_workbook_path=ESCORT_WORKBOOK_PATH,
        week_start=date(2026, 1, 5),
        changes_json="[]",
    )
    result = run_scheduler(build.snapshot)
    plan = prepare_generated_division_roster_export(
        division_layout=build.division,
        dataset=result.dataset,
        version=result.version,
        generated=result.generated,
    )
    return demo_api.DemoRun(
        run_id="phase1b-publication-api",
        build=build,
        result=result,
        review_version=plan.review_version,
        export_report=plan.report,
        export_plan=plan,
        upload_names={"hc_workbook": "hc.xlsx", "escort_workbook": "escort.xlsx"},
        master_data_version=10,
    )


def _controlled_run(source, *, state: str, run_id: str):
    """Use the real preflight on a minimal, explicitly demand-free test week."""

    run = deepcopy(source)
    run.run_id = run_id
    generated = GeneratedDemands(week_start=run.build.snapshot.week_start)
    current = run.review_version.model_copy(deep=True, update={
        "entries": [],
        "unassigned": [],
        "audit_items": [],
        "demand_dispositions": [],
        "reconciliation": None,
        "trigger_events": [],
    })
    if state == "draft":
        current.audit_items = [AuditItem(
            id="test-draft-audit",
            kind=AuditKind.TEMPLATE_ISSUE,
            severity=Severity.WARNING,
            blocking=False,
            reason="受控測試：仍待非阻塞主管審核",
        )]
    plan = prepare_generated_division_roster_export(
        division_layout=run.build.division,
        dataset=run.result.dataset,
        version=current,
        generated=generated,
    )
    assert plan.report.publication_state == state
    run.result.generated = generated
    run.review_version = plan.review_version
    run.export_report = plan.report
    run.export_plan = plan
    return run


def _install_run(tmp_path: Path, run, monkeypatch):
    monkeypatch.setenv("ROSTER_EXPORT_DIR", str(tmp_path / "exports"))
    state_service._STATE = AppState(db_path=tmp_path / "publication.db")
    record = state_service.get_state().store.create_weekly_run(
        demo_api._weekly_run_record(run)
    )
    return TestClient(app), record


def _publish_payload(record, *, actor="主管甲"):
    return {
        "actor": actor,
        "source_version_id": record.current_version_id,
        "content_hash": record.latest_content_hash,
    }


@pytest.mark.parametrize("state", ["blocked", "draft"])
def test_d01_non_ready_never_writes_final_or_record(
    tmp_path,
    monkeypatch,
    publication_demo_run,
    state,
):
    previous = state_service.get_state()
    run = (
        deepcopy(publication_demo_run)
        if state == "blocked"
        else _controlled_run(publication_demo_run, state="draft", run_id="draft-run")
    )
    run.run_id = f"d01-{state}"
    try:
        client, record = _install_run(tmp_path, run, monkeypatch)
        response = client.post(
            f"/api/demo/weekly-roster/{record.run_id}/publish",
            json=_publish_payload(record),
        )
        assert response.status_code == 409, response.text
        detail = response.json()["detail"]
        assert detail["code"] == "PUBLICATION_NOT_READY"
        assert detail["publication_state"] == state
        assert detail["reasons"]
        stored = state_service.get_state().store.get_weekly_run(record.run_id)
        assert stored is not None and stored.publications == []
        assert not list((tmp_path / "exports").rglob("*.xlsx"))
    finally:
        state_service._STATE = previous


def test_d02_ready_publishes_openable_staff_final(
    tmp_path,
    monkeypatch,
    publication_demo_run,
):
    previous = state_service.get_state()
    run = _controlled_run(publication_demo_run, state="ready", run_id="d02-ready")
    try:
        client, record = _install_run(tmp_path, run, monkeypatch)
        response = client.post(
            f"/api/demo/weekly-roster/{record.run_id}/publish",
            json=_publish_payload(record, actor="  主管甲  "),
        )
        assert response.status_code == 200, response.text
        body = response.json()
        publication = body["publication"]
        assert publication["filename"] == "照顧員工作分工表_正式版.xlsx"
        assert publication["actor"] == "主管甲"
        assert publication["source_version_id"] == record.current_version_id
        assert publication["content_hash"] == record.latest_content_hash
        assert len(publication["artifact_sha256"]) == 64
        assert "artifact_path" not in publication
        assert all("artifact_path" not in item for item in body["publications"])

        stored = state_service.get_state().store.get_weekly_run(record.run_id)
        assert stored is not None and len(stored.publications) == 1
        artifact = Path(stored.publications[0].artifact_path)
        assert artifact.is_file()
        assert hashlib.sha256(artifact.read_bytes()).hexdigest() == publication[
            "artifact_sha256"
        ]
        load_workbook(artifact).close()

        download = client.get(body["final_export_url"])
        assert download.status_code == 200
        assert "正式版" in unquote(download.headers["content-disposition"])
        load_workbook(BytesIO(download.content)).close()
    finally:
        state_service._STATE = previous


def test_d03_stale_version_hash_and_invalid_body_are_rejected(
    tmp_path,
    monkeypatch,
    publication_demo_run,
):
    previous = state_service.get_state()
    run = _controlled_run(publication_demo_run, state="ready", run_id="d03-ready")
    try:
        client, record = _install_run(tmp_path, run, monkeypatch)
        for patch, expected in (
            ({"source_version_id": "stale-version"}, "STALE_SCHEDULE_VERSION"),
            ({"content_hash": "stale-hash"}, "STALE_CONTENT_HASH"),
        ):
            payload = {**_publish_payload(record), **patch}
            response = client.post(
                f"/api/demo/weekly-roster/{record.run_id}/publish", json=payload
            )
            assert response.status_code == 409
            assert response.json()["detail"]["code"] == expected
        invalid = client.post(
            f"/api/demo/weekly-roster/{record.run_id}/publish",
            json={**_publish_payload(record, actor=" "), "unexpected": True},
        )
        assert invalid.status_code == 422
        missing = client.post(
            "/api/demo/weekly-roster/not-a-run/publish",
            json=_publish_payload(record),
        )
        assert missing.status_code == 404
        assert not list((tmp_path / "exports").rglob("*.xlsx"))
    finally:
        state_service._STATE = previous


def test_d04_exact_retry_and_concurrency_keep_one_restart_safe_artifact(
    tmp_path,
    monkeypatch,
    publication_demo_run,
):
    previous = state_service.get_state()
    run = _controlled_run(publication_demo_run, state="ready", run_id="d04-ready")
    try:
        client, record = _install_run(tmp_path, run, monkeypatch)

        def publish(actor):
            return TestClient(app).post(
                f"/api/demo/weekly-roster/{record.run_id}/publish",
                json=_publish_payload(record, actor=actor),
            )

        with ThreadPoolExecutor(max_workers=2) as pool:
            responses = list(pool.map(publish, ["主管甲", "主管乙"]))
        assert [item.status_code for item in responses] == [200, 200]
        bodies = [item.json() for item in responses]
        assert len({item["publication"]["publication_id"] for item in bodies}) == 1
        assert sorted(item["idempotent_replay"] for item in bodies) == [False, True]

        db_path = state_service.get_state().store.db_path
        state_service._STATE = AppState(db_path=db_path, load_existing=True)
        restarted = client.get(f"/api/demo/weekly-roster/{record.run_id}")
        assert restarted.status_code == 200, restarted.text
        assert len(restarted.json()["publications"]) == 1
        assert restarted.json()["final_export_url"]
        stored = state_service.get_state().store.get_weekly_run(record.run_id)
        assert stored is not None and Path(stored.publications[0].artifact_path).is_file()
    finally:
        state_service._STATE = previous


def test_d04_persistence_failure_removes_uncommitted_artifact(
    tmp_path,
    monkeypatch,
    publication_demo_run,
):
    previous = state_service.get_state()
    run = _controlled_run(
        publication_demo_run, state="ready", run_id="d04-persist-failure"
    )
    try:
        client, record = _install_run(tmp_path, run, monkeypatch)
        store = state_service.get_state().store
        original = store.save_weekly_run_publication

        def fail_persistence(_publication):
            raise ValueError("deterministic persistence failure")

        monkeypatch.setattr(store, "save_weekly_run_publication", fail_persistence)
        response = client.post(
            f"/api/demo/weekly-roster/{record.run_id}/publish",
            json=_publish_payload(record),
        )
        assert response.status_code == 409
        assert not list((tmp_path / "exports").rglob("*.xlsx"))
        monkeypatch.setattr(store, "save_weekly_run_publication", original)
        stored = store.get_weekly_run(record.run_id)
        assert stored is not None and stored.publications == []
    finally:
        state_service._STATE = previous


@pytest.mark.parametrize("damage", ["tampered", "missing"])
def test_d04_tampered_or_missing_artifact_fails_closed_after_restart(
    tmp_path,
    monkeypatch,
    publication_demo_run,
    damage,
):
    previous = state_service.get_state()
    run = _controlled_run(
        publication_demo_run, state="ready", run_id=f"d04-{damage}"
    )
    try:
        client, record = _install_run(tmp_path, run, monkeypatch)
        published = client.post(
            f"/api/demo/weekly-roster/{record.run_id}/publish",
            json=_publish_payload(record),
        )
        assert published.status_code == 200
        stored = state_service.get_state().store.get_weekly_run(record.run_id)
        assert stored is not None
        artifact = Path(stored.publications[0].artifact_path)
        if damage == "tampered":
            artifact.write_bytes(b"tampered")
        else:
            artifact.unlink()

        db_path = state_service.get_state().store.db_path
        state_service._STATE = AppState(db_path=db_path, load_existing=True)
        restored = client.get(f"/api/demo/weekly-roster/{record.run_id}")
        assert restored.status_code == 409
        assert restored.json()["detail"]["code"] in {
            "PUBLICATION_ARTIFACT_CORRUPT",
            "PUBLICATION_ARTIFACT_MISSING",
        }
        download = client.get(published.json()["final_export_url"])
        assert download.status_code == 409
    finally:
        state_service._STATE = previous


def test_d05_review_export_and_frontend_server_state_contract(
    tmp_path,
    monkeypatch,
    publication_demo_run,
):
    previous = state_service.get_state()
    try:
        client, record = _install_run(tmp_path, deepcopy(publication_demo_run), monkeypatch)
        review = client.get(f"/api/demo/weekly-roster/{record.run_id}/export")
        assert review.status_code == 200
        disposition = unquote(review.headers["content-disposition"])
        assert "審核草稿" in disposition
        assert "正式版" not in disposition

        html = (Path(__file__).resolve().parents[2] / "frontend" / "index.html").read_text(
            encoding="utf-8"
        )
        for token in (
            "下載審核草稿",
            "發佈正式版",
            "下載正式版",
            "/publish",
            "run.reconciliation?.publication_state",
            'publicationState === "ready"',
        ):
            assert token in html
        assert "review_export_allowed ===" not in html
        script = html.rsplit("<script>", 1)[1].split("</script>", 1)[0]
        checked = subprocess.run(
            ["node", "--check"],
            input=script,
            text=True,
            capture_output=True,
            check=False,
        )
        assert checked.returncode == 0, checked.stderr
    finally:
        state_service._STATE = previous
