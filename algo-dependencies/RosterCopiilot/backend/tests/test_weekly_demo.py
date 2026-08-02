"""User-facing weekly roster demo flow tests."""
from __future__ import annotations

from copy import deepcopy
from datetime import date
from io import BytesIO
from urllib.parse import unquote

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest

from app.api import demo as demo_api
from app.domain import canonical_json
from app.exporter import prepare_generated_division_roster_export
from app.main import app
from app.scheduler import run_scheduler, version_content_hash
from app.services import state as state_service
from app.services.weekly_demo import WeeklyRosterDemoBuilder

from fixtures.paths import (
    DIVISION_WORKBOOK_PATH,
    ESCORT_WORKBOOK_PATH,
    HC_TIMETABLE_WORKBOOK_PATH,
)


@pytest.fixture(scope="module")
def provenance_demo_run():
    build = WeeklyRosterDemoBuilder().build(
        hc_workbook_path=HC_TIMETABLE_WORKBOOK_PATH,
        escort_workbook_path=ESCORT_WORKBOOK_PATH,
        week_start=date(2026, 1, 5),
        changes_json="[]",
    )
    result = run_scheduler(build.snapshot)
    plan = prepare_generated_division_roster_export(
        division_layout=build.division,
        dataset=result.dataset,
        version=result.version,
        generated=result.generated,
    )
    return demo_api.DemoRun(
        run_id="registry-negative",
        build=build,
        result=result,
        review_version=plan.review_version,
        export_report=plan.report,
        export_plan=plan,
        upload_names={"hc_workbook": "hc.xlsx", "escort_workbook": "escort.xlsx"},
    )


def test_weekly_demo_builder_uses_internal_division_and_uploaded_escort_week():
    build = WeeklyRosterDemoBuilder().build(
        hc_workbook_path=HC_TIMETABLE_WORKBOOK_PATH,
        escort_workbook_path=ESCORT_WORKBOOK_PATH,
        week_start=date(2026, 1, 5),
        changes_json="[]",
    )

    assert len(build.division.workers) == 46
    assert build.source_counts["division_fixed"] >= 300
    assert build.source_counts["escort_selected"] == 21
    assert build.source_counts.get("hc_selected", 0) == 0
    assert build.warnings

    result = run_scheduler(build.snapshot)
    assert result.version.entries
    assert result.generated.counts_by_kind["escort"] == 21
    assert not result.violations


def test_weekly_demo_builder_filters_hc_target_week_and_records_changes():
    changes = [{
        "type": "leave",
        "change_date": "2026-05-04",
        "period": "AM",
        "worker_alias": "輝",
        "reason": "測試請假",
    }]
    build = WeeklyRosterDemoBuilder().build(
        hc_workbook_path=HC_TIMETABLE_WORKBOOK_PATH,
        escort_workbook_path=ESCORT_WORKBOOK_PATH,
        week_start=date(2026, 5, 4),
        changes_json={"changes": changes},
    )

    assert build.source_counts["hc_selected"] > 0
    assert build.source_counts.get("escort_selected", 0) == 0
    assert build.source_counts["temporary_changes"] == 1
    assert build.snapshot.change_events[0].worker_id
    assert build.snapshot.availability[0].is_available is False


def test_weekly_demo_builder_aligns_non_monday_week_start():
    build = WeeklyRosterDemoBuilder().build(
        hc_workbook_path=HC_TIMETABLE_WORKBOOK_PATH,
        escort_workbook_path=ESCORT_WORKBOOK_PATH,
        week_start=date(2026, 1, 7),  # Wednesday
        changes_json="[]",
    )

    assert build.snapshot.week_start == date(2026, 1, 5)
    assert any("自動對齊" in warning for warning in build.warnings)
    assert build.source_counts["escort_selected"] == 21


def test_weekly_demo_uploaded_week_selection_changes_generated_demand():
    kwargs = dict(
        hc_workbook_path=HC_TIMETABLE_WORKBOOK_PATH,
        escort_workbook_path=ESCORT_WORKBOOK_PATH,
        changes_json="[]",
    )
    escort_week = WeeklyRosterDemoBuilder().build(week_start=date(2026, 1, 5), **kwargs)
    empty_week = WeeklyRosterDemoBuilder().build(week_start=date(2026, 3, 2), **kwargs)

    assert escort_week.source_counts["escort_selected"] == 21
    assert empty_week.source_counts.get("escort_selected", 0) == 0

    escort_result = run_scheduler(escort_week.snapshot)
    empty_result = run_scheduler(empty_week.snapshot)
    assert escort_result.generated.counts_by_kind.get("escort", 0) == 21
    assert empty_result.generated.counts_by_kind.get("escort", 0) == 0


def test_weekly_demo_leave_change_removes_worker_from_that_day():
    week_start = date(2026, 1, 5)
    builder = WeeklyRosterDemoBuilder()
    kwargs = dict(
        hc_workbook_path=HC_TIMETABLE_WORKBOOK_PATH,
        escort_workbook_path=ESCORT_WORKBOOK_PATH,
        week_start=week_start,
    )

    baseline = run_scheduler(builder.build(changes_json="[]", **kwargs).snapshot)
    scheduled_monday = [
        entry for entry in baseline.version.entries
        if entry.worker_id and entry.schedule_date == week_start
        and entry.status.value in ("scheduled", "needs_review")
    ]
    assert scheduled_monday
    target_worker = scheduled_monday[0].worker_id

    changed_build = builder.build(
        changes_json=[{
            "type": "leave",
            "change_date": week_start.isoformat(),
            "worker_id": target_worker,
            "reason": "測試全日請假",
        }],
        **kwargs,
    )
    changed = run_scheduler(changed_build.snapshot)

    assert changed_build.change_summaries
    assert changed.reports, "leave change must produce an impact report"
    still_scheduled = [
        entry for entry in changed.version.entries
        if entry.worker_id == target_worker and entry.schedule_date == week_start
        and entry.status.value in ("scheduled", "needs_review")
    ]
    assert not still_scheduled
    assert not changed.violations


def test_weekly_demo_new_escort_change_adds_demand_and_audit_material():
    week_start = date(2026, 1, 5)
    builder = WeeklyRosterDemoBuilder()
    kwargs = dict(
        hc_workbook_path=HC_TIMETABLE_WORKBOOK_PATH,
        escort_workbook_path=ESCORT_WORKBOOK_PATH,
        week_start=week_start,
    )

    baseline = run_scheduler(builder.build(changes_json="[]", **kwargs).snapshot)
    changed = run_scheduler(builder.build(
        changes_json=[{
            "type": "escort_new",
            "change_date": "2026-01-07",
            "period": "AM",
            "elder_alias": "測試長者",
            "destination": "東區醫院",
            "reason": "臨時新增覆診",
        }],
        **kwargs,
    ).snapshot)

    assert len(changed.version.entries) == len(baseline.version.entries) + 1
    assert changed.reports
    assert len(changed.version.audit_items) > len(baseline.version.audit_items)


def test_weekly_demo_api_generates_and_exports_ngo_workbook(tmp_path, monkeypatch):
    monkeypatch.setenv("ROSTER_EXPORT_DIR", str(tmp_path / "exports"))
    client = TestClient(app)

    with HC_TIMETABLE_WORKBOOK_PATH.open("rb") as hc, ESCORT_WORKBOOK_PATH.open("rb") as escort:
        response = client.post(
            "/api/demo/weekly-roster",
            data={
                "week_start": "2026-01-05",
                "changes_json": "[]",
            },
            files={
                "hc_workbook": (
                    HC_TIMETABLE_WORKBOOK_PATH.name,
                    hc,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                "escort_workbook": (
                    ESCORT_WORKBOOK_PATH.name,
                    escort,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["fixed_base"]["worker_columns"] == 46
    assert body["parse_summary"]["escort_selected_for_week"] == 21
    assert body["generation_summary"]["hard_constraint_violations"] == 0
    assert body["export_url"].startswith("/api/demo/weekly-roster/")
    assert body["review_export_allowed"] is True
    assert "unassigned_items" in body
    assert len(body["unassigned_items"]) == body["generation_summary"]["unassigned"]
    assert "impact_reports" in body
    assert "change_summary" in body
    assert body["reconciliation"] == body["version"]["reconciliation"]
    assert body["reconciliation"] == body["export_report"]["reconciliation"]
    assert body["audit_items"] == body["version"]["audit_items"]
    assert body["publication_state"] == body["reconciliation"]["publication_state"]
    assert body["generation_summary"]["needs_review"] == body["reconciliation"]["needs_review"]
    assert body["generation_summary"]["unassigned"] == body["reconciliation"]["unassigned"]
    assert body["generation_summary"]["export_failures"] == body["reconciliation"]["export_failure_count"]
    _assert_api_provenance_resolves(body)

    # GET must reuse the exact POST-prepared plan.  A fresh preflight here
    # would call this patched function and fail the request.
    monkeypatch.setattr(
        "app.exporter.division_writer.prepare_generated_division_roster_export",
        lambda **kwargs: (_ for _ in ()).throw(
            AssertionError("download attempted a fresh export preflight")
        ),
    )

    export_response = client.get(body["export_url"])
    assert export_response.status_code == 200, export_response.text
    assert "審核草稿" in unquote(export_response.headers["content-disposition"])
    workbook = load_workbook(BytesIO(export_response.content))
    assert {"恆常服務", "RC_變更摘要", "RC_審核", "RC_未分配", "RC_meta"} <= set(workbook.sheetnames)
    expected_reconciliation = canonical_json(body["reconciliation"])
    for sheet_name in ("RC_變更摘要", "RC_審核", "RC_未分配", "RC_meta"):
        assert _embedded_reconciliation(workbook[sheet_name]) == expected_reconciliation
    meta = _sheet_key_values(workbook["RC_meta"])
    assert meta["reconciliation.version_id"] == body["reconciliation"]["version_id"]
    assert meta["reconciliation.content_hash"] == body["reconciliation"]["content_hash"]

    template = load_workbook(DIVISION_WORKBOOK_PATH)
    exported = workbook["恆常服務"]
    original = template["恆常服務"]
    assert len(exported.merged_cells.ranges) == len(original.merged_cells.ranges)

    generated_cells = []
    for row in exported.iter_rows(min_row=4, max_row=90, min_col=3, max_col=49):
        for cell in row:
            if cell.comment and "RosterCopiilot" in cell.comment.text:
                generated_cells.append(cell)
    assert generated_cells


def test_weekly_demo_api_uses_persisted_master_data(tmp_path):
    """A master-data override must change the user-facing demo result."""
    previous_state = state_service.get_state()
    state_service.reset_state(db_path=tmp_path / "roster.db")
    client = TestClient(app)

    def build_run():
        with HC_TIMETABLE_WORKBOOK_PATH.open("rb") as hc, ESCORT_WORKBOOK_PATH.open("rb") as escort:
            response = client.post(
                "/api/demo/weekly-roster",
                data={"week_start": "2026-01-05", "changes_json": "[]"},
                files={
                    "hc_workbook": (
                        HC_TIMETABLE_WORKBOOK_PATH.name,
                        hc,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                    "escort_workbook": (
                        ESCORT_WORKBOOK_PATH.name,
                        escort,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                },
            )
        assert response.status_code == 200, response.text
        return response.json()

    try:
        baseline = build_run()
        target = next(
            entry for entry in baseline["version"]["entries"]
            if entry["schedule_date"] == "2026-01-05"
            and entry["period"] == "AM"
            and entry["worker_id"]
            and entry["status"] in {"scheduled", "needs_review"}
        )

        override = client.post(
            "/api/master-data/manual-overrides",
            json={
                "id": "test-forbid-monday-am",
                "scope": "week",
                "pin": {
                    "worker_id": target["worker_id"],
                    "weekday": 1,
                    "period": "AM",
                },
                "action": "forbid_assignment",
                "reason": "API integration regression test",
                "effective_from": "2026-01-05",
                "effective_to": "2026-01-10",
            },
        )
        assert override.status_code == 201, override.text
        assert override.json()["version"] == 2

        updated = build_run()
        blocked_assignments = [
            entry for entry in updated["version"]["entries"]
            if entry["schedule_date"] == "2026-01-05"
            and entry["period"] == "AM"
            and entry["worker_id"] == target["worker_id"]
            and entry["status"] in {"scheduled", "needs_review"}
        ]
        assert not blocked_assignments
        assert updated["review_export_allowed"] is True
    finally:
        state_service._STATE = previous_state


def test_api_registry_does_not_self_register_missing_producer_evidence(
    provenance_demo_run,
):
    run = deepcopy(provenance_demo_run)
    missing_id = next(
        ref
        for disposition in run.export_report.reconciliation.dispositions
        for ref in disposition.source_ref_ids
    )
    run.result.generated.source_evidence = [
        item for item in run.result.generated.source_evidence if item.id != missing_id
    ]

    with pytest.raises(RuntimeError, match=missing_id):
        demo_api._provenance_registries(run)


def test_api_registry_rejects_same_evidence_id_with_different_payload(
    provenance_demo_run,
):
    run = deepcopy(provenance_demo_run)
    entry = next(
        item for item in run.review_version.entries
        if any(evidence.confidence == "high" for evidence in item.source_evidence)
    )
    index = next(
        index for index, evidence in enumerate(entry.source_evidence)
        if evidence.confidence == "high"
    )
    tampered = entry.source_evidence[index].model_copy(deep=True)
    tampered.confidence = "seed"
    entry.source_evidence[index] = tampered
    recomputed = version_content_hash(run.review_version)
    run.review_version.reconciliation.content_hash = recomputed
    run.export_report.reconciliation.content_hash = recomputed

    with pytest.raises(RuntimeError, match=f"payload differs.*{tampered.id}"):
        demo_api._provenance_registries(run)


def test_api_registry_rejects_duplicate_root_evidence_ids(provenance_demo_run):
    run = deepcopy(provenance_demo_run)
    duplicate = run.result.generated.source_evidence[0].model_copy(deep=True)
    run.result.generated.source_evidence.append(duplicate)

    with pytest.raises(RuntimeError, match=f"duplicate source evidence.*{duplicate.id}"):
        demo_api._provenance_registries(run)


def test_api_registry_rejects_demand_gap_payload_drift(provenance_demo_run):
    run = deepcopy(provenance_demo_run)
    demand = next(
        item for item in run.result.generated.weekly_demands if item.data_gaps
    )
    tampered = demand.data_gaps[0].model_copy(deep=True)
    tampered.message += " tampered"
    demand.data_gaps[0] = tampered

    with pytest.raises(RuntimeError, match=f"payload differs.*{tampered.id}"):
        demo_api._provenance_registries(run)


def test_api_registry_rejects_unknown_demand_gap_id(provenance_demo_run):
    run = deepcopy(provenance_demo_run)
    demand = run.result.generated.weekly_demands[0]
    demand.data_gap_ids.append("gap_00000000000000000000")

    with pytest.raises(RuntimeError, match=f"{demand.demand_id} data-gap IDs drifted"):
        demo_api._provenance_registries(run)


def test_api_registry_rejects_invalid_primary_evidence(provenance_demo_run):
    run = deepcopy(provenance_demo_run)
    demand = run.result.generated.weekly_demands[0]
    demand.primary_source_evidence_id = None

    with pytest.raises(RuntimeError, match=f"{demand.demand_id} has invalid primary evidence"):
        demo_api._provenance_registries(run)


def test_api_registry_rejects_unknown_embedded_entry_gap(provenance_demo_run):
    run = deepcopy(provenance_demo_run)
    embedded = next(
        entry
        for audit in run.review_version.audit_items
        for entry in [audit.original_entry, audit.suggested_entry, *audit.alternatives]
        if entry is not None
    )
    embedded.data_gap_ids.append("gap_00000000000000000000")
    recomputed = version_content_hash(run.review_version)
    run.review_version.reconciliation.content_hash = recomputed
    run.export_report.reconciliation.content_hash = recomputed

    with pytest.raises(RuntimeError, match=f"entry {embedded.id} data gaps"):
        demo_api._provenance_registries(run)


def test_api_payload_rejects_outer_publication_state_drift(provenance_demo_run):
    run = deepcopy(provenance_demo_run)
    run.export_report = run.export_report.model_copy(deep=True)
    run.export_report.publication_state = "draft"

    with pytest.raises(RuntimeError, match="export report conflicts with export plan"):
        demo_api._response_payload(run)


def test_api_payload_rejects_outer_review_export_allowed_drift(
    provenance_demo_run,
):
    run = deepcopy(provenance_demo_run)
    run.export_report = run.export_report.model_copy(deep=True)
    run.export_report.review_export_allowed = False

    with pytest.raises(RuntimeError, match="export report conflicts with export plan"):
        demo_api._response_payload(run)


def test_api_payload_rejects_plan_placement_cell_drift(provenance_demo_run):
    run = deepcopy(provenance_demo_run)
    run.export_plan.report.placements[0].assignment_cell = "恆常服務!A1"

    with pytest.raises(RuntimeError, match="weekly demo export plan is invalid"):
        demo_api._response_payload(run)


def _assert_api_provenance_resolves(body: dict) -> None:
    evidence_ids = {item["id"] for item in body["source_evidence"]}
    gap_ids = {item["id"] for item in body["data_gaps"]}
    audit_ids = {item["id"] for item in body["audit_items"]}
    disposition_by_demand = {
        item["demand_id"]: item for item in body["demand_dispositions"]
    }
    version_entry_ids = {item["id"] for item in body["version"]["entries"]}
    embedded_entry_ids = {
        entry["id"]
        for audit in body["audit_items"]
        for entry in [audit.get("original_entry"), audit.get("suggested_entry"),
                      *audit.get("alternatives", [])]
        if entry is not None
    }
    embedded_entry_ids |= {
        entry["id"]
        for audit in body["audit_items"]
        for step in audit.get("chain", [])
        for entry in [step.get("entry_before"), step.get("entry_after")]
        if entry is not None
    }
    entry_ids = version_entry_ids | embedded_entry_ids

    assert body["source_evidence"] == sorted(
        body["source_evidence"], key=lambda item: item["id"]
    )
    assert body["data_gaps"] == sorted(body["data_gaps"], key=lambda item: item["id"])
    for gap in body["data_gaps"]:
        assert set(gap["source_ref_ids"]) <= evidence_ids
    for disposition in body["demand_dispositions"]:
        assert set(disposition["source_ref_ids"]) <= evidence_ids
        assert set(disposition["audit_ids"]) <= audit_ids
        if disposition["entry_id"]:
            assert disposition["entry_id"] in version_entry_ids
    for entry in body["version"]["entries"]:
        assert entry["demand_id"] in disposition_by_demand
        assert {item["id"] for item in entry["source_evidence"]} <= evidence_ids
        assert set(entry["data_gap_ids"]) <= gap_ids
        assert set(entry["audit_ids"]) <= audit_ids
    for audit in body["audit_items"]:
        assert set(audit["demand_ids"]) <= set(disposition_by_demand)
        assert set(audit["entry_ids"]) <= entry_ids
        assert set(audit["data_gap_ids"]) <= gap_ids
        assert set(audit["evidence_refs"]) <= evidence_ids
    for placement in body["export_report"]["placements"]:
        assert placement["version_id"] == body["reconciliation"]["version_id"]
        assert placement["entry_id"] in version_entry_ids
        assert placement["demand_id"] in disposition_by_demand
        assert placement["disposition"] == disposition_by_demand[
            placement["demand_id"]
        ]["disposition"]
        assert set(placement["audit_ids"]) <= audit_ids
        assert set(placement["data_gap_ids"]) <= gap_ids
        assert set(placement["source_evidence_ids"]) <= evidence_ids
        assert placement["assignment_cell"]


def _embedded_reconciliation(ws) -> str:
    manifest_col = next(
        cell.column for cell in ws[1]
        if cell.value == "RC_reconciliation_key"
    )
    parts = []
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, manifest_col).value
        value = ws.cell(row, manifest_col + 1).value
        if isinstance(key, str) and key.startswith("reconciliation_json_"):
            parts.append((key, value or ""))
    return "".join(value for _, value in sorted(parts))


def _sheet_key_values(ws) -> dict[str, object]:
    return {
        str(ws.cell(row, 1).value): ws.cell(row, 2).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 1).value is not None
    }
