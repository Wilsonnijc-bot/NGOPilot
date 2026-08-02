"""Export bridge tests: scheduler entries → division-style assignment grid.

Covers ENGINEERING_SPEC.md §10.5 (division workbook export from scheduled
entries) via the template-free path the scheduler fixture needs.
"""
import openpyxl

from app.domain import EntryStatus
from app.exporter import build_assignment_grid_workbook, entry_label, grid_labels
from app.scheduler import representative_snapshot, run_scheduler


def test_grid_labels_place_entries_by_worker_and_slot():
    result = run_scheduler(representative_snapshot())
    labels = grid_labels(result.version, workers=result.snapshot.workers)
    labels.pop("__names__")
    # every worker column exists even if idle
    assert {w.id for w in result.snapshot.workers} <= set(labels)
    # a placed entry lands in its (weekday, period) cell
    placed = [e for e in result.version.entries
              if e.status in (EntryStatus.SCHEDULED, EntryStatus.NEEDS_REVIEW)
              and e.worker_id]
    sample = placed[0]
    slot = labels[sample.worker_id][(sample.weekday, sample.period.value)]
    assert any(sample.service_code.value in label for label in slot)


def test_grid_workbook_is_openable_and_covers_workers(tmp_path):
    result = run_scheduler(representative_snapshot())
    wb = build_assignment_grid_workbook(result.version, workers=result.snapshot.workers)
    path = tmp_path / "grid.xlsx"
    wb.save(path)

    reloaded = openpyxl.load_workbook(path)
    assert reloaded.sheetnames == ["排班格", "RC_未分配"]
    ws = reloaded["排班格"]
    # header worker column + 6 weekdays × AM/PM
    assert ws.max_column == 1 + 6 * 2
    assert ws.cell(1, 1).value == "同工"
    # one row per worker
    assert ws.max_row == 1 + len(result.snapshot.workers)


def test_unassigned_work_appears_in_its_own_sheet():
    result = run_scheduler(representative_snapshot())
    wb = build_assignment_grid_workbook(result.version, workers=result.snapshot.workers)
    ws = wb["RC_未分配"]
    ids = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    unassigned_ids = {e.id for e in result.version.entries
                      if e.status == EntryStatus.UNASSIGNED}
    assert unassigned_ids and unassigned_ids <= ids


def test_needs_review_cells_are_marked_additively():
    result = run_scheduler(representative_snapshot())
    wb = build_assignment_grid_workbook(result.version, workers=result.snapshot.workers)
    ws = wb["排班格"]
    marked = [ws.cell(r, c) for r in range(2, ws.max_row + 1)
              for c in range(2, ws.max_column + 1)
              if ws.cell(r, c).comment is not None]
    review_entries = [e for e in result.version.entries
                      if e.status == EntryStatus.NEEDS_REVIEW and e.worker_id]
    # a review marker exists iff there is review-required placed work
    assert bool(marked) == bool(review_entries)


def test_entry_label_uses_ngo_cell_grammar():
    result = run_scheduler(representative_snapshot())
    duty = next(e for e in result.version.entries if e.center == "AMC"
                and e.status == EntryStatus.SCHEDULED)
    assert entry_label(duty).startswith("AMC")
