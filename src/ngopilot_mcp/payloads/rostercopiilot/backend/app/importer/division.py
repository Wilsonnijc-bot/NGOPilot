"""Importer for the care-worker work division workbook (照顧員工作分工表).

Scope of this module: the ``恆常服務`` (regular services) sheet — the master
weekly template. Layout and grammar are documented in
docs/spec/excel_semantics.md §1.1, field vocabulary in
docs/spec/data_dictionary.md §1/§3, corrections in
docs/records/fact_check_report_2026-07-01.md (E1: 46 worker columns; E4:
Saturday A/B row; E5: stacked alternating-week cases share one slot).

Parsing philosophy (excel_io_contract.md §6):
- every non-empty cell is classified exactly once (reconciliation, not trust);
- anything the grammar cannot parse becomes an ImportAmbiguity, never a drop;
- inferred facts (gray = departed, trailing name = case manager) are labelled
  inferred/candidate, never asserted as business truth.
"""
from __future__ import annotations

import datetime as _dt
import re
import unicodedata
from dataclasses import replace
from pathlib import Path

from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .division_models import (
    ACTIVE_HEADER_FILLS,
    CounterObservation,
    DEPARTED_HEADER_FILLS,
    DivisionImportResult,
    FixedServiceCandidate,
    ParsedAssignment,
    ParsedDetail,
    PeriodBlock,
    RawScheduleCell,
    WeekdayBlock,
    WorkerColumn,
)
from .models import CellRef, ImportAmbiguity, SourceRef
from .workbook_utils import (
    effective_used_range,
    load_workbook,
    read_fill_color,
    require_sheet,
)

DOC_REF = "docs/spec/excel_semantics.md §1; docs/spec/excel_io_contract.md §1"
REGULAR_SERVICES_SHEET = "恆常服務"
TRANSFER_LOG_PREFIX = "個案轉移紀錄"
SKILLS_SHEET = "新同工跟服務紀錄表"

HEADER_ROW = 2
FIRST_WORKER_COLUMN = 3  # column C
COUNTER_HEADER_MAIN = "個案數量"
COUNTER_HEADER_TOTAL = "總數"

WEEKDAY_CHARS = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6}
PERIOD_CHARS = {"上": "AM", "下": "PM"}

# canonical domain ServiceCode values the importer may assert directly
CANONICAL_CODES = {"E+RO": "E+RO", "HC": "HC", "PC": "PC", "B": "B",
                   "ESC": "ESC", "D": "D", "AMC": "AMC", "MRC": "MRC", "GC": "GC"}
# raw field codes we recognise but do not force into a canonical code
FIELD_CODES = ("E+RO", "PC+E", "HC:+PC", "HC+PC", "Esc", "HC", "PC", "RO", "B", "E")
FIELD_TO_CANONICAL = {"E+RO": "E+RO", "HC": "HC", "PC": "PC", "B": "B", "Esc": "ESC"}
KNOWN_UNITS = {"EH", "IH", "ED", "HSS", "AMC", "MRC", "GC",
               "MRCV", "GCV", "AMCV", "CCSV", "V", "HW", "CC", "B"}

LOGISTICS_PREFIXES = (
    "執牌", "跟車", "中心跟車", "特別探訪", "DECC", "康雲計劃", "康云計劃",
    "電話通知服務更改", "寫飯紙", "清潔及維修飯車", "清潔飯袋", "購買中心生果茶點",
    "午膳",
)
DUTY_ROLE_TOKENS = {
    "主": "lead", "協": "assist", "帶活動": "activity",
    "派藥": "medication", "清潔": "cleaning",
}

_TIME_RANGE_RE = re.compile(
    r"(\d{1,2})[:：](\d{2})\s*-\s*(\d{1,2})[:：](\d{2})")
_BARE_TIME_RE = re.compile(r"^(\d{1,2})[:：](\d{2})(?::\d{2})?$")
_PAREN_RE = re.compile(r"[（(]([^()（）]*)[)）]")
_WEEK_PATTERN_RE = re.compile(
    r"^(?P<weekday>[一二三四五六])?\s*(?P<weeks>\d(?:\s*[,，]\s*\d)*)\s*(?P<tail>長周)?$")
_SATURDAY_TEAM_RE = re.compile(r"^([AB])\s*[:：]?\s*(.*)$")
_FULL_DAY_SPAN_RE = re.compile(r"^(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$")


# ---------------------------------------------------------------------------
# text helpers
# ---------------------------------------------------------------------------

def _normalize(value: object) -> str:
    """Raw cell value -> normalized single-line text (fullwidth folded)."""
    if value is None:
        return ""
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, _dt.datetime):
        return value.isoformat()
    text = str(value)
    text = unicodedata.normalize("NFKC", text)  # fullwidth ASCII -> halfwidth
    text = text.replace("：", ":").replace("，", ",")
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _raw_str(value: object) -> str:
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M:%S")
    return str(value)


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


# ---------------------------------------------------------------------------
# grammar: assignment cells
# ---------------------------------------------------------------------------

def _parse_week_pattern(raw: str, block_weekday: int
                        ) -> tuple[tuple[int, ...] | None, tuple[str, ...]]:
    """Parse a pattern-suffix body like ``1,3`` / ``四1,3`` / ``六1長周``.

    Returns (weeks or None, warnings). Anything with 長周/unknown text keeps
    weeks=None and yields a warning (do not over-interpret; see Q-A4/Q-A6).
    """
    m = _WEEK_PATTERN_RE.match(raw.strip())
    if not m:
        return None, (f"unrecognised week pattern text: {raw!r}",)
    warnings: list[str] = []
    if m.group("weekday"):
        wd = WEEKDAY_CHARS.get(m.group("weekday"))
        if wd != block_weekday:
            warnings.append(
                f"pattern weekday prefix {m.group('weekday')!r} does not match "
                f"the block weekday {block_weekday}")
    if m.group("tail"):  # 長周 — semantics unconfirmed (Q-A4)
        return None, tuple(warnings + [f"unconfirmed pattern token 長周 in {raw!r}"])
    weeks = tuple(sorted(int(x) for x in re.split(r"[,，]", m.group("weeks"))))
    if any(w < 1 or w > 5 for w in weeks):
        return None, tuple(warnings + [f"weeks out of range in {raw!r}"])
    return weeks, tuple(warnings)


def _match_field_code(text: str) -> tuple[str, str] | None:
    """Match a leading field-service code; return (code, remainder)."""
    for code in FIELD_CODES:
        if not text.startswith(code):
            continue
        rest = text[len(code):]
        if code in ("E", "B", "RO", "PC", "HC") and rest[:1] not in (":", "："):
            # single/short codes require an explicit colon to avoid matching
            # arbitrary text ("E+ROF玲" is handled by the E+RO branch below)
            if code == "E+RO":
                pass
            else:
                continue
        return code, rest
    # tolerated no-colon variant seen in the real file: `E+ROF玲(EH)`
    if text.startswith("E+RO"):
        return "E+RO", text[4:]
    return None


def _parse_field_remainder(rest: str, block_weekday: int) -> dict:
    """Parse `` : Elder(UNIT)(pattern) trailing-note`` after a field code."""
    out: dict = {"elder_alias": None, "unit": None, "week_pattern_raw": None,
                 "week_pattern_weeks": None, "inline_note": None,
                 "warnings": []}
    body = rest.lstrip(" :：.").strip()
    if not body:
        out["warnings"].append("incomplete assignment: code without case")
        return out

    parens = [(m.start(), m.group(1).strip()) for m in _PAREN_RE.finditer(body)]
    head_end = parens[0][0] if parens else len(body)
    head = body[:head_end].strip(" .")
    tail = body[_last_paren_end(body):].strip() if parens else ""

    notes: list[str] = []
    for _, content in parens:
        if out["unit"] is None and content in KNOWN_UNITS:
            out["unit"] = content
            continue
        weeks, warns = _parse_week_pattern(content, block_weekday)
        if weeks is not None or (warns and "unconfirmed" in warns[0]):
            out["week_pattern_raw"] = content
            out["week_pattern_weeks"] = weeks
            out["warnings"].extend(warns)
        else:
            notes.append(f"({content})")

    # elder alias may carry a trailing pattern glued without parens; keep raw
    out["elder_alias"] = head or None
    if tail:
        notes.append(tail)
    if notes:
        out["inline_note"] = " ".join(notes)
    if out["elder_alias"] is None:
        out["warnings"].append("no elder alias found")
    return out


def _last_paren_end(body: str) -> int:
    end = 0
    for m in _PAREN_RE.finditer(body):
        end = m.end()
    return end


def _parse_duty(text: str) -> dict | None:
    m = re.match(r"^(AMC|MRC|GC)\b(.*)$", text)
    if not m:
        return None
    center, rest = m.group(1), m.group(2).strip()
    role = None
    notes: list[str] = []
    rest_norm = rest
    if rest_norm.startswith("+CC"):
        notes.append("+CC")
        rest_norm = rest_norm[3:].strip()
    for m2 in _PAREN_RE.finditer(rest_norm):
        token = m2.group(1).strip()
        role = DUTY_ROLE_TOKENS.get(token, role)
        if token not in DUTY_ROLE_TOKENS:
            notes.append(f"({token})")
    bare = _PAREN_RE.sub("", rest_norm).strip()
    if bare:
        if bare in DUTY_ROLE_TOKENS:
            role = DUTY_ROLE_TOKENS[bare]
        else:
            notes.append(bare)  # start-time hints like `9:00` / `1:00`
    return {"duty_center": center, "duty_role": role,
            "inline_note": " ".join(notes) or None}


def _classify_assignment(text: str, block_weekday: int) -> dict | None:
    """Try the assignment grammar; return record fields or None."""
    if text == "OFF":
        return {"kind": "off", "service_code_raw": "OFF", "confidence": "high"}
    if text == "ESC":
        return {"kind": "escort_slot", "service_code_raw": "ESC",
                "service_code": "ESC", "confidence": "high"}
    duty = _parse_duty(text)
    if duty is not None:
        return {"kind": "center_duty", "service_code_raw": duty["duty_center"],
                "service_code": CANONICAL_CODES[duty["duty_center"]],
                "confidence": "high", **duty}
    if text == "D" or re.match(r"^D\s*[:：(]", text):
        place = None
        note = None
        pm = _PAREN_RE.search(text)
        if pm:
            place = pm.group(1).strip()
        cm = re.match(r"^D\s*[:：]\s*(.+)$", text)
        if cm:
            note = cm.group(1).strip()
        return {"kind": "meal", "service_code_raw": "D", "service_code": "D",
                "route_or_place": place, "inline_note": note,
                "confidence": "high"}
    if text.startswith("執牌"):
        pm = _PAREN_RE.search(text)
        return {"kind": "kitchen", "service_code_raw": "執牌",
                "route_or_place": pm.group(1).strip() if pm else None,
                "confidence": "high"}
    for prefix in LOGISTICS_PREFIXES:
        if text.startswith(prefix):
            return {"kind": "logistics", "service_code_raw": prefix,
                    "inline_note": text if text != prefix else None,
                    "confidence": "high"}
    # paren-wrapped variant seen in the real file: `(特別探訪)HSS`
    m = re.match(r"^\((特別探訪[^)]*)\)\s*(.*)$", text)
    if m:
        return {"kind": "logistics", "service_code_raw": "特別探訪",
                "inline_note": m.group(2).strip() or None,
                "confidence": "medium"}
    matched = _match_field_code(text)
    if matched is not None:
        code, rest = matched
        fields = _parse_field_remainder(rest, block_weekday)
        warnings = fields.pop("warnings")
        confidence = "high"
        if warnings:
            confidence = "medium"
        if fields.get("elder_alias") is None:
            confidence = "low"
        return {"kind": "field_service", "service_code_raw": code,
                "service_code": FIELD_TO_CANONICAL.get(code),
                "confidence": confidence, "warnings": tuple(warnings), **fields}
    return None


# ---------------------------------------------------------------------------
# grammar: detail cells
# ---------------------------------------------------------------------------

def _parse_detail(cell: CellRef, value: object) -> ParsedDetail:
    raw = _raw_str(value)
    if isinstance(value, _dt.time):
        return ParsedDetail(cell=cell, raw_text=raw,
                            start_time=value.strftime("%H:%M"))
    text = _normalize(value)
    start = end = None
    m = _TIME_RANGE_RE.search(text)
    remainder = text
    if m:
        start = f"{int(m.group(1)):02d}:{m.group(2)}"
        end = f"{int(m.group(3)):02d}:{m.group(4)}"
        remainder = (text[:m.start()] + " " + text[m.end():]).strip()
    else:
        bm = _BARE_TIME_RE.match(text)
        if bm:
            return ParsedDetail(cell=cell, raw_text=raw,
                                start_time=f"{int(bm.group(1)):02d}:{bm.group(2)}")
    district = None
    role_note = None
    rest_parts: list[str] = []
    for pm in _PAREN_RE.finditer(remainder):
        token = pm.group(1).strip()
        if token in DUTY_ROLE_TOKENS:
            role_note = token
        elif district is None and re.search(r"[一-鿿]", token):
            district = token
        else:
            rest_parts.append(f"({token})")
    trailing = _PAREN_RE.sub("", remainder).strip()
    if trailing:
        rest_parts.append(trailing)
    trailing_label = " ".join(rest_parts).strip() or None
    return ParsedDetail(cell=cell, raw_text=raw, start_time=start, end_time=end,
                        district=district, role_note=role_note,
                        trailing_label=trailing_label)


def _looks_like_detail(text: str) -> bool:
    return bool(_TIME_RANGE_RE.search(text) or _BARE_TIME_RE.match(text)
                or _PAREN_RE.fullmatch(text))


# ---------------------------------------------------------------------------
# main importer
# ---------------------------------------------------------------------------

class DivisionSheetImporter:
    """Stateful single-pass importer for the ``恆常服務`` sheet."""

    parser_name = "division.regular_services"
    doc_ref = DOC_REF

    def __init__(self, workbook_path: Path, worksheet: Worksheet):
        self.path = workbook_path
        self.ws = worksheet
        self.used = effective_used_range(worksheet)
        self.workers: dict[int, WorkerColumn] = {}
        self.gap_columns: list[int] = []
        self.counter_columns: list[int] = []
        self.assignments: list[ParsedAssignment] = []
        self.counters: list[CounterObservation] = []
        self.ambiguities: list[ImportAmbiguity] = []
        self.raw_cells: dict[str, RawScheduleCell] = {}   # coordinate -> cell
        self.blocks: list[WeekdayBlock] = []
        self.hours_row: int | None = None
        self.saturday_row: int | None = None

    # ------------------------------------------------------------- plumbing
    def _ref(self, row: int, col: int) -> CellRef:
        return CellRef(sheet_name=self.ws.title, row=row, column=col)

    def _source(self, row: int, col: int, value: object = None) -> SourceRef:
        return SourceRef(workbook_path=self.path, sheet_name=self.ws.title,
                         cell=self._ref(row, col), raw_value=_raw_str(value),
                         doc_ref=self.doc_ref)

    def _classify_cell(self, row: int, col: int, value: object, category: str,
                       *, fill: str | None = None) -> None:
        ref = self._ref(row, col)
        if ref.coordinate in self.raw_cells:
            return  # first classification wins; never double-count
        self.raw_cells[ref.coordinate] = RawScheduleCell(
            cell=ref, raw_text=_raw_str(value), category=category, fill=fill)

    def _flag(self, code: str, message: str, row: int, col: int, value: object,
              *, severity: str = "warning", hint: str | None = None) -> None:
        self.ambiguities.append(ImportAmbiguity(
            code=code, message=message,
            source=self._source(row, col, value),
            severity=severity, raw_value=_raw_str(value),  # type: ignore[arg-type]
            resolution_hint=hint))

    # --------------------------------------------------------------- headers
    def parse_headers(self) -> None:
        assert self.used.max_column is not None
        # counter columns: anchored on the 個案數量 / 總數 headers (row 2);
        # 個案數量 is merged AY2:AZ2, so the anchored column + merged span
        # + everything up to 總數 belongs to the counter group.
        anchor_main = anchor_total = None
        for col in range(FIRST_WORKER_COLUMN, self.used.max_column + 1):
            v = _normalize(self.ws.cell(HEADER_ROW, col).value)
            if v == COUNTER_HEADER_MAIN:
                anchor_main = col
            elif v == COUNTER_HEADER_TOTAL:
                anchor_total = col
        if anchor_main is not None and anchor_total is not None:
            self.counter_columns = list(range(anchor_main, anchor_total + 1))
        elif anchor_main is not None:
            self.counter_columns = [anchor_main, anchor_main + 1]
            self._flag("STRUCTURE_UNEXPECTED", "總數 header not found",
                       HEADER_ROW, anchor_main, COUNTER_HEADER_MAIN)

        last_named = FIRST_WORKER_COLUMN
        for col in range(FIRST_WORKER_COLUMN, self.used.max_column + 1):
            if col in self.counter_columns:
                cell = self.ws.cell(HEADER_ROW, col)
                if not _is_blank(cell.value):  # AZ2 is blank (merged AY2:AZ2)
                    self._classify_cell(HEADER_ROW, col, cell.value,
                                        "counter_header")
                continue
            cell = self.ws.cell(HEADER_ROW, col)
            raw = cell.value
            if _is_blank(raw):
                self.gap_columns.append(col)
                continue
            text = _normalize(raw)
            tags = tuple(t.strip() for t in re.findall(r"\(([^)]*)\)", text))
            name = re.sub(r"\([^)]*\)", "", text).strip()
            fill = read_fill_color(cell)
            if fill in ACTIVE_HEADER_FILLS:
                status = "active"
            elif fill in DEPARTED_HEADER_FILLS:
                status = "departed_inferred"
            else:
                status = "unknown"
            self.workers[col] = WorkerColumn(
                column=col, column_letter=get_column_letter(col),
                raw_header=_raw_str(raw), display_name=name, tags=tags,
                header_fill=fill, status_inferred=status)  # type: ignore[arg-type]
            self._classify_cell(HEADER_ROW, col, raw, "header", fill=fill)
            last_named = col
        # gap columns are only meaningful inside the worker span
        self.gap_columns = [c for c in self.gap_columns if c < last_named]
        # structural header cells in columns A/B (星期 merged A2:B2)
        for col in (1, 2):
            v = self.ws.cell(HEADER_ROW, col).value
            if not _is_blank(v):
                self._classify_cell(HEADER_ROW, col, v, "structural")

    # ---------------------------------------------------------------- blocks
    def detect_blocks(self) -> None:
        weekday_ranges: list[tuple[int, int, str]] = []
        period_ranges: list[tuple[int, int, str]] = []
        for rng in self.ws.merged_cells.ranges:
            anchor = self.ws.cell(rng.min_row, rng.min_col).value
            text = _normalize(anchor)
            if rng.min_col == 1 and rng.max_col == 1 and text[:1] in WEEKDAY_CHARS:
                weekday_ranges.append((rng.min_row, rng.max_row, text))
            elif rng.min_col == 2 and rng.max_col == 2 and text in PERIOD_CHARS:
                period_ranges.append((rng.min_row, rng.max_row, text))
        weekday_ranges.sort()
        period_ranges.sort()

        for w_start, w_end, w_label in weekday_ranges:
            weekday = WEEKDAY_CHARS[w_label[0]]
            periods = []
            for p_start, p_end, p_label in period_ranges:
                if p_start < w_start or p_end > w_end:
                    continue
                rows = tuple(range(p_start, p_end + 1))
                # slot-pair geometry (verified on the real file): rows
                # [0]/[1] = session-1 assignment/detail, [3]/[4] = session 2.
                assignment_rows = tuple(
                    rows[i] for i in (0, 3) if i < len(rows))
                detail_rows = tuple(rows[i] for i in (1, 4) if i < len(rows))
                extra = tuple(r for r in rows
                              if r not in assignment_rows and r not in detail_rows)
                periods.append(PeriodBlock(
                    period=PERIOD_CHARS[p_label], rows=rows,
                    assignment_rows=assignment_rows, detail_rows=detail_rows,
                    extra_rows=extra))
            self.blocks.append(WeekdayBlock(
                weekday=weekday, label_raw=w_label, row_start=w_start,
                row_end=w_end, periods=tuple(periods)))
            # classify the merged label cells
            self._classify_cell(w_start, 1, w_label, "structural")
            for p_start, _, p_label in period_ranges:
                if w_start <= p_start <= w_end:
                    self._classify_cell(p_start, 2, p_label, "structural")

    # --------------------------------------------------- special row finders
    def detect_special_rows(self) -> None:
        assert self.used.max_row is not None
        for row in range(HEADER_ROW + 1, self.used.max_row + 1):
            span_hits = 0
            team_hits = 0
            nonblank = 0
            for col in self.workers:
                v = self.ws.cell(row, col).value
                if _is_blank(v):
                    continue
                nonblank += 1
                text = _normalize(v)
                m = _FULL_DAY_SPAN_RE.match(text)
                if m and int(m.group(1)) <= 10 and int(m.group(3)) >= 16:
                    span_hits += 1
                if _SATURDAY_TEAM_RE.match(text) and len(text) <= 12:
                    team_hits += 1
            if span_hits >= 8 and self.hours_row is None:
                self.hours_row = row
            if team_hits >= 6 and nonblank - team_hits <= 2 \
                    and self.saturday_row is None:
                self.saturday_row = row

    def parse_hours_row(self) -> None:
        if self.hours_row is None:
            self._flag("STRUCTURE_UNEXPECTED", "working-hours row not found",
                       0, 0, "", severity="warning")
            return
        row = self.hours_row
        for col in list(self.workers):
            v = self.ws.cell(row, col).value
            if _is_blank(v):
                continue
            text = _normalize(v)
            m = _FULL_DAY_SPAN_RE.match(text)
            wc = self.workers[col]
            if m:
                self.workers[col] = replace(
                    wc, work_hours_raw=text,
                    work_start=f"{int(m.group(1)):02d}:{m.group(2)}",
                    work_end=f"{int(m.group(3)):02d}:{m.group(4)}")
                self._classify_cell(row, col, v, "work_hours")
            else:
                self._classify_cell(row, col, v, "extra_note")
                self._flag("EXTRA_ROW_NOTE",
                           f"unparsed note on working-hours row: {text!r}",
                           row, col, v, severity="info")
        # any other cells on that row (cols A/B or beyond workers)
        self._sweep_row(row, "work_hours_row")

    def parse_saturday_row(self) -> None:
        if self.saturday_row is None:
            self._flag("STRUCTURE_UNEXPECTED", "Saturday A/B team row not found",
                       0, 0, "", severity="warning")
            return
        row = self.saturday_row
        for col in list(self.workers):
            v = self.ws.cell(row, col).value
            if _is_blank(v):
                continue
            text = _normalize(v)
            m = _SATURDAY_TEAM_RE.match(text)
            wc = self.workers[col]
            if m:
                names = m.group(2).strip() or None
                self.workers[col] = replace(
                    wc, saturday_raw=_raw_str(v), saturday_team=m.group(1),  # type: ignore[arg-type]
                    saturday_names_raw=names)
                self._classify_cell(row, col, v, "saturday_team")
                if names:
                    self._flag(
                        "SATURDAY_NAMES_UNCONFIRMED",
                        f"Saturday team cell carries names {names!r}; their "
                        "meaning (partners vs covered-by) is unconfirmed (Q-A5)",
                        row, col, v, severity="info",
                        hint="ask NGO what the names after A/B mean")
            else:
                self._classify_cell(row, col, v, "extra_note")
                self._flag("EXTRA_ROW_NOTE",
                           f"unparsed Saturday-row cell: {text!r}", row, col, v,
                           severity="info")
        self._sweep_row(row, "saturday_row")

    def _sweep_row(self, row: int, context: str) -> None:
        """Classify leftover non-empty cells of a special row."""
        assert self.used.max_column is not None
        for col in range(1, self.used.max_column + 1):
            v = self.ws.cell(row, col).value
            if _is_blank(v):
                continue
            ref = self._ref(row, col).coordinate
            if ref in self.raw_cells:
                continue
            self._classify_cell(row, col, v, "extra_note")
            self._flag("EXTRA_ROW_NOTE",
                       f"unclassified cell on {context}: {_normalize(v)!r}",
                       row, col, v, severity="info")

    # ------------------------------------------------------------ main grid
    def parse_blocks(self) -> None:
        for block in self.blocks:
            for pb in block.periods:
                for si, arow in enumerate(pb.assignment_rows, start=1):
                    drow = pb.detail_rows[si - 1] if si - 1 < len(pb.detail_rows) else None
                    self._parse_session(block.weekday, pb.period, si, arow, drow)
                for erow in pb.extra_rows:
                    if erow in (self.hours_row, self.saturday_row):
                        continue
                    self._parse_extra_row(block.weekday, pb.period, erow)

    def _parse_session(self, weekday: int, period: str, session: int,
                       arow: int, drow: int | None) -> None:
        for col, worker in self.workers.items():
            a_val = self.ws.cell(arow, col).value
            d_val = self.ws.cell(drow, col).value if drow else None
            assignment = None
            if not _is_blank(a_val):
                assignment = self._make_assignment(
                    weekday, period, session, arow, col, worker, a_val)
            if _is_blank(d_val):
                continue
            d_text = _normalize(d_val)
            # a detail-position cell may itself be an assignment: stacked
            # alternating-week cases (fact-check E5) or a shifted duty cell
            fields = _classify_assignment(d_text, weekday)
            if fields is not None and (
                    fields["kind"] == "field_service"
                    or (assignment is None and fields["kind"] != "off")):
                stacked = assignment is not None
                self._emit_assignment(weekday, period, session, drow, col,
                                      worker, d_val, fields, stacked=stacked)
                continue
            detail = _parse_detail(self._ref(drow, col), d_val)
            if assignment is not None:
                idx = self.assignments.index(assignment)
                self.assignments[idx] = replace(assignment, detail=detail)
                self._classify_cell(drow, col, d_val, "detail")
            else:
                self._classify_cell(drow, col, d_val, "orphan_detail")
                self._flag("ORPHAN_DETAIL",
                           f"detail text without an assignment above: {d_text!r}",
                           drow, col, d_val, severity="info",
                           hint="may belong to a neighbouring slot; confirm")

    def _make_assignment(self, weekday: int, period: str, session: int | None,
                         row: int, col: int, worker: WorkerColumn,
                         value: object, *, overflow: bool = False
                         ) -> ParsedAssignment | None:
        text = _normalize(value)
        fields = _classify_assignment(text, weekday)
        if fields is None:
            self._classify_cell(row, col, value, "ambiguous")
            code = "BARE_NAME" if re.fullmatch(r"[一-鿿A-Za-z]{1,6}", text) \
                else "UNPARSED_ASSIGNMENT"
            self._flag(code, f"assignment cell did not match the grammar: {text!r}",
                       row, col, value,
                       hint="verify with NGO / extend grammar vocabulary")
            return None
        if fields["kind"] == "field_service" and fields.get("elder_alias") is None:
            self._classify_cell(row, col, value, "ambiguous")
            self._flag("INCOMPLETE_ASSIGNMENT",
                       f"service code without a case: {text!r}", row, col, value,
                       hint="cyan incomplete cell — ask the roster owner")
            return None
        return self._emit_assignment(weekday, period, session, row, col, worker,
                                     value, fields, overflow=overflow)

    def _emit_assignment(self, weekday: int, period: str, session: int | None,
                         row: int, col: int, worker: WorkerColumn, value: object,
                         fields: dict, *, stacked: bool = False,
                         overflow: bool = False) -> ParsedAssignment:
        fill = read_fill_color(self.ws.cell(row, col))
        warnings = tuple(fields.pop("warnings", ()))
        assignment = ParsedAssignment(
            cell=self._ref(row, col), worker_column=col,
            worker_alias=worker.display_name, weekday=weekday, period=period,
            session_index=None if overflow else session,
            raw_text=_raw_str(value), fill=fill, stacked=stacked,
            overflow=overflow, warnings=warnings, **fields)
        self.assignments.append(assignment)
        self._classify_cell(row, col, value, "assignment", fill=fill)
        for w in warnings:
            severity = "warning" if "unconfirmed" in w or "unrecognised" in w else "info"
            code = ("UNKNOWN_WEEK_PATTERN"
                    if "長周" in w or "pattern" in w else "GRAMMAR_NOTE")
            self._flag(code, w, row, col, value, severity=severity)
        return assignment

    def _parse_extra_row(self, weekday: int, period: str, row: int) -> None:
        for col, worker in self.workers.items():
            v = self.ws.cell(row, col).value
            if _is_blank(v):
                continue
            text = _normalize(v)
            fields = _classify_assignment(text, weekday)
            if fields is not None and fields["kind"] != "off":
                self._emit_assignment(weekday, period, None, row, col, worker,
                                      v, fields, overflow=True)
                continue
            if fields is not None and fields["kind"] == "off":
                self._emit_assignment(weekday, period, None, row, col, worker,
                                      v, fields, overflow=True)
                continue
            self._classify_cell(row, col, v, "extra_note")
            severity = "info" if _looks_like_detail(text) else "warning"
            self._flag("EXTRA_ROW_NOTE",
                       f"extra-row cell not parsed as assignment: {text!r}",
                       row, col, v, severity=severity,
                       hint="often extended time windows or handover notes")

    # ---------------------------------------------------------------- counters
    def parse_counters(self) -> None:
        if not self.counter_columns:
            return
        col_main = self.counter_columns[0]
        col_other = self.counter_columns[1] if len(self.counter_columns) > 1 else None
        col_total = self.counter_columns[-1] if len(self.counter_columns) > 2 else None
        for block in self.blocks:
            for pb in block.periods:
                for si, arow in enumerate(pb.assignment_rows, start=1):
                    label_row = arow + 1
                    ero_exp = _as_int(self.ws.cell(arow, col_main).value)
                    other_exp = _as_int(self.ws.cell(arow, col_other).value) \
                        if col_other else None
                    total_exp = _as_int(self.ws.cell(arow, col_total).value) \
                        if col_total else None
                    other_label = _normalize(
                        self.ws.cell(label_row, col_other).value) or None \
                        if col_other else None
                    row_assignments = [a for a in self.assignments
                                       if a.cell.row == arow]
                    ero = sum(1 for a in row_assignments
                              if a.service_code_raw == "E+RO")
                    esc = sum(1 for a in row_assignments
                              if a.kind == "escort_slot"
                              or a.service_code_raw == "Esc")
                    d = sum(1 for a in row_assignments if a.kind == "meal")
                    matches: bool | None = None
                    if ero_exp is not None:
                        counted_other = {"Esc": esc, "D": d}.get(other_label or "", None)
                        matches = (ero == ero_exp) and (
                            other_exp is None or counted_other is None
                            or counted_other == other_exp)
                        if not matches:
                            self._flag(
                                "COUNTER_MISMATCH",
                                f"row {arow}: sheet counters say E+RO={ero_exp}"
                                f", {other_label}={other_exp}; importer counted "
                                f"E+RO={ero}, Esc={esc}, D={d}",
                                arow, col_main,
                                self.ws.cell(arow, col_main).value,
                                severity="warning",
                                hint="validation data — reconcile grammar vs sheet")
                    self.counters.append(CounterObservation(
                        row=arow, weekday=block.weekday, period=pb.period,
                        session_index=si, ero_expected=ero_exp,
                        other_label=other_label, other_expected=other_exp,
                        total_expected=total_exp, ero_counted=ero,
                        esc_counted=esc, d_counted=d, matches=matches))
                    # classify counter + label cells
                    for r in (arow, label_row):
                        for c in self.counter_columns:
                            v = self.ws.cell(r, c).value
                            if not _is_blank(v):
                                cat = "counter" if r == arow else "counter_label"
                                self._classify_cell(r, c, v, cat)

    # ------------------------------------------------------------- comments
    def parse_comments(self) -> None:
        for row in self.ws.iter_rows(min_row=1,
                                     max_row=self.used.max_row or 1,
                                     max_col=self.used.max_column or 1):
            for cell in row:
                comment = getattr(cell, "comment", None)
                if comment is not None and comment.text and comment.text.strip():
                    self._flag(
                        "CELL_COMMENT",
                        f"cell comment carries schedule information: "
                        f"{comment.text.strip()!r}",
                        cell.row, cell.column, cell.value, severity="info",
                        hint="comments often hold effective-dated changes")

    # ---------------------------------------------------------- final sweep
    def sweep_unclassified(self) -> None:
        """Every remaining non-empty cell becomes structural or an ambiguity."""
        assert self.used.max_row is not None and self.used.max_column is not None
        block_rows = {r for b in self.blocks for p in b.periods for r in p.rows}
        for row in range(1, self.used.max_row + 1):
            for col in range(1, self.used.max_column + 1):
                v = self.ws.cell(row, col).value
                if _is_blank(v):
                    continue
                ref = self._ref(row, col).coordinate
                if ref in self.raw_cells:
                    continue
                if row == 1 or (row == HEADER_ROW):
                    self._classify_cell(row, col, v, "structural")
                    continue
                if col in (1, 2):
                    self._classify_cell(row, col, v, "structural")
                    continue
                category = "out_of_block" if row not in block_rows else "extra_note"
                self._classify_cell(row, col, v, category)
                self._flag("OUT_OF_BLOCK_NOTE" if category == "out_of_block"
                           else "EXTRA_ROW_NOTE",
                           f"unclassified cell: {_normalize(v)!r}", row, col, v,
                           severity="info")

    # ----------------------------------------------------------------- run
    def run(self) -> DivisionImportResult:
        self.parse_headers()
        self.detect_blocks()
        self.detect_special_rows()
        self.parse_hours_row()
        self.parse_saturday_row()
        self.parse_blocks()
        self.parse_counters()
        self.parse_comments()
        self.sweep_unclassified()

        nonempty = self.used.value_count
        classified = len(self.raw_cells)
        silently_dropped = nonempty - classified
        counter_mismatches = sum(1 for c in self.counters if c.matches is False)
        candidates = self.fixed_service_candidates()

        summary = {
            "parser_name": self.parser_name,
            "status": "ok" if silently_dropped == 0 else "partial",
            "worker_count": len(self.workers),
            "gap_columns": [get_column_letter(c) for c in self.gap_columns],
            "counter_columns": [get_column_letter(c) for c in self.counter_columns],
            "weekday_blocks": len(self.blocks),
            "assignment_count": len(self.assignments),
            "stacked_assignment_count": sum(1 for a in self.assignments if a.stacked),
            "overflow_assignment_count": sum(1 for a in self.assignments if a.overflow),
            "fixed_service_candidate_count": len(candidates),
            "counter_rows": len(self.counters),
            "counter_mismatch_count": counter_mismatches,
            "ambiguity_count": len(self.ambiguities),
            "nonempty_cells": nonempty,
            "classified_cells": classified,
            "silently_dropped_cells": silently_dropped,
            "hours_row": self.hours_row,
            "saturday_row": self.saturday_row,
            "doc_ref": self.doc_ref,
        }
        return DivisionImportResult(
            workbook_path=str(self.path),
            sheet_name=self.ws.title,
            imported_at=_dt.datetime.now(_dt.timezone.utc).isoformat(),
            declared_max_row=self.ws.max_row,
            declared_max_column=self.ws.max_column,
            used_range={
                "min_row": self.used.min_row, "min_column": self.used.min_column,
                "max_row": self.used.max_row, "max_column": self.used.max_column,
                "value_count": self.used.value_count,
            },
            workers=tuple(self.workers[c] for c in sorted(self.workers)),
            gap_columns=tuple(get_column_letter(c) for c in self.gap_columns),
            counter_columns=tuple(get_column_letter(c) for c in self.counter_columns),
            weekday_blocks=tuple(self.blocks),
            assignments=tuple(self.assignments),
            fixed_service_candidates=tuple(candidates),
            counters=tuple(self.counters),
            raw_cells=tuple(self.raw_cells.values()),
            ambiguities=tuple(self.ambiguities),
            summary=summary,
        )

    # ----------------------------------------------------------- candidates
    def fixed_service_candidates(self) -> list[FixedServiceCandidate]:
        out: list[FixedServiceCandidate] = []
        for a in self.assignments:
            if a.kind != "field_service":
                continue
            detail = a.detail
            out.append(FixedServiceCandidate(
                service_code=a.service_code,
                service_code_raw=a.service_code_raw or "",
                weekday=a.weekday, period=a.period,
                session_index=a.session_index,
                worker_alias=a.worker_alias,
                worker_column_letter=get_column_letter(a.worker_column),
                elder_alias=a.elder_alias, unit=a.unit,
                week_pattern_raw=a.week_pattern_raw,
                week_pattern_weeks=a.week_pattern_weeks,
                start_time=detail.start_time if detail else None,
                end_time=detail.end_time if detail else None,
                district=detail.district if detail else None,
                case_manager_candidate=detail.trailing_label if detail else None,
                inline_note=a.inline_note,
                stacked=a.stacked,
                source_ref=a.cell.label,
                confidence=a.confidence,
                warnings=a.warnings,
            ))
        return out


def _as_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    return int(text) if text.isdigit() else None


# ---------------------------------------------------------------------------
# public API
# ---------------------------------------------------------------------------

def parse_division_workbook(path: Path | str) -> DivisionImportResult:
    """Parse the division workbook's ``恆常服務`` sheet from a file path."""
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=True)
    return parse_regular_services_sheet(wb, workbook_path=workbook_path)


def parse_regular_services_sheet(workbook: Workbook, *,
                                 workbook_path: Path | None = None
                                 ) -> DivisionImportResult:
    ws = require_sheet(workbook, REGULAR_SERVICES_SHEET)
    importer = DivisionSheetImporter(workbook_path or Path("<memory>"), ws)
    return importer.run()
