"""Fail-closed two-week manual-workbook comparison harness.

The harness consumes stored API payloads.  It never calls the scheduler and it
never interprets an empty/manual workbook cell as new scheduling demand.  The
roster owner's workbook remains the operational source of truth; a separate
ledger supplies only explicit disposition observations and classifications.
"""
from __future__ import annotations

import csv
import json
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from pydantic import (
    BaseModel,
    ConfigDict,
    Field,
    ValidationError,
    field_validator,
    model_validator,
)

from ..domain import (
    DemandReconciliationReport,
    ScheduleVersion,
    canonical_json,
    stable_id,
)
from ..exporter.division_writer import render_export_placement_values
from ..scheduler.reconciliation import version_content_hash


SCHEMA_VERSION = "1.0"
CSV_LEDGER_COLUMNS = (
    "schema_version",
    "week_start",
    "row_type",
    "demand_id",
    "manual_disposition",
    "reference",
    "manual_key",
    "cell",
    "diff_id",
    "entry_id",
    "cell_or_ref",
    "generated_exists",
    "manual_exists",
    "generated_value_json",
    "manual_value_json",
    "category",
    "note",
    "reviewer",
    "reviewed_at",
    "blocking_reason",
)
Disposition = Literal[
    "scheduled",
    "needs_review",
    "unassigned",
    "confirmed_cancelled",
    "suppressed_with_audit",
]
DiffCategory = Literal["expected", "reviewer_approved", "blocking"]
Scope = Literal["fixture_smoke", "real_parallel_run"]


class ParallelRunValidationError(ValueError):
    """Structured input/comparison failure safe to serialize to CLI JSON."""

    def __init__(
        self,
        code: str,
        message: str,
        *,
        week_start: str | None = None,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.week_start = week_start
        self.details = details or {}


class _StrictModel(BaseModel):
    model_config = ConfigDict(extra="forbid")


class RosterOwnerSignoff(_StrictModel):
    reviewer: str
    signed_at: datetime
    evidence_ref: str | None = None

    @field_validator("reviewer", "evidence_ref", mode="before")
    @classmethod
    def _strip_optional(cls, value: Any) -> Any:
        return value.strip() if isinstance(value, str) else value

    @model_validator(mode="after")
    def _reviewer_required(self) -> "RosterOwnerSignoff":
        if not self.reviewer:
            raise ValueError("roster-owner signoff reviewer must not be empty")
        return self


class NgoMasterDataEvidence(_StrictModel):
    confirmed: bool = False
    evidence_refs: list[str] = Field(default_factory=list)

    @field_validator("evidence_refs")
    @classmethod
    def _canonical_refs(cls, values: list[str]) -> list[str]:
        cleaned = sorted({value.strip() for value in values if value.strip()})
        if len(cleaned) != len(values):
            raise ValueError("NGO master-data evidence refs must be unique and non-empty")
        return cleaned

    @model_validator(mode="after")
    def _confirmation_needs_evidence(self) -> "NgoMasterDataEvidence":
        if self.confirmed and not self.evidence_refs:
            raise ValueError("confirmed NGO master data requires evidence_refs")
        return self


class WeekCase(_StrictModel):
    week_start: date
    generated_run: str
    manual_workbook: str
    comparison_ledger: str
    roster_owner_signoff: RosterOwnerSignoff | None = None

    @field_validator("generated_run", "manual_workbook", "comparison_ledger", mode="before")
    @classmethod
    def _input_ref_required(cls, value: Any) -> Any:
        if not isinstance(value, str) or not value.strip():
            raise ValueError("week input references must be non-empty strings")
        return value.strip()

    @model_validator(mode="after")
    def _monday_only(self) -> "WeekCase":
        if self.week_start.weekday() != 0:
            raise ValueError("week_start must be a Monday")
        return self


class ParallelRunManifest(_StrictModel):
    schema_version: Literal["1.0"] = SCHEMA_VERSION
    scope: Scope
    weeks: list[WeekCase]
    ngo_master_data: NgoMasterDataEvidence = Field(default_factory=NgoMasterDataEvidence)

    @model_validator(mode="after")
    def _exactly_two_unique_weeks(self) -> "ParallelRunManifest":
        if len(self.weeks) != 2:
            raise ValueError("parallel-run manifest requires exactly two week cases")
        starts = [item.week_start for item in self.weeks]
        if len(set(starts)) != 2:
            raise ValueError("parallel-run week_start values must be unique")
        return self


class DispositionObservation(_StrictModel):
    demand_id: str
    manual_disposition: Disposition
    reference: str

    @field_validator("demand_id", "reference", mode="before")
    @classmethod
    def _required_text(cls, value: Any) -> Any:
        if not isinstance(value, str) or not value.strip():
            raise ValueError("disposition observation fields must not be empty")
        return value.strip()


class ManualOnlyObservation(_StrictModel):
    manual_key: str
    cell: str

    @field_validator("manual_key", "cell", mode="before")
    @classmethod
    def _required_text(cls, value: Any) -> Any:
        if not isinstance(value, str) or not value.strip():
            raise ValueError("manual-only fields must not be empty")
        return value.strip()


class LedgerDiff(_StrictModel):
    diff_id: str
    week_start: date
    demand_id: str | None = None
    manual_key: str | None = None
    entry_id: str | None = None
    cell_or_ref: str
    generated_exists: bool
    manual_exists: bool
    generated_value: Any = Field(...)
    manual_value: Any = Field(...)
    category: DiffCategory
    note: str
    reviewer: str | None = None
    reviewed_at: datetime | None = None
    blocking_reason: str | None = None

    @field_validator(
        "diff_id",
        "demand_id",
        "manual_key",
        "entry_id",
        "cell_or_ref",
        "note",
        "reviewer",
        "blocking_reason",
        mode="before",
    )
    @classmethod
    def _strip_text(cls, value: Any) -> Any:
        return value.strip() if isinstance(value, str) else value

    @model_validator(mode="after")
    def _classification_contract(self) -> "LedgerDiff":
        if not self.diff_id or not self.cell_or_ref or not self.note:
            raise ValueError("diff_id, cell_or_ref, and note must not be empty")
        if bool(self.demand_id) == bool(self.manual_key):
            raise ValueError("each diff needs exactly one demand_id or manual_key")
        if self.category == "reviewer_approved" and (
            not self.reviewer or self.reviewed_at is None
        ):
            raise ValueError(
                "reviewer_approved diff requires reviewer and reviewed_at"
            )
        if self.category == "blocking" and not self.blocking_reason:
            raise ValueError("blocking diff requires blocking_reason")
        return self


class ComparisonLedger(_StrictModel):
    schema_version: Literal["1.0"] = SCHEMA_VERSION
    week_start: date
    disposition_comparisons: list[DispositionObservation] = Field(default_factory=list)
    manual_only: list[ManualOnlyObservation] = Field(default_factory=list)
    diffs: list[LedgerDiff] = Field(default_factory=list)

    @model_validator(mode="after")
    def _unique_identity_rows(self) -> "ComparisonLedger":
        demand_ids = [item.demand_id for item in self.disposition_comparisons]
        manual_keys = [item.manual_key for item in self.manual_only]
        manual_cells = [item.cell for item in self.manual_only]
        diff_ids = [item.diff_id for item in self.diffs]
        for label, values in (
            ("disposition demand IDs", demand_ids),
            ("manual-only keys", manual_keys),
            ("manual-only cells", manual_cells),
            ("diff IDs", diff_ids),
        ):
            if len(values) != len(set(values)):
                raise ValueError(f"ledger contains duplicate {label}")
        return self


class RunPlacement(_StrictModel):
    demand_id: str
    entry_id: str
    disposition: Disposition
    version_id: str
    status: str
    worker_name: str
    schedule_date: date
    period: str
    service_code: str
    target: str
    assignment_cell: str
    detail_cell: str | None = None
    audit_ids: list[str] = Field(default_factory=list)
    data_gap_ids: list[str] = Field(default_factory=list)
    source_evidence_ids: list[str] = Field(default_factory=list)
    source_refs: list[str] = Field(default_factory=list)


@dataclass(frozen=True)
class _LoadedRun:
    version: ScheduleVersion
    reconciliation: DemandReconciliationReport
    placements: tuple[RunPlacement, ...]
    decision_actions: tuple[str, ...]


@dataclass(frozen=True)
class _WeekEvaluation:
    metrics: dict[str, Any]
    blocking_ids: tuple[str, ...]
    engineering_issues: tuple[str, ...]


def evaluate_parallel_run(manifest_path: str | Path) -> dict[str, Any]:
    """Evaluate two explicit weeks and return deterministic report data."""

    path = Path(manifest_path)
    manifest_payload = _load_json(path, "MANIFEST_READ_FAILED", "manifest")
    manifest = _validate_model(
        ParallelRunManifest,
        manifest_payload,
        code="MANIFEST_SCHEMA_INVALID",
        label="manifest",
    )
    base = path.resolve().parent
    evaluations = [
        _evaluate_week(base, week)
        for week in sorted(manifest.weeks, key=lambda item: item.week_start)
    ]

    weekly = [item.metrics for item in evaluations]
    engineering_issues = sorted({
        issue for item in evaluations for issue in item.engineering_issues
    })
    blocking_ids = sorted({
        diff_id for item in evaluations for diff_id in item.blocking_ids
    })
    engineering_state = "passed" if not engineering_issues else "blocked"
    comparison_reasons = []
    if engineering_issues:
        comparison_reasons.append("engineering_gate_blocked")
    if blocking_ids:
        comparison_reasons.append("blocking_differences")
    comparison_state = "passed" if not comparison_reasons else "blocked"

    total_metrics = _aggregate_metrics(weekly)
    ngo_gate, ngo_missing = _ngo_gate(
        manifest,
        engineering_state=engineering_state,
        comparison_state=comparison_state,
    )
    claims_ngo_acceptance = ngo_gate == "accepted"
    overall = (
        "passed"
        if engineering_state == "passed" and comparison_state == "passed"
        else "blocked"
    )
    return {
        "schema_version": SCHEMA_VERSION,
        "status": overall,
        "scope": manifest.scope,
        "engineering_gate": {
            "state": engineering_state,
            "issues": engineering_issues,
        },
        "comparison_gate": {
            "state": comparison_state,
            "reasons": comparison_reasons,
            "blocking_ids": blocking_ids,
        },
        "ngo_gate": ngo_gate,
        "ngo_gate_missing": ngo_missing,
        "claims_ngo_acceptance": claims_ngo_acceptance,
        "weekly": weekly,
        "total": total_metrics,
        "interpretation": (
            "NGO gate acceptance records evidence for the two-week parallel run; "
            "it is not a staff-readiness or automatic-publication claim."
        ),
    }


def failure_report(error: ParallelRunValidationError) -> dict[str, Any]:
    """Return deterministic structured CLI output for a fail-closed error."""

    detail: dict[str, Any] = {
        "code": error.code,
        "message": error.message,
    }
    if error.week_start is not None:
        detail["week_start"] = error.week_start
    if error.details:
        detail["details"] = error.details
    return {
        "schema_version": SCHEMA_VERSION,
        "status": "failed",
        "engineering_gate": {"state": "blocked"},
        "comparison_gate": {"state": "not_evaluated"},
        "ngo_gate": "not_evaluated",
        "claims_ngo_acceptance": False,
        "error": detail,
    }


def canonical_report_json(report: dict[str, Any], *, pretty: bool = False) -> str:
    """Serialize reports reproducibly; no runtime or filesystem metadata is added."""

    compact = canonical_json(report)
    if not pretty:
        return compact + "\n"
    return json.dumps(
        json.loads(compact),
        ensure_ascii=False,
        sort_keys=True,
        indent=2,
    ) + "\n"


def _evaluate_week(base: Path, week: WeekCase) -> _WeekEvaluation:
    week_text = week.week_start.isoformat()
    run_path = _required_input(base, week.generated_run, "generated_run", week_text)
    workbook_path = _required_input(
        base, week.manual_workbook, "manual_workbook", week_text
    )
    ledger_path = _required_input(
        base, week.comparison_ledger, "comparison_ledger", week_text
    )
    run = _load_run(run_path, week.week_start)
    ledger_payload = _load_ledger(ledger_path, week_text)
    ledger = _validate_model(
        ComparisonLedger,
        ledger_payload,
        code="LEDGER_SCHEMA_INVALID",
        label="comparison ledger",
        week_start=week_text,
    )
    if ledger.week_start != week.week_start:
        raise ParallelRunValidationError(
            "LEDGER_WEEK_MISMATCH",
            "comparison ledger week_start does not match manifest",
            week_start=week_text,
        )

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    except Exception as exc:  # openpyxl emits backend-specific exception types
        raise ParallelRunValidationError(
            "MANUAL_WORKBOOK_INVALID",
            "manual comparison workbook could not be opened",
            week_start=week_text,
        ) from exc
    try:
        actual_diffs, cell_counts, disposition_counts = _collect_actual_diffs(
            run,
            ledger,
            workbook,
            week.week_start,
        )
    finally:
        workbook.close()

    classified = _validate_classifications(actual_diffs, ledger, week.week_start)
    categories = Counter(item.category for item in classified)
    blocking_ids = tuple(sorted(
        item.diff_id for item in classified if item.category == "blocking"
    ))
    decisions = Counter(run.decision_actions)
    decision_denominator = sum(decisions[action] for action in ("approve", "reject", "edit"))
    approval_ratio = (
        decisions["approve"] / decision_denominator
        if decision_denominator
        else None
    )
    reconciliation = run.reconciliation
    disposition_values = {
        name: int(getattr(reconciliation, name))
        for name in (
            "scheduled",
            "needs_review",
            "unassigned",
            "confirmed_cancelled",
            "suppressed_with_audit",
        )
    }
    engineering_issues: list[str] = []
    if reconciliation.errors:
        engineering_issues.append(f"{week_text}:reconciliation_errors")
    if reconciliation.hard_violation_count:
        engineering_issues.append(f"{week_text}:hard_constraint_violations")
    if reconciliation.export_failure_count:
        engineering_issues.append(f"{week_text}:export_failures")

    metrics = {
        "week_start": week_text,
        "demand_counts": {
            "total": reconciliation.weekly_demand_total,
            **disposition_values,
        },
        "disposition_comparison_counts": disposition_counts,
        "reconciliation_counts": {
            "disposition_total": sum(disposition_values.values()),
            "pending_audit_total": reconciliation.pending_audit_counts.get("total", 0),
            "decided_audit_total": reconciliation.decided_audit_counts.get("total", 0),
            "placement_count": reconciliation.placement_count,
            "changed_cell_count": reconciliation.changed_cell_count,
            "error_count": len(reconciliation.errors),
        },
        "placement_cell_comparison_counts": cell_counts,
        "diff_category_counts": {
            "total": len(classified),
            "expected": categories["expected"],
            "reviewer_approved": categories["reviewer_approved"],
            "blocking": categories["blocking"],
            "manual_only": sum(1 for item in classified if item.manual_key),
        },
        "uncategorized_ids": [],
        "blocking_ids": list(blocking_ids),
        "decision_counts": {
            "total": sum(decisions.values()),
            "approve": decisions["approve"],
            "reject": decisions["reject"],
            "edit": decisions["edit"],
            "revalidate": decisions["revalidate"],
        },
        "unchanged_approval_ratio": {
            "value": approval_ratio,
            "numerator": decisions["approve"],
            "denominator": decision_denominator,
            "explanation": (
                "approve decisions divided by approve, reject, and edit decisions"
                if decision_denominator
                else "no approve, reject, or edit decisions"
            ),
        },
        "hard_violation_count": reconciliation.hard_violation_count,
        "export_failure_count": reconciliation.export_failure_count,
    }
    return _WeekEvaluation(
        metrics=metrics,
        blocking_ids=blocking_ids,
        engineering_issues=tuple(engineering_issues),
    )


def _load_run(path: Path, expected_week: date) -> _LoadedRun:
    week_text = expected_week.isoformat()
    payload = _load_json(
        path,
        "GENERATED_RUN_READ_FAILED",
        "generated run",
        week_start=week_text,
    )
    if not isinstance(payload, dict):
        raise ParallelRunValidationError(
            "GENERATED_RUN_SCHEMA_INVALID",
            "generated run must be a JSON object",
            week_start=week_text,
        )
    if payload.get("week_start") != week_text:
        raise ParallelRunValidationError(
            "GENERATED_RUN_WEEK_MISMATCH",
            "generated run week_start does not match manifest",
            week_start=week_text,
        )
    version = _validate_model(
        ScheduleVersion,
        payload.get("version"),
        code="GENERATED_VERSION_INVALID",
        label="generated run version",
        week_start=week_text,
    )
    reconciliation = _validate_model(
        DemandReconciliationReport,
        payload.get("reconciliation"),
        code="GENERATED_RECONCILIATION_INVALID",
        label="generated run reconciliation",
        week_start=week_text,
    )
    export_report = payload.get("export_report")
    if not isinstance(export_report, dict):
        raise ParallelRunValidationError(
            "GENERATED_EXPORT_REPORT_INVALID",
            "generated run export_report must be an object",
            week_start=week_text,
        )
    export_reconciliation = _validate_model(
        DemandReconciliationReport,
        export_report.get("reconciliation"),
        code="GENERATED_EXPORT_REPORT_INVALID",
        label="export report reconciliation",
        week_start=week_text,
    )
    raw_placements = export_report.get("placements")
    if not isinstance(raw_placements, list):
        raise ParallelRunValidationError(
            "GENERATED_EXPORT_REPORT_INVALID",
            "generated run export_report.placements must be a list",
            week_start=week_text,
        )
    placements = tuple(
        _validate_model(
            RunPlacement,
            item,
            code="GENERATED_PLACEMENT_INVALID",
            label=f"placement {index}",
            week_start=week_text,
        )
        for index, item in enumerate(raw_placements)
    )
    raw_violations = export_report.get("validator_violations")
    raw_failures = export_report.get("export_failures")
    if not isinstance(raw_violations, list) or not isinstance(raw_failures, list):
        raise ParallelRunValidationError(
            "GENERATED_EXPORT_REPORT_INVALID",
            "export report violations and failures must be lists",
            week_start=week_text,
        )
    if version.week_start != expected_week:
        raise ParallelRunValidationError(
            "GENERATED_VERSION_WEEK_MISMATCH",
            "generated version week_start does not match manifest",
            week_start=week_text,
        )
    if version.reconciliation is None:
        raise ParallelRunValidationError(
            "GENERATED_VERSION_RECONCILIATION_MISSING",
            "generated version has no reconciliation report",
            week_start=week_text,
        )
    for label, candidate in (
        ("version", version.reconciliation),
        ("export_report", export_reconciliation),
    ):
        if canonical_json(candidate) != canonical_json(reconciliation):
            raise ParallelRunValidationError(
                "GENERATED_RECONCILIATION_MISMATCH",
                f"{label} reconciliation differs from top-level reconciliation",
                week_start=week_text,
            )
    if reconciliation.version_id != version.id:
        raise ParallelRunValidationError(
            "GENERATED_VERSION_ID_MISMATCH",
            "reconciliation version_id does not match generated version",
            week_start=week_text,
        )
    expected_hash = version_content_hash(version)
    if reconciliation.content_hash != expected_hash:
        raise ParallelRunValidationError(
            "GENERATED_CONTENT_HASH_MISMATCH",
            "reconciliation content_hash does not match generated version",
            week_start=week_text,
        )
    if canonical_json(version.demand_dispositions) != canonical_json(
        reconciliation.dispositions
    ):
        raise ParallelRunValidationError(
            "GENERATED_DISPOSITION_MISMATCH",
            "version demand_dispositions differ from reconciliation",
            week_start=week_text,
        )
    _validate_reconciliation_counts(reconciliation, week_text)
    entry_ids = [item.id for item in version.entries]
    if len(entry_ids) != len(set(entry_ids)):
        raise ParallelRunValidationError(
            "GENERATED_ENTRY_DUPLICATE",
            "generated version contains duplicate entry IDs",
            week_start=week_text,
        )
    entries = {item.id: item for item in version.entries}
    dispositions = {item.demand_id: item for item in reconciliation.dispositions}
    placement_demands: set[str] = set()
    placement_cells: set[str] = set()
    for placement in placements:
        if placement.demand_id in placement_demands:
            raise ParallelRunValidationError(
                "GENERATED_PLACEMENT_DUPLICATE",
                "multiple placements reference the same demand",
                week_start=week_text,
            )
        placement_demands.add(placement.demand_id)
        entry = entries.get(placement.entry_id)
        disposition = dispositions.get(placement.demand_id)
        if entry is None or disposition is None:
            raise ParallelRunValidationError(
                "GENERATED_PLACEMENT_LINK_INVALID",
                "placement references a missing entry or demand",
                week_start=week_text,
            )
        expected_worker = entry.worker_name or entry.worker_id or ""
        expected_target = (
            entry.elder_name or entry.center or entry.route or entry.destination or ""
        )
        if (
            entry.demand_id != placement.demand_id
            or disposition.entry_id != placement.entry_id
            or disposition.disposition != placement.disposition
            or placement.version_id != version.id
            or placement.status != entry.status.value
            or placement.worker_name != expected_worker
            or placement.schedule_date != entry.schedule_date
            or placement.period != entry.period.value
            or placement.service_code != entry.service_code.value
            or placement.target != expected_target
        ):
            raise ParallelRunValidationError(
                "GENERATED_PLACEMENT_MISMATCH",
                "placement fields do not align with version entry/disposition",
                week_start=week_text,
                details={"demand_id": placement.demand_id, "entry_id": placement.entry_id},
            )
        for raw_cell in (placement.assignment_cell, placement.detail_cell):
            if raw_cell is None:
                continue
            normalized = _normalize_cell_ref(raw_cell, week_text)
            if normalized in placement_cells:
                raise ParallelRunValidationError(
                    "GENERATED_CELL_DUPLICATE",
                    "multiple placement fields reference the same workbook cell",
                    week_start=week_text,
                    details={"cell": normalized},
                )
            placement_cells.add(normalized)
    if reconciliation.placement_count != len(placements):
        raise ParallelRunValidationError(
            "GENERATED_PLACEMENT_COUNT_MISMATCH",
            "reconciliation placement_count differs from export placements",
            week_start=week_text,
        )
    if (
        len(raw_violations) != reconciliation.hard_violation_count
        or len(raw_failures) != reconciliation.export_failure_count
        or export_report.get("publication_state") != reconciliation.publication_state
    ):
        raise ParallelRunValidationError(
            "GENERATED_EXPORT_REPORT_MISMATCH",
            "export report counts/state differ from reconciliation",
            week_start=week_text,
        )
    _validate_run_reconciliation_structure(version, reconciliation, placements, week_text)
    raw_decisions = payload.get("review_decisions", [])
    if not isinstance(raw_decisions, list):
        raise ParallelRunValidationError(
            "GENERATED_DECISIONS_INVALID",
            "generated run review_decisions must be a list",
            week_start=week_text,
        )
    actions: list[str] = []
    for decision in raw_decisions:
        action = decision.get("action") if isinstance(decision, dict) else None
        if action not in {"approve", "reject", "edit", "revalidate"}:
            raise ParallelRunValidationError(
                "GENERATED_DECISIONS_INVALID",
                "review decision has an unsupported action",
                week_start=week_text,
            )
        actions.append(action)
    return _LoadedRun(
        version=version,
        reconciliation=reconciliation,
        placements=placements,
        decision_actions=tuple(actions),
    )


def _validate_reconciliation_counts(
    report: DemandReconciliationReport,
    week_text: str,
) -> None:
    demand_ids = [item.demand_id for item in report.dispositions]
    if len(demand_ids) != len(set(demand_ids)):
        raise ParallelRunValidationError(
            "GENERATED_DISPOSITION_DUPLICATE",
            "reconciliation contains duplicate demand dispositions",
            week_start=week_text,
        )
    counted = Counter(item.disposition for item in report.dispositions)
    fields = (
        "scheduled",
        "needs_review",
        "unassigned",
        "confirmed_cancelled",
        "suppressed_with_audit",
    )
    if (
        report.weekly_demand_total != len(report.dispositions)
        or report.weekly_demand_total != sum(
            int(getattr(report, name)) for name in fields
        )
        or any(counted[name] != getattr(report, name) for name in fields)
    ):
        raise ParallelRunValidationError(
            "GENERATED_DEMAND_CONSERVATION_INVALID",
            "reconciliation demand/disposition counts do not conserve",
            week_start=week_text,
        )


def _validate_run_reconciliation_structure(
    version: ScheduleVersion,
    report: DemandReconciliationReport,
    placements: tuple[RunPlacement, ...],
    week_text: str,
) -> None:
    entries = {item.id: item for item in version.entries}
    expected_status = {
        "scheduled": "scheduled",
        "needs_review": "needs_review",
        "unassigned": "unassigned",
        "confirmed_cancelled": "cancelled",
    }
    active: list[str] = []
    review: list[str] = []
    unassigned: list[str] = []
    cancellations: list[str] = []
    suppressions: list[str] = []
    audits = {item.id: item for item in version.audit_items}
    if len(audits) != len(version.audit_items):
        raise ParallelRunValidationError(
            "GENERATED_AUDIT_DUPLICATE",
            "generated version contains duplicate audit IDs",
            week_start=week_text,
        )
    for disposition in report.dispositions:
        if disposition.disposition == "suppressed_with_audit":
            if disposition.entry_id is not None:
                raise ParallelRunValidationError(
                    "GENERATED_DISPOSITION_LINK_INVALID",
                    "suppressed disposition must not point to an active entry",
                    week_start=week_text,
                    details={"demand_id": disposition.demand_id},
                )
            suppressions.append(disposition.demand_id)
        else:
            entry = entries.get(disposition.entry_id or "")
            if (
                entry is None
                or entry.demand_id != disposition.demand_id
                or entry.status.value != expected_status[disposition.disposition]
            ):
                raise ParallelRunValidationError(
                    "GENERATED_DISPOSITION_LINK_INVALID",
                    "disposition entry link/status is inconsistent",
                    week_start=week_text,
                    details={"demand_id": disposition.demand_id},
                )
            if disposition.disposition in {"scheduled", "needs_review"}:
                active.append(entry.id)
            if disposition.disposition == "needs_review":
                review.append(entry.id)
            elif disposition.disposition == "unassigned":
                unassigned.append(entry.id)
            elif disposition.disposition == "confirmed_cancelled":
                cancellations.append(entry.id)
        for audit_id in disposition.audit_ids:
            audit = audits.get(audit_id)
            if (
                audit is None
                or disposition.demand_id not in audit.demand_ids
                or (
                    disposition.entry_id is not None
                    and disposition.entry_id not in audit.entry_ids
                )
            ):
                raise ParallelRunValidationError(
                    "GENERATED_AUDIT_LINK_INVALID",
                    "disposition audit link is missing or non-reciprocal",
                    week_start=week_text,
                    details={
                        "demand_id": disposition.demand_id,
                        "audit_id": audit_id,
                    },
                )
    expected_lists = {
        "active_entry_ids": sorted(active),
        "review_entry_ids": sorted(review),
        "unassigned_entry_ids": sorted(unassigned),
        "cancellation_entry_ids": sorted(cancellations),
        "suppression_demand_ids": sorted(suppressions),
    }
    for field, expected in expected_lists.items():
        if sorted(getattr(report, field)) != expected:
            raise ParallelRunValidationError(
                "GENERATED_RECONCILIATION_INDEX_MISMATCH",
                f"reconciliation {field} does not match dispositions",
                week_start=week_text,
            )
    if sorted(item.id for item in version.unassigned) != sorted(unassigned):
        raise ParallelRunValidationError(
            "GENERATED_UNASSIGNED_INDEX_MISMATCH",
            "version unassigned list differs from reconciliation",
            week_start=week_text,
        )
    pending = [item for item in version.audit_items if item.status.value == "pending"]
    decided = [item for item in version.audit_items if item.status.value != "pending"]
    if (
        _audit_counts(pending) != report.pending_audit_counts
        or _audit_counts(decided) != report.decided_audit_counts
    ):
        raise ParallelRunValidationError(
            "GENERATED_AUDIT_COUNT_MISMATCH",
            "reconciliation audit counts differ from version audit items",
            week_start=week_text,
        )
    expected_changed_cells = sum(
        1 + int(item.detail_cell is not None) for item in placements
    )
    if report.changed_cell_count != expected_changed_cells:
        raise ParallelRunValidationError(
            "GENERATED_CHANGED_CELL_COUNT_MISMATCH",
            "reconciliation changed_cell_count differs from placement cells",
            week_start=week_text,
        )
    expected_publication_state = "ready"
    if (
        report.errors
        or report.hard_violation_count
        or report.export_failure_count
        or report.unassigned
        or report.pending_audit_counts.get("blocking:true", 0)
    ):
        expected_publication_state = "blocked"
    elif report.needs_review or report.pending_audit_counts.get("total", 0):
        expected_publication_state = "draft"
    if report.publication_state != expected_publication_state:
        raise ParallelRunValidationError(
            "GENERATED_PUBLICATION_STATE_MISMATCH",
            "reconciliation publication_state is inconsistent",
            week_start=week_text,
        )


def _audit_counts(audits: list[Any]) -> dict[str, int]:
    counts: Counter[str] = Counter()
    for audit in audits:
        counts["total"] += 1
        counts[f"blocking:{str(audit.blocking).lower()}"] += 1
        counts[f"severity:{audit.severity.value}"] += 1
        counts[f"kind:{audit.kind.value}"] += 1
    return dict(sorted(counts.items()))


def _collect_actual_diffs(
    run: _LoadedRun,
    ledger: ComparisonLedger,
    workbook: Any,
    week_start: date,
) -> tuple[list[dict[str, Any]], dict[str, int], dict[str, int]]:
    week_text = week_start.isoformat()
    entries = {item.id: item for item in run.version.entries}
    actual: list[dict[str, Any]] = []
    compared_cells: set[str] = set()
    cell_total = 0
    cell_differences = 0
    placement_demands: set[str] = set()
    for placement in sorted(run.placements, key=lambda item: (item.demand_id, item.entry_id)):
        placement_demands.add(placement.demand_id)
        entry = entries[placement.entry_id]
        assignment_value, detail_value = render_export_placement_values(entry)
        comparisons = [(placement.assignment_cell, assignment_value)]
        if placement.detail_cell is not None:
            comparisons.append((placement.detail_cell, detail_value))
        for raw_cell, expected in comparisons:
            cell = _normalize_cell_ref(raw_cell, week_text)
            compared_cells.add(cell)
            generated_value = _json_cell_value(expected)
            manual_value = _read_workbook_cell(workbook, cell, week_text)
            generated_exists = generated_value is not None
            manual_exists = manual_value is not None
            cell_total += 1
            if canonical_json(
                {"exists": generated_exists, "value": generated_value}
            ) == canonical_json({"exists": manual_exists, "value": manual_value}):
                continue
            cell_differences += 1
            actual.append(_actual_diff(
                week_start=week_start,
                kind="placement_cell",
                demand_id=placement.demand_id,
                entry_id=placement.entry_id,
                cell_or_ref=cell,
                generated_exists=generated_exists,
                manual_exists=manual_exists,
                generated_value=generated_value,
                manual_value=manual_value,
            ))

    nonplacement = {
        item.demand_id: item
        for item in run.reconciliation.dispositions
        if item.demand_id not in placement_demands
    }
    observations = {item.demand_id: item for item in ledger.disposition_comparisons}
    missing = sorted(set(nonplacement) - set(observations))
    extra = sorted(set(observations) - set(nonplacement))
    if missing or extra:
        raise ParallelRunValidationError(
            "LEDGER_DISPOSITION_COVERAGE_INVALID",
            "ledger must cover every and only non-placement demand",
            week_start=week_text,
            details={"missing_demand_ids": missing, "extra_demand_ids": extra},
        )
    disposition_differences = 0
    for demand_id in sorted(nonplacement):
        generated = nonplacement[demand_id]
        observation = observations[demand_id]
        if generated.disposition == observation.manual_disposition:
            continue
        disposition_differences += 1
        actual.append(_actual_diff(
            week_start=week_start,
            kind="disposition",
            demand_id=demand_id,
            entry_id=generated.entry_id,
            cell_or_ref=observation.reference,
            generated_exists=True,
            manual_exists=True,
            generated_value=generated.disposition,
            manual_value=observation.manual_disposition,
        ))

    manual_only_count = 0
    for item in sorted(ledger.manual_only, key=lambda row: row.manual_key):
        cell = _normalize_cell_ref(item.cell, week_text)
        if cell in compared_cells:
            raise ParallelRunValidationError(
                "LEDGER_MANUAL_ONLY_OVERLAP",
                "manual-only cell overlaps a generated placement comparison",
                week_start=week_text,
                details={"manual_key": item.manual_key, "cell": cell},
            )
        manual_value = _read_workbook_cell(workbook, cell, week_text)
        if manual_value is None:
            raise ParallelRunValidationError(
                "LEDGER_MANUAL_ONLY_EMPTY",
                "manual-only reference must point to a non-empty workbook cell",
                week_start=week_text,
                details={"manual_key": item.manual_key, "cell": cell},
            )
        manual_only_count += 1
        actual.append(_actual_diff(
            week_start=week_start,
            kind="manual_only",
            manual_key=item.manual_key,
            cell_or_ref=cell,
            generated_exists=False,
            manual_exists=True,
            generated_value=None,
            manual_value=manual_value,
        ))

    return (
        sorted(actual, key=lambda item: item["diff_id"]),
        {
            "placement_count": len(run.placements),
            "cell_total": cell_total,
            "matched": cell_total - cell_differences,
            "differences": cell_differences,
        },
        {
            "nonplacement_total": len(nonplacement),
            "matched": len(nonplacement) - disposition_differences,
            "differences": disposition_differences,
            "manual_only": manual_only_count,
        },
    )


def _actual_diff(
    *,
    week_start: date,
    kind: Literal["placement_cell", "disposition", "manual_only"],
    cell_or_ref: str,
    generated_exists: bool,
    manual_exists: bool,
    generated_value: Any,
    manual_value: Any,
    demand_id: str | None = None,
    manual_key: str | None = None,
    entry_id: str | None = None,
) -> dict[str, Any]:
    identity = {
        "week_start": week_start,
        "kind": kind,
        "demand_id": demand_id,
        "manual_key": manual_key,
        "entry_id": entry_id,
    }
    if kind in {"placement_cell", "manual_only"}:
        identity["cell"] = cell_or_ref
    return {
        "diff_id": stable_id("dif_", "parallel_run_diff", identity),
        "week_start": week_start.isoformat(),
        "demand_id": demand_id,
        "manual_key": manual_key,
        "entry_id": entry_id,
        "cell_or_ref": cell_or_ref,
        "generated_exists": generated_exists,
        "manual_exists": manual_exists,
        "generated_value": generated_value,
        "manual_value": manual_value,
    }


def _validate_classifications(
    actual_diffs: list[dict[str, Any]],
    ledger: ComparisonLedger,
    week_start: date,
) -> list[LedgerDiff]:
    week_text = week_start.isoformat()
    actual_by_id = {item["diff_id"]: item for item in actual_diffs}
    ledger_by_id = {item.diff_id: item for item in ledger.diffs}
    missing = sorted(set(actual_by_id) - set(ledger_by_id))
    extra = sorted(set(ledger_by_id) - set(actual_by_id))
    if missing or extra:
        templates = [
            {**actual_by_id[diff_id], "classification_required": True}
            for diff_id in missing
        ]
        raise ParallelRunValidationError(
            "LEDGER_DIFF_SET_MISMATCH",
            "ledger has missing or unknown diff classifications",
            week_start=week_text,
            details={
                "uncategorized_ids": missing,
                "unknown_diff_ids": extra,
                "uncategorized_diffs": templates,
            },
        )
    core_fields = tuple(next(iter(actual_by_id.values())).keys()) if actual_by_id else ()
    for diff_id in sorted(actual_by_id):
        row = ledger_by_id[diff_id]
        if row.week_start != week_start:
            raise ParallelRunValidationError(
                "LEDGER_DIFF_WEEK_MISMATCH",
                "ledger diff week_start does not match its week case",
                week_start=week_text,
                details={"diff_id": diff_id},
            )
        dumped = row.model_dump(mode="json")
        mismatched = [
            field
            for field in core_fields
            if canonical_json(dumped.get(field))
            != canonical_json(actual_by_id[diff_id].get(field))
        ]
        if mismatched:
            raise ParallelRunValidationError(
                "LEDGER_DIFF_MISMATCH",
                "ledger diff facts do not match generated/manual inputs",
                week_start=week_text,
                details={"diff_id": diff_id, "fields": mismatched},
            )
    return [ledger_by_id[diff_id] for diff_id in sorted(ledger_by_id)]


def _aggregate_metrics(weekly: list[dict[str, Any]]) -> dict[str, Any]:
    disposition_names = (
        "scheduled",
        "needs_review",
        "unassigned",
        "confirmed_cancelled",
        "suppressed_with_audit",
    )
    demand_counts = {
        "total": sum(item["demand_counts"]["total"] for item in weekly),
        **{
            name: sum(item["demand_counts"][name] for item in weekly)
            for name in disposition_names
        },
    }
    decision_counts = {
        name: sum(item["decision_counts"][name] for item in weekly)
        for name in ("total", "approve", "reject", "edit", "revalidate")
    }
    denominator = sum(decision_counts[name] for name in ("approve", "reject", "edit"))
    return {
        "week_count": len(weekly),
        "demand_counts": demand_counts,
        "disposition_comparison_counts": _sum_named_counts(
            weekly,
            "disposition_comparison_counts",
            ("nonplacement_total", "matched", "differences", "manual_only"),
        ),
        "reconciliation_counts": _sum_named_counts(
            weekly,
            "reconciliation_counts",
            (
                "disposition_total",
                "pending_audit_total",
                "decided_audit_total",
                "placement_count",
                "changed_cell_count",
                "error_count",
            ),
        ),
        "placement_cell_comparison_counts": _sum_named_counts(
            weekly,
            "placement_cell_comparison_counts",
            ("placement_count", "cell_total", "matched", "differences"),
        ),
        "diff_category_counts": _sum_named_counts(
            weekly,
            "diff_category_counts",
            ("total", "expected", "reviewer_approved", "blocking", "manual_only"),
        ),
        "uncategorized_ids": [],
        "blocking_ids": sorted({
            diff_id for item in weekly for diff_id in item["blocking_ids"]
        }),
        "decision_counts": decision_counts,
        "unchanged_approval_ratio": {
            "value": decision_counts["approve"] / denominator if denominator else None,
            "numerator": decision_counts["approve"],
            "denominator": denominator,
            "explanation": (
                "approve decisions divided by approve, reject, and edit decisions"
                if denominator
                else "no approve, reject, or edit decisions"
            ),
        },
        "hard_violation_count": sum(item["hard_violation_count"] for item in weekly),
        "export_failure_count": sum(item["export_failure_count"] for item in weekly),
    }


def _sum_named_counts(
    weekly: list[dict[str, Any]],
    section: str,
    names: tuple[str, ...],
) -> dict[str, int]:
    return {
        name: sum(item[section][name] for item in weekly)
        for name in names
    }


def _ngo_gate(
    manifest: ParallelRunManifest,
    *,
    engineering_state: str,
    comparison_state: str,
) -> tuple[str, list[str]]:
    if manifest.scope == "fixture_smoke":
        return "not_evaluated", []
    missing: list[str] = []
    if engineering_state != "passed":
        missing.append("engineering_gate")
    if comparison_state != "passed":
        missing.append("comparison_gate")
    if not manifest.ngo_master_data.confirmed:
        missing.append("ngo_confirmed_master_data")
    signed = {item.week_start for item in manifest.weeks if item.roster_owner_signoff}
    for item in sorted(manifest.weeks, key=lambda row: row.week_start):
        if item.week_start not in signed:
            missing.append(f"roster_owner_signoff:{item.week_start.isoformat()}")
    return ("accepted", []) if not missing else ("pending", missing)


def _read_workbook_cell(workbook: Any, cell_ref: str, week_text: str) -> Any:
    sheet_name, coordinate = cell_ref.rsplit("!", 1)
    if sheet_name not in workbook.sheetnames:
        raise ParallelRunValidationError(
            "MANUAL_WORKBOOK_SHEET_MISSING",
            "manual workbook is missing a referenced sheet",
            week_start=week_text,
            details={"cell": cell_ref},
        )
    try:
        value = workbook[sheet_name][coordinate].value
    except Exception as exc:
        raise ParallelRunValidationError(
            "MANUAL_WORKBOOK_CELL_INVALID",
            "manual workbook cell reference could not be read",
            week_start=week_text,
            details={"cell": cell_ref},
        ) from exc
    return _json_cell_value(value)


def _json_cell_value(value: Any) -> Any:
    if value == "" or value is None:
        return None
    try:
        return json.loads(canonical_json(value))
    except (TypeError, ValueError) as exc:
        raise ParallelRunValidationError(
            "MANUAL_WORKBOOK_VALUE_UNSUPPORTED",
            "workbook cell value cannot be represented in canonical JSON",
        ) from exc


def _normalize_cell_ref(value: str, week_text: str) -> str:
    if not isinstance(value, str) or "!" not in value:
        raise ParallelRunValidationError(
            "CELL_REFERENCE_INVALID",
            "cell reference must use Sheet!A1 form",
            week_start=week_text,
        )
    sheet, coordinate = value.rsplit("!", 1)
    sheet = sheet.strip()
    if len(sheet) >= 2 and sheet[0] == sheet[-1] == "'":
        sheet = sheet[1:-1].replace("''", "'")
    if not sheet or "[" in sheet or "]" in sheet:
        raise ParallelRunValidationError(
            "CELL_REFERENCE_INVALID",
            "cell reference has an invalid sheet name",
            week_start=week_text,
        )
    try:
        column, row = coordinate_from_string(coordinate.replace("$", ""))
    except ValueError as exc:
        raise ParallelRunValidationError(
            "CELL_REFERENCE_INVALID",
            "cell reference has an invalid coordinate",
            week_start=week_text,
        ) from exc
    return f"{sheet}!{column.upper()}{row}"


def _required_input(base: Path, reference: str, label: str, week_text: str) -> Path:
    path = Path(reference)
    resolved = path.resolve() if path.is_absolute() else (base / path).resolve()
    if not resolved.is_file():
        raise ParallelRunValidationError(
            "REQUIRED_INPUT_MISSING",
            f"required {label} input is missing",
            week_start=week_text,
            details={"input": label},
        )
    return resolved


def _load_ledger(path: Path, week_text: str) -> Any:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(
            path,
            "LEDGER_READ_FAILED",
            "comparison ledger",
            week_start=week_text,
        )
    if suffix != ".csv":
        raise ParallelRunValidationError(
            "LEDGER_FORMAT_UNSUPPORTED",
            "comparison ledger must be JSON or CSV",
            week_start=week_text,
        )
    return _load_csv_ledger(path, week_text)


def _load_csv_ledger(path: Path, week_text: str) -> dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None or set(reader.fieldnames) != set(
                CSV_LEDGER_COLUMNS
            ):
                raise ParallelRunValidationError(
                    "LEDGER_CSV_HEADER_INVALID",
                    "CSV ledger header does not match the stable schema",
                    week_start=week_text,
                    details={"required_columns": list(CSV_LEDGER_COLUMNS)},
                )
            rows = list(reader)
    except ParallelRunValidationError:
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        raise ParallelRunValidationError(
            "LEDGER_READ_FAILED",
            "comparison ledger is missing or invalid CSV",
            week_start=week_text,
        ) from exc
    metadata = [row for row in rows if _csv_text(row.get("row_type")) == "meta"]
    if len(metadata) != 1:
        raise ParallelRunValidationError(
            "LEDGER_CSV_META_INVALID",
            "CSV ledger requires exactly one meta row",
            week_start=week_text,
        )
    meta = metadata[0]
    payload: dict[str, Any] = {
        "schema_version": _csv_text(meta.get("schema_version")),
        "week_start": _csv_text(meta.get("week_start")),
        "disposition_comparisons": [],
        "manual_only": [],
        "diffs": [],
    }
    for row_number, row in enumerate(rows, start=2):
        row_type = _csv_text(row.get("row_type"))
        if row_type == "meta":
            continue
        if row_type == "disposition":
            payload["disposition_comparisons"].append({
                "demand_id": _csv_text(row.get("demand_id")),
                "manual_disposition": _csv_text(row.get("manual_disposition")),
                "reference": _csv_text(row.get("reference")),
            })
            continue
        if row_type == "manual_only":
            payload["manual_only"].append({
                "manual_key": _csv_text(row.get("manual_key")),
                "cell": _csv_text(row.get("cell")),
            })
            continue
        if row_type == "diff":
            payload["diffs"].append(_csv_diff_row(row, row_number, week_text))
            continue
        raise ParallelRunValidationError(
            "LEDGER_CSV_ROW_TYPE_INVALID",
            "CSV ledger contains an unsupported row_type",
            week_start=week_text,
            details={"row": row_number},
        )
    return payload


def _csv_diff_row(
    row: dict[str | None, str | None],
    row_number: int,
    week_text: str,
) -> dict[str, Any]:
    try:
        generated_value = json.loads(_csv_text(row.get("generated_value_json")))
        manual_value = json.loads(_csv_text(row.get("manual_value_json")))
    except json.JSONDecodeError as exc:
        raise ParallelRunValidationError(
            "LEDGER_CSV_VALUE_INVALID",
            "CSV diff values must be valid JSON",
            week_start=week_text,
            details={"row": row_number},
        ) from exc
    return {
        "diff_id": _csv_text(row.get("diff_id")),
        "week_start": _csv_text(row.get("week_start")) or week_text,
        "demand_id": _csv_optional(row.get("demand_id")),
        "manual_key": _csv_optional(row.get("manual_key")),
        "entry_id": _csv_optional(row.get("entry_id")),
        "cell_or_ref": _csv_text(row.get("cell_or_ref")),
        "generated_exists": _csv_bool(
            row.get("generated_exists"), row_number, week_text
        ),
        "manual_exists": _csv_bool(
            row.get("manual_exists"), row_number, week_text
        ),
        "generated_value": generated_value,
        "manual_value": manual_value,
        "category": _csv_text(row.get("category")),
        "note": _csv_text(row.get("note")),
        "reviewer": _csv_optional(row.get("reviewer")),
        "reviewed_at": _csv_optional(row.get("reviewed_at")),
        "blocking_reason": _csv_optional(row.get("blocking_reason")),
    }


def _csv_bool(value: str | None, row_number: int, week_text: str) -> bool:
    normalized = _csv_text(value).lower()
    if normalized not in {"true", "false"}:
        raise ParallelRunValidationError(
            "LEDGER_CSV_BOOLEAN_INVALID",
            "CSV diff existence fields must be true or false",
            week_start=week_text,
            details={"row": row_number},
        )
    return normalized == "true"


def _csv_text(value: str | None) -> str:
    return value.strip() if isinstance(value, str) else ""


def _csv_optional(value: str | None) -> str | None:
    normalized = _csv_text(value)
    return normalized or None


def _load_json(
    path: Path,
    code: str,
    label: str,
    *,
    week_start: str | None = None,
) -> Any:
    try:
        with path.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ParallelRunValidationError(
            code,
            f"{label} is missing or invalid JSON",
            week_start=week_start,
        ) from exc


def _validate_model(
    model: type[BaseModel],
    payload: Any,
    *,
    code: str,
    label: str,
    week_start: str | None = None,
) -> Any:
    try:
        return model.model_validate(payload)
    except ValidationError as exc:
        errors = [
            {
                "location": ".".join(str(part) for part in item["loc"]),
                "message": item["msg"],
                "type": item["type"],
            }
            for item in exc.errors(include_url=False, include_input=False)
        ]
        raise ParallelRunValidationError(
            code,
            f"{label} schema validation failed",
            week_start=week_start,
            details={"validation_errors": errors},
        ) from exc
