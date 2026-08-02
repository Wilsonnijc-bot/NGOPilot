"""NGO-format division workbook writer.

This exporter starts from the original division workbook, keeps the NGO's
visual layout intact, and appends RosterCopiilot review sheets. Cell-level
changes are marked with borders/comments so existing business fill colours are
not overwritten.
"""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
import os
from pathlib import Path
from typing import Iterable, Literal

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.workbook.workbook import Workbook
from pydantic import BaseModel, Field

from ..domain import (
    AuditItem,
    AuditKind,
    AuditStatus,
    DemandReconciliationReport,
    EntryStatus,
    HardViolation,
    ManualReviewReason,
    MockDataset,
    Period,
    ReviewReasonCode,
    ScheduleEntry,
    ScheduleVersion,
    Severity,
    canonical_json,
    content_fingerprint,
)
from ..engine import validate_entries
from ..importer.division_models import DivisionImportResult
from ..scheduler.generator import GeneratedDemands
from ..scheduler.reconciliation import (
    finalize_version_provenance,
    reconcile_weekly_demands,
    version_content_hash,
)

RC_SHEETS = ("RC_變更摘要", "RC_審核", "RC_未分配", "RC_meta")
MARK_SIDE = Side(style="medium", color="C00000")
MARK_BORDER = Border(left=MARK_SIDE, right=MARK_SIDE, top=MARK_SIDE, bottom=MARK_SIDE)


PublicationState = Literal["blocked", "draft", "ready"]


class ExportPlacementFailure(BaseModel):
    """Machine-readable reason an active entry cannot be written to the grid."""

    id: str
    demand_id: str | None = None
    entry_id: str | None = None
    audit_id: str | None = None
    code: str
    message: str
    proposed_disposition: str = "留在 RC_未分配，先由主管覆核"
    source_ref: str | None = None
    schedule_date: str | None = None
    period: str | None = None
    service_code: str | None = None
    worker_name: str | None = None
    target: str | None = None


class ExportUnassignedItem(BaseModel):
    """One row shared by API/UI and the RC_未分配 sheet."""

    id: str
    demand_id: str | None = None
    entry_id: str | None = None
    audit_id: str | None = None
    audit_ids: list[str] = Field(default_factory=list)
    schedule_date: str
    period: str
    service_code: str
    target: str
    code: str
    message: str
    source_ref: str = ""
    next_action: str = "主管覆核後人工分配或修正資料"
    is_export_failure: bool = False
    worker_name: str | None = None
    explanation: str | None = None
    disposition: str = "unassigned"


class ExportPlacement(BaseModel):
    """A safe grid write decided by the preflight manifest."""

    demand_id: str
    entry_id: str
    disposition: str
    version_id: str
    status: str
    worker_name: str
    schedule_date: str
    period: str
    service_code: str
    target: str
    assignment_cell: str
    detail_cell: str | None = None
    audit_ids: list[str] = Field(default_factory=list)
    data_gap_ids: list[str] = Field(default_factory=list)
    source_evidence_ids: list[str] = Field(default_factory=list)
    source_refs: list[str] = Field(default_factory=list)


class GeneratedDivisionExportReport(BaseModel):
    """Preflight/publication report used by API, UI, RC sheets, and errors."""

    reconciliation: DemandReconciliationReport
    publication_state: PublicationState
    publication_label: str
    review_export_allowed: bool
    export_block_reasons: list[str] = Field(default_factory=list)
    validator_violations: list[HardViolation] = Field(default_factory=list)
    export_failures: list[ExportPlacementFailure] = Field(default_factory=list)
    placements: list[ExportPlacement] = Field(default_factory=list)
    unassigned_items: list[ExportUnassignedItem] = Field(default_factory=list)
    pending_audit_count: int = 0
    pending_blocking_audit_count: int = 0
    needs_review_count: int = 0
    changed_cell_count: int = 0


class ExportPreflightError(ValueError):
    """Raised when the staff-facing grid must not be mutated."""

    def __init__(self, report: GeneratedDivisionExportReport) -> None:
        self.report = report
        reasons = "；".join(report.export_block_reasons) or "export preflight failed"
        super().__init__(reasons)


@dataclass
class CellMarker:
    note: str
    marker_kind: Literal["changed", "review"] = "changed"


@dataclass
class PlacementTarget:
    entry: ScheduleEntry
    assignment_row: int
    detail_row: int | None
    col: int
    session: int
    occupied_sessions: list[int]
    assignment_ref: str
    detail_ref: str | None
    audit_ids: list[str] = field(default_factory=list)


@dataclass
class GeneratedDivisionExportPlan:
    report: GeneratedDivisionExportReport
    review_version: ScheduleVersion
    placements: list[PlacementTarget]
    changed_cells: dict[str, CellMarker]
    integrity_hash: str = ""


def _plan_integrity_hash(plan: GeneratedDivisionExportPlan) -> str:
    return content_fingerprint({
        "review_version_hash": version_content_hash(plan.review_version),
        "report": plan.report.model_dump(mode="json"),
        "placements": [
            {
                "entry": target.entry.model_dump(mode="json"),
                "assignment_row": target.assignment_row,
                "detail_row": target.detail_row,
                "col": target.col,
                "session": target.session,
                "occupied_sessions": target.occupied_sessions,
                "assignment_ref": target.assignment_ref,
                "detail_ref": target.detail_ref,
                "audit_ids": target.audit_ids,
            }
            for target in plan.placements
        ],
        "changed_cells": {
            ref: {"note": marker.note, "marker_kind": marker.marker_kind}
            for ref, marker in sorted(plan.changed_cells.items())
        },
    })


def validate_prepared_division_export_plan(
    plan: GeneratedDivisionExportPlan,
    version: ScheduleVersion,
) -> None:
    """Reject a stale or internally split export plan before workbook writes."""

    reconciliation = plan.review_version.reconciliation
    if version.id != plan.review_version.id:
        raise ValueError("prepared export plan version does not match requested version")
    if reconciliation is None:
        raise ValueError("prepared export plan has no canonical reconciliation")
    recomputed_hash = version_content_hash(plan.review_version)
    if reconciliation.content_hash != recomputed_hash:
        raise ValueError("prepared export plan reconciliation content hash is invalid")
    if plan.report.reconciliation.content_hash != recomputed_hash:
        raise ValueError("prepared export report content hash is invalid")
    if version_content_hash(version) != recomputed_hash:
        raise ValueError("prepared export plan content does not match requested version")
    if canonical_json(reconciliation.model_dump(mode="json")) != canonical_json(
        plan.report.reconciliation.model_dump(mode="json")
    ):
        raise ValueError("prepared export plan carries conflicting reconciliations")
    if canonical_json(
        [item.model_dump(mode="json") for item in plan.review_version.demand_dispositions]
    ) != canonical_json(
        [item.model_dump(mode="json") for item in reconciliation.dispositions]
    ):
        raise ValueError("prepared export plan disposition manifest drifted")
    if plan.report.publication_state != reconciliation.publication_state:
        raise ValueError("prepared export plan publication state drifted")
    if plan.report.changed_cell_count != reconciliation.changed_cell_count:
        raise ValueError("prepared export plan changed-cell count drifted")
    if plan.report.pending_audit_count != reconciliation.pending_audit_counts.get(
        "total", 0
    ):
        raise ValueError("prepared export plan pending-audit count drifted")
    if (
        plan.report.pending_blocking_audit_count
        != reconciliation.pending_audit_counts.get("blocking:true", 0)
    ):
        raise ValueError("prepared export plan blocking-audit count drifted")
    if plan.report.needs_review_count != reconciliation.needs_review:
        raise ValueError("prepared export plan needs-review count drifted")
    if len(plan.report.validator_violations) != reconciliation.hard_violation_count:
        raise ValueError("prepared export plan validator count drifted")
    if len(plan.report.export_failures) != reconciliation.export_failure_count:
        raise ValueError("prepared export plan failure count drifted")
    if len(plan.placements) != reconciliation.placement_count:
        raise ValueError("prepared export plan placement count drifted")

    entries = {entry.id: entry for entry in plan.review_version.entries}
    audits = {audit.id: audit for audit in plan.review_version.audit_items}
    dispositions = {
        item.demand_id: item for item in reconciliation.dispositions
    }
    manifests = {placement.entry_id: placement for placement in plan.report.placements}
    if (
        len(plan.report.placements) != len(plan.placements)
        or len(manifests) != len(plan.placements)
    ):
        raise ValueError("prepared export placement manifest is incomplete")
    reciprocal_audits = _audit_ids_by_entry(plan.review_version)
    expected_cells: set[str] = set()
    for target in plan.placements:
        entry = entries.get(target.entry.id)
        manifest = manifests.get(target.entry.id)
        if entry is None or manifest is None:
            raise ValueError("prepared export placement references a missing entry")
        if canonical_json(target.entry.model_dump(mode="json")) != canonical_json(
            entry.model_dump(mode="json")
        ):
            raise ValueError("prepared export target entry payload drifted")
        disposition = dispositions.get(entry.demand_id or "")
        if (
            not entry.demand_id
            or disposition is None
            or disposition.entry_id != entry.id
            or manifest.demand_id != entry.demand_id
            or manifest.disposition != disposition.disposition
        ):
            raise ValueError("prepared export placement has no canonical disposition")
        if manifest.version_id != plan.review_version.id:
            raise ValueError("prepared export placement has a stale version ID")
        canonical_audit_ids = reciprocal_audits.get(entry.id, [])
        if sorted(entry.audit_ids) != canonical_audit_ids:
            raise ValueError("prepared export entry audit links are non-reciprocal")
        if target.audit_ids != canonical_audit_ids:
            raise ValueError("prepared export target audit links drifted")
        if entry.status == EntryStatus.NEEDS_REVIEW and not canonical_audit_ids:
            raise ValueError("prepared needs-review placement has no canonical audit")
        if not target.occupied_sessions or target.occupied_sessions != sorted(
            set(target.occupied_sessions)
        ):
            raise ValueError("prepared export occupied sessions drifted")
        if target.session != target.occupied_sessions[0]:
            raise ValueError("prepared export write session drifted")
        if target.assignment_row < 1 or target.col < 1 or "!" not in target.assignment_ref:
            raise ValueError("prepared export assignment target is invalid")
        sheet_name, assignment_coordinate = target.assignment_ref.rsplit("!", 1)
        if assignment_coordinate != _cell_ref(target.assignment_row, target.col):
            raise ValueError("prepared export assignment cell drifted")
        if target.detail_row is None:
            if target.detail_ref is not None:
                raise ValueError("prepared export detail cell drifted")
        else:
            if target.detail_ref != (
                f"{sheet_name}!{_cell_ref(target.detail_row, target.col)}"
            ):
                raise ValueError("prepared export detail cell drifted")
        expected_manifest = ExportPlacement(
            demand_id=entry.demand_id,
            entry_id=entry.id,
            disposition=disposition.disposition,
            version_id=plan.review_version.id,
            status=entry.status.value,
            worker_name=entry.worker_name or entry.worker_id or "",
            schedule_date=entry.schedule_date.isoformat(),
            period=entry.period.value,
            service_code=entry.service_code.value,
            target=_entry_target(entry),
            assignment_cell=target.assignment_ref,
            detail_cell=target.detail_ref,
            audit_ids=canonical_audit_ids,
            data_gap_ids=sorted(entry.data_gap_ids),
            source_evidence_ids=sorted(
                evidence.id for evidence in entry.source_evidence
            ),
            source_refs=_entry_source_refs(entry),
        )
        if canonical_json(manifest.model_dump(mode="json")) != canonical_json(
            expected_manifest.model_dump(mode="json")
        ):
            raise ValueError("prepared export placement payload drifted")
        marker_kind = (
            "review" if entry.status == EntryStatus.NEEDS_REVIEW else "changed"
        )
        expected_notes = {
            target.assignment_ref: _placement_note(
                entry, canonical_audit_ids, audits, detail=False
            ),
        }
        if target.detail_ref:
            expected_notes[target.detail_ref] = _placement_note(
                entry, canonical_audit_ids, audits, detail=True
            )
        for ref, expected_note in expected_notes.items():
            if ref in expected_cells:
                raise ValueError("prepared export cells overlap")
            expected_cells.add(ref)
            marker = plan.changed_cells.get(ref)
            if (
                marker is None
                or marker.marker_kind != marker_kind
                or marker.note != expected_note
            ):
                raise ValueError("prepared export changed-cell marker drifted")
    if expected_cells != set(plan.changed_cells):
        raise ValueError("prepared export changed-cell manifest drifted")
    if not plan.integrity_hash or plan.integrity_hash != _plan_integrity_hash(plan):
        raise ValueError("prepared export plan integrity hash is invalid")


def _header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="305496")


def _reset_rc_sheets(wb: Workbook) -> None:
    for name in RC_SHEETS:
        if name in wb.sheetnames:
            del wb[name]


def _write_rows(ws, headers: list[str], rows: Iterable[list[object]]) -> None:
    for col, header in enumerate(headers, start=1):
        _header(ws.cell(1, col, header))
    for row_idx, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            safe_value = _formula_safe(value) if isinstance(value, str) else value
            ws.cell(row_idx, col, safe_value)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 22
    ws.freeze_panes = "A2"


def _attach_reconciliation_manifest(
    ws,
    report: GeneratedDivisionExportReport | None,
    *,
    start_col: int,
) -> None:
    """Attach the exact canonical report without recomputing sheet totals.

    Excel cells cap text at 32,767 characters.  The full disposition manifest
    can exceed that, so every RC sheet receives the same deterministic 30k
    chunks in two hidden columns.  Concatenating the ordered chunks reproduces
    the API's canonical reconciliation payload byte-for-byte.
    """

    if report is None:
        return
    from openpyxl.utils import get_column_letter

    payload = canonical_json(report.reconciliation.model_dump(mode="json"))
    chunks = [payload[index:index + 30_000] for index in range(0, len(payload), 30_000)]
    ws.cell(1, start_col, "RC_reconciliation_key")
    ws.cell(1, start_col + 1, "RC_reconciliation_value")
    ws.cell(2, start_col, "version_id")
    ws.cell(2, start_col + 1, report.reconciliation.version_id or "")
    ws.cell(3, start_col, "content_hash")
    ws.cell(3, start_col + 1, report.reconciliation.content_hash or "")
    for index, chunk in enumerate(chunks, start=1):
        ws.cell(index + 3, start_col, f"reconciliation_json_{index:04d}")
        ws.cell(index + 3, start_col + 1, chunk)
    ws.column_dimensions[get_column_letter(start_col)].hidden = True
    ws.column_dimensions[get_column_letter(start_col + 1)].hidden = True


def _entry_label(entry) -> str:
    if entry is None:
        return ""
    target = entry.elder_name or entry.center or entry.route or entry.destination or ""
    return (
        f"{entry.schedule_date} {entry.period.value} "
        f"{entry.service_code.value}:{target} -> {entry.worker_name or entry.worker_id or '未分配'}"
    )


def _reconciliation_count_rows(
    report: GeneratedDivisionExportReport,
) -> list[list[object]]:
    reconciliation = report.reconciliation
    return [
        ["reconciliation.version_id", reconciliation.version_id or ""],
        ["reconciliation.content_hash", reconciliation.content_hash or ""],
        ["reconciliation.weekly_demand_total", reconciliation.weekly_demand_total],
        ["reconciliation.scheduled", reconciliation.scheduled],
        ["reconciliation.needs_review", reconciliation.needs_review],
        ["reconciliation.unassigned", reconciliation.unassigned],
        ["reconciliation.confirmed_cancelled", reconciliation.confirmed_cancelled],
        ["reconciliation.suppressed_with_audit", reconciliation.suppressed_with_audit],
        ["reconciliation.pending_audit_total",
         reconciliation.pending_audit_counts.get("total", 0)],
        ["reconciliation.decided_audit_total",
         reconciliation.decided_audit_counts.get("total", 0)],
        ["reconciliation.placement_count", reconciliation.placement_count],
        ["reconciliation.changed_cell_count", reconciliation.changed_cell_count],
        ["reconciliation.hard_violation_count", reconciliation.hard_violation_count],
        ["reconciliation.export_failure_count", reconciliation.export_failure_count],
        ["reconciliation.error_count", len(reconciliation.errors)],
        ["reconciliation.publication_state", reconciliation.publication_state],
    ]


def _add_summary_sheet(
    wb: Workbook,
    version: ScheduleVersion | None,
    *,
    source_summary: dict[str, object] | None = None,
    report: GeneratedDivisionExportReport | None = None,
) -> None:
    ws = wb.create_sheet("RC_變更摘要")
    rows = []
    if source_summary:
        for key, value in source_summary.items():
            rows.append([key, value])
    if version is not None:
        for key, value in sorted(version.summary.items()):
            rows.append([key, value])
    if report is not None:
        rows.extend([
            ["發放狀態", report.publication_label],
            ["可寫入審核草稿主表", "是" if report.review_export_allowed else "否"],
            ["阻塞原因", "；".join(report.export_block_reasons)],
            ["主表格寫入格數", report.changed_cell_count],
            ["未分配/落格失敗項目", len(report.unassigned_items)],
            ["待審項目", report.needs_review_count],
            ["待處理阻塞審核", report.pending_blocking_audit_count],
        ])
        rows.extend(_reconciliation_count_rows(report))
    rows.append(["generated_at", datetime.now().isoformat(timespec="seconds")])
    rows.append(["source", "RosterCopiilot Phase 1 NGO-format exporter"])
    _write_rows(ws, ["項目", "值"], rows)
    _attach_reconciliation_manifest(ws, report, start_col=4)


def _add_review_sheet(
    wb: Workbook,
    version: ScheduleVersion | None,
    *,
    report: GeneratedDivisionExportReport | None = None,
) -> None:
    ws = wb.create_sheet("RC_審核")
    if version is None:
        rows = [["", "", "", "", "", "", "", "", "", "", "", "", "no schedule version supplied"]]
    else:
        order = {"high": 0, "warning": 1, "info": 2}
        items = sorted(
            version.audit_items,
            key=lambda a: (a.status != AuditStatus.PENDING, not a.blocking,
                           order[a.severity.value], a.id),
        )
        rows = [[
            item.id,
            version.id,
            item.status.value,
            "是" if item.blocking else "",
            item.severity.value,
            item.kind.value,
            item.reason,
            _reason_lines(item.reasons),
            ", ".join(item.demand_ids),
            ", ".join(item.entry_ids),
            ", ".join(item.data_gap_ids),
            ", ".join(item.evidence_refs),
            _entry_label(item.original_entry),
            _entry_label(item.suggested_entry),
            item.human_note or "",
        ] for item in items]
    _write_rows(
        ws,
        [
            "ID", "版本", "狀態", "阻塞", "風險", "類型", "原因", "結構化理由",
            "需求ID", "相關項目", "資料缺口ID", "來源證據ID",
            "原安排", "建議安排", "人工備註",
        ],
        rows,
    )
    _attach_reconciliation_manifest(ws, report, start_col=18)


def _add_unassigned_sheet(
    wb: Workbook,
    version: ScheduleVersion | None,
    *,
    unassigned_items: list[ExportUnassignedItem] | None = None,
    extra_rows: list[list[object]] | None = None,
    report: GeneratedDivisionExportReport | None = None,
) -> None:
    ws = wb.create_sheet("RC_未分配")
    rows: list[list[object]] = []
    if unassigned_items is not None:
        for item in unassigned_items:
            rows.append([
                item.id,
                item.demand_id or "",
                item.entry_id or "",
                item.audit_id or "",
                item.disposition,
                item.schedule_date,
                item.period,
                item.service_code,
                item.target,
                item.code,
                item.message,
                item.source_ref,
                item.next_action,
            ])
    elif version is not None:
        audit_by_entry = _audit_ids_by_entry(version)
        for entry in version.entries:
            if entry.status != EntryStatus.UNASSIGNED:
                continue
            reason = _primary_reason(entry.review_reasons)
            audit_ids = audit_by_entry.get(entry.id, [])
            rows.append([
                entry.id,
                entry.demand_id or "",
                entry.id,
                ", ".join(audit_ids),
                "unassigned",
                entry.schedule_date.isoformat(),
                entry.period.value,
                entry.service_code.value,
                entry.elder_name or entry.center or entry.route or "",
                reason.code.value if reason else ReviewReasonCode.NO_QUALIFIED_WORKER.value,
                reason.message if reason else entry.explanation or "未能自動分配",
                _entry_source_ref(entry),
                "主管覆核後人工分配或修正資料",
            ])
    if extra_rows:
        rows.extend(extra_rows)
    _write_rows(
        ws,
        ["ID", "需求ID", "項目ID", "審核ID", "Disposition", "日期", "時段",
         "服務", "對象", "原因代碼", "原因說明", "來源", "下一步"],
        rows,
    )
    _attach_reconciliation_manifest(ws, report, start_col=16)


def _add_meta_sheet(
    wb: Workbook,
    *,
    template_path: Path,
    version: ScheduleVersion | None,
    changed_cells: dict[str, object],
    report: GeneratedDivisionExportReport | None = None,
) -> None:
    ws = wb.create_sheet("RC_meta")
    rows = [
        ["template_path", str(template_path)],
        ["version_id", version.id if version else ""],
        ["version_kind", version.kind.value if version else ""],
        ["changed_cell_count", len(changed_cells)],
        ["publication_state", report.publication_state if report else ""],
        ["publication_label", report.publication_label if report else ""],
        ["review_export_allowed", report.review_export_allowed if report else ""],
        ["export_failure_count", len(report.export_failures) if report else 0],
        ["unassigned_review_count", len(report.unassigned_items) if report else ""],
        ["pending_blocking_audit_count", report.pending_blocking_audit_count if report else ""],
        ["business_fill_policy", "changed cells use additive border/comment; fill colors are preserved"],
        ["business_comment_policy", "existing business comments are retained; RosterCopiilot notes are appended once"],
        ["assignment_export_policy", "validator or placement preflight failures block staff-facing grid mutation"],
    ]
    if report is not None:
        rows.extend(_reconciliation_count_rows(report))
    _write_rows(ws, ["key", "value"], rows)
    _attach_reconciliation_manifest(ws, report, start_col=4)


def _mark_changed_cells(wb: Workbook, changed_cells: dict[str, str | CellMarker]) -> None:
    for ref, marker in changed_cells.items():
        if "!" not in ref:
            continue
        sheet_name, coordinate = ref.split("!", 1)
        if sheet_name not in wb.sheetnames:
            continue
        cell = wb[sheet_name][coordinate]
        note = marker.note if isinstance(marker, CellMarker) else f"RosterCopiilot: {marker}"
        _add_marker_border(cell)
        _append_rc_comment(cell, note)


def build_ngo_division_workbook(
    *,
    template_path: Path,
    dataset: MockDataset | None = None,
    version: ScheduleVersion | None = None,
    changed_cells: dict[str, str] | None = None,
) -> Workbook:
    wb = load_workbook(template_path)
    changed = changed_cells or {}
    _reset_rc_sheets(wb)
    _mark_changed_cells(wb, changed)
    _add_summary_sheet(wb, version)
    _add_review_sheet(wb, version)
    _add_unassigned_sheet(wb, version)
    _add_meta_sheet(wb, template_path=template_path, version=version,
                    changed_cells=changed)
    return wb


def build_generated_division_roster_workbook(
    *,
    template_path: Path,
    division_layout: DivisionImportResult,
    dataset: MockDataset | None = None,
    version: ScheduleVersion,
    generated: GeneratedDemands | None = None,
    prepared_plan: GeneratedDivisionExportPlan | None = None,
    source_summary: dict[str, object] | None = None,
) -> Workbook:
    """Write a generated weekly roster back into the NGO division template.

    The writer preserves workbook layout and cell styles. It clears only the
    worker assignment/detail cells in the ``恆常服務`` schedule grid, writes
    generated entries into safe worker/day/period slots, and records overflow in
    ``RC_未分配``. Changed cells are marked with comments/borders rather than
    replacing business fill colours.
    """

    if prepared_plan is None:
        if generated is None:
            raise ValueError("generated demands are required for fresh export preflight")
        plan = prepare_generated_division_roster_export(
            division_layout=division_layout,
            dataset=dataset,
            version=version,
            generated=generated,
        )
    else:
        plan = prepared_plan
        validate_prepared_division_export_plan(plan, version)
    if not plan.report.review_export_allowed:
        raise ExportPreflightError(plan.report)

    wb = load_workbook(template_path)
    sheet_name = division_layout.sheet_name
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"template sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    _reset_rc_sheets(wb)
    changed_cells = plan.changed_cells

    _clear_schedule_grid(ws, division_layout)
    for placement in plan.placements:
        entry = placement.entry
        assignment_row = placement.assignment_row
        detail_row = placement.detail_row
        col = placement.col
        assignment_cell = ws.cell(assignment_row, col)
        detail_cell = ws.cell(detail_row, col) if detail_row else None
        assignment_value, detail_value = render_export_placement_values(entry)
        assignment_cell.value = assignment_value
        if detail_cell is not None:
            detail_cell.value = detail_value

    _refresh_counter_columns(ws, division_layout)
    _mark_changed_cells(wb, changed_cells)
    _add_summary_sheet(
        wb,
        plan.review_version,
        source_summary={
            **(source_summary or {}),
            "寫入主表格數": len(changed_cells),
            "未能落格項目": len(plan.report.export_failures),
            "發放狀態": plan.report.publication_label,
        },
        report=plan.report,
    )
    _add_review_sheet(wb, plan.review_version, report=plan.report)
    _add_unassigned_sheet(
        wb,
        plan.review_version,
        unassigned_items=plan.report.unassigned_items,
        report=plan.report,
    )
    _add_meta_sheet(
        wb,
        template_path=template_path,
        version=plan.review_version,
        changed_cells=changed_cells,
        report=plan.report,
    )
    return wb


def prepare_generated_division_roster_export(
    *,
    division_layout: DivisionImportResult,
    dataset: MockDataset | None,
    version: ScheduleVersion,
    generated: GeneratedDemands,
) -> GeneratedDivisionExportPlan:
    """Preflight a generated workbook export without mutating the template.

    This is the safety boundary for staff-facing grid writes. It validates the
    accepted roster, builds the exact cell manifest, reconciles placement
    failures into audit/unassigned rows, and computes publication state before
    the workbook is opened or cleared.
    """

    if generated is None:
        raise ValueError("generated demands are required for canonical reconciliation")

    # Export works on an immutable child copy.  The scheduler version remains
    # untouched; every export-derived fact is finalized through the same
    # canonical provenance/reconciliation functions as the scheduler.
    review_version = version.model_copy(deep=True)
    finalize_version_provenance(review_version, generated)

    worker_columns = _worker_column_map(division_layout, dataset)
    slot_rows = _slot_row_map(division_layout)
    failures: list[ExportPlacementFailure] = []
    placements: list[PlacementTarget] = []
    failure_audits: list[tuple[ExportPlacementFailure, AuditItem]] = []

    validator_violations: list[HardViolation] = []
    if dataset is None:
        failures.append(ExportPlacementFailure(
            id="export-failure-001",
            code="dataset_missing",
            message="缺少排班資料集，無法在匯出前重新驗證硬性規則。",
        ))
    else:
        validator_violations = validate_entries(
            dataset,
            review_version.entries,
            _leaves_from_trigger_events(review_version),
        )
        for violation in validator_violations:
            entry = review_version.entry_by_id(violation.entry_id)
            failures.append(_failure_from_validator_violation(
                len(failures) + 1,
                violation,
                entry,
            ))

    invalid_entry_ids = {
        violation.entry_id for violation in validator_violations
    }
    occupied: dict[tuple[int, int, str, int], str] = {}
    for entry in review_version.entries:
        if entry.status not in (EntryStatus.SCHEDULED, EntryStatus.NEEDS_REVIEW):
            continue
        if dataset is None or entry.id in invalid_entry_ids:
            continue
        failure = _entry_preflight_failure(
            entry,
            worker_columns,
            slot_rows,
            occupied,
            len(failures) + 1,
        )
        if failure is not None:
            failures.append(failure)
            continue

        col = _entry_worker_column(entry, worker_columns)
        rows = slot_rows[(entry.weekday, entry.period.value)]
        occupied_sessions = _required_sessions(entry, rows)
        assert col is not None and occupied_sessions
        write_session = occupied_sessions[0]
        assignment_row, detail_row = rows[write_session]

        for session in occupied_sessions:
            occupied[(col, entry.weekday, entry.period.value, session)] = entry.id
        placements.append(PlacementTarget(
            entry=entry,
            assignment_row=assignment_row,
            detail_row=detail_row,
            col=col,
            session=write_session,
            occupied_sessions=occupied_sessions,
            assignment_ref=f"{division_layout.sheet_name}!{_cell_ref(assignment_row, col)}",
            detail_ref=(f"{division_layout.sheet_name}!{_cell_ref(detail_row, col)}"
                        if detail_row else None),
            audit_ids=[],
        ))

    for failure in failures:
        audit = _audit_from_failure(
            f"pending-export-failure-{failure.id}",
            failure,
            review_version,
        )
        review_version.audit_items.append(audit)
        failure_audits.append((failure, audit))

    # Canonical finalization owns audit IDs, deduplication, and reciprocal
    # links.  Exporter-local IDs must never leak into the plan or workbook.
    finalize_version_provenance(review_version, generated)
    for failure, audit in failure_audits:
        failure.audit_id = audit.id

    audit_by_entry = _audit_ids_by_entry(review_version)
    audit_by_id = {audit.id: audit for audit in review_version.audit_items}
    for placement in placements:
        placement.audit_ids = list(audit_by_entry.get(placement.entry.id, []))

    tentative_changed_cells = _changed_cells_for_placements(placements, audit_by_id)
    preliminary_reconciliation = reconcile_weekly_demands(
        review_version,
        generated,
        hard_violation_count=len(validator_violations),
        export_failure_count=len(failures),
        placement_count=len(placements),
        changed_cell_count=len(tentative_changed_cells),
    )
    disposition_by_entry = {
        disposition.entry_id: disposition
        for disposition in preliminary_reconciliation.dispositions
        if disposition.entry_id is not None
    }
    # A mapped cell is not an export placement until it has one canonical
    # terminal disposition tied to the same entry.  Invalid conservation may
    # still return a diagnostic plan, but it never leaks orphan cell writes.
    placements = [
        placement for placement in placements
        if placement.entry.id in disposition_by_entry
        and disposition_by_entry[placement.entry.id].disposition
        in {"scheduled", "needs_review"}
    ]
    changed_cells = _changed_cells_for_placements(placements, audit_by_id)
    reconciliation = reconcile_weekly_demands(
        review_version,
        generated,
        hard_violation_count=len(validator_violations),
        export_failure_count=len(failures),
        placement_count=len(placements),
        changed_cell_count=len(changed_cells),
    )
    unassigned_items = _unassigned_items(
        review_version,
        failures,
        reconciliation,
    )
    placement_models = [
        ExportPlacement(
            demand_id=p.entry.demand_id or "",
            entry_id=p.entry.id,
            disposition=disposition_by_entry[p.entry.id].disposition,
            version_id=review_version.id,
            status=p.entry.status.value,
            worker_name=p.entry.worker_name or p.entry.worker_id or "",
            schedule_date=p.entry.schedule_date.isoformat(),
            period=p.entry.period.value,
            service_code=p.entry.service_code.value,
            target=_entry_target(p.entry),
            assignment_cell=p.assignment_ref,
            detail_cell=p.detail_ref,
            audit_ids=p.audit_ids,
            data_gap_ids=sorted(p.entry.data_gap_ids),
            source_evidence_ids=sorted(
                evidence.id for evidence in p.entry.source_evidence
            ),
            source_refs=_entry_source_refs(p.entry),
        )
        for p in placements
    ]

    pending_audit_count = reconciliation.pending_audit_counts.get("total", 0)
    pending_blocking_count = reconciliation.pending_audit_counts.get(
        "blocking:true", 0
    )
    publication_state = reconciliation.publication_state
    block_reasons = _export_block_reasons(
        validator_violations=validator_violations,
        failures=failures,
        reconciliation=reconciliation,
    )
    if publication_state == "blocked" and not block_reasons:
        if pending_blocking_count:
            block_reasons.append(f"仍有 {pending_blocking_count} 項阻塞審核未處理")
        if reconciliation.unassigned:
            block_reasons.append(f"仍有 {reconciliation.unassigned} 項未分配或需人工處理")

    report = GeneratedDivisionExportReport(
        reconciliation=reconciliation,
        publication_state=publication_state,
        publication_label=_publication_label(publication_state),
        review_export_allowed=not failures and not reconciliation.errors,
        export_block_reasons=block_reasons,
        validator_violations=validator_violations,
        export_failures=failures,
        placements=placement_models,
        unassigned_items=unassigned_items,
        pending_audit_count=pending_audit_count,
        pending_blocking_audit_count=pending_blocking_count,
        needs_review_count=reconciliation.needs_review,
        changed_cell_count=reconciliation.changed_cell_count,
    )
    plan = GeneratedDivisionExportPlan(
        report=report,
        review_version=review_version,
        placements=placements,
        changed_cells=changed_cells,
    )
    plan.integrity_hash = _plan_integrity_hash(plan)
    # Fresh preflight returns a structured fail-closed report for malformed
    # lineage.  Strict plan integrity is required before any reusable plan can
    # mutate a workbook; valid review drafts must pass it immediately here.
    if plan.report.review_export_allowed:
        validate_prepared_division_export_plan(plan, review_version)
    return plan


def save_ngo_division_workbook(
    *,
    template_path: Path,
    dataset: MockDataset | None = None,
    version: ScheduleVersion | None = None,
    output_dir: Path | None = None,
    changed_cells: dict[str, str] | None = None,
) -> Path:
    if output_dir is None:
        raw_dir = os.getenv("ROSTER_EXPORT_DIR")
        output_dir = (
            Path(raw_dir)
            if raw_dir
            else Path(__file__).resolve().parents[3] / "data" / "exports"
        )
        if not output_dir.is_absolute():
            output_dir = Path(__file__).resolve().parents[3] / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    version_id = version.id if version else "no_version"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"ngo_division_{version_id}_{ts}.xlsx"
    build_ngo_division_workbook(
        template_path=template_path,
        dataset=dataset,
        version=version,
        changed_cells=changed_cells,
    ).save(path)
    return path


def save_generated_division_roster_workbook(
    *,
    template_path: Path,
    division_layout: DivisionImportResult,
    dataset: MockDataset | None = None,
    version: ScheduleVersion,
    generated: GeneratedDemands | None = None,
    prepared_plan: GeneratedDivisionExportPlan | None = None,
    output_dir: Path | None = None,
    source_summary: dict[str, object] | None = None,
) -> Path:
    if output_dir is None:
        raw_dir = os.getenv("ROSTER_EXPORT_DIR")
        output_dir = (
            Path(raw_dir)
            if raw_dir
            else Path(__file__).resolve().parents[3] / "data" / "exports"
        )
        if not output_dir.is_absolute():
            output_dir = Path(__file__).resolve().parents[3] / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    export_version = prepared_plan.review_version if prepared_plan else version
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"照顧員工作分工表_審核草稿_{export_version.id}_{ts}.xlsx"
    build_generated_division_roster_workbook(
        template_path=template_path,
        division_layout=division_layout,
        dataset=dataset,
        version=version,
        generated=generated,
        prepared_plan=prepared_plan,
        source_summary=source_summary,
    ).save(path)
    return path


# --------------------------------------------------------------------------
# Export preflight / RC-sheet helpers
# --------------------------------------------------------------------------


def _entry_preflight_failure(
    entry: ScheduleEntry,
    worker_columns: dict[str, int],
    slot_rows: dict[tuple[int, str], dict[int, tuple[int, int | None]]],
    occupied: dict[tuple[int, int, str, int], str],
    index: int,
) -> ExportPlacementFailure | None:
    if not entry.demand_id:
        return _entry_failure(
            index,
            entry,
            "demand_link_missing",
            "排班項目沒有 canonical demand ID，已停止寫入主表。",
        )
    if not entry.source_evidence:
        return _entry_failure(
            index,
            entry,
            "source_evidence_missing",
            "排班項目沒有 structured source evidence，已停止寫入主表。",
        )
    hard_bypassed = "supervisor_hard_bypass" in entry.constraint_flags
    if entry.status == EntryStatus.SCHEDULED and not hard_bypassed and (
        entry.review_reasons or entry.constraint_flags
    ):
        return _entry_failure(
            index,
            entry,
            "scheduled_entry_has_review_state",
            "已編排項目仍帶有審核理由或未核實標記，不能以普通格位寫入主表。",
        )
    if not entry.worker_id and not entry.worker_name:
        return _entry_failure(
            index,
            entry,
            "worker_missing",
            "排班項目沒有同工，不能寫入原分工表同工欄。",
        )
    col = _entry_worker_column(entry, worker_columns)
    if col is None:
        return _entry_failure(
            index,
            entry,
            "worker_unmapped",
            "未能對應原分工表的同工欄。",
        )
    rows = slot_rows.get((entry.weekday, entry.period.value))
    if rows is None:
        return _entry_failure(
            index,
            entry,
            "slot_unmapped",
            "未能對應原分工表的星期/時段格。",
        )
    required = _required_sessions(entry, rows)
    if not required:
        return _entry_failure(
            index,
            entry,
            "unsupported_cell_grammar",
            "排班項目的 session 不能安全對應原表格兩段式格位。",
        )
    collisions = [
        occupied[(col, entry.weekday, entry.period.value, session)]
        for session in required
        if (col, entry.weekday, entry.period.value, session) in occupied
    ]
    if collisions:
        return _entry_failure(
            index,
            entry,
            "cell_collision",
            f"目標格位已由 {', '.join(sorted(set(collisions)))} 佔用，已停止寫入避免覆蓋。",
        )
    return None


def _required_sessions(
    entry: ScheduleEntry,
    rows: dict[int, tuple[int, int | None]],
) -> list[int] | None:
    if entry.session_index is None:
        sessions = sorted(rows)
        return sessions if len(sessions) >= 2 else None
    if entry.session_index not in rows:
        return None
    return [entry.session_index]


def _entry_failure(
    index: int,
    entry: ScheduleEntry,
    code: str,
    message: str,
) -> ExportPlacementFailure:
    return ExportPlacementFailure(
        id=f"export-failure-{index:03d}",
        demand_id=entry.demand_id,
        entry_id=entry.id,
        code=code,
        message=message,
        source_ref=_entry_source_ref(entry),
        schedule_date=entry.schedule_date.isoformat(),
        period=entry.period.value,
        service_code=entry.service_code.value,
        worker_name=entry.worker_name or entry.worker_id,
        target=_entry_target(entry),
    )


def _failure_from_validator_violation(
    index: int,
    violation: HardViolation,
    entry: ScheduleEntry | None,
) -> ExportPlacementFailure:
    return ExportPlacementFailure(
        id=f"export-failure-{index:03d}",
        demand_id=entry.demand_id if entry else None,
        entry_id=violation.entry_id,
        code=f"validator_{violation.code.value}",
        message=f"匯出前硬性規則驗證失敗：{violation.message}",
        proposed_disposition="停止寫入主表；先修正排班或資料後再匯出",
        source_ref=_entry_source_ref(entry) if entry else "",
        schedule_date=entry.schedule_date.isoformat() if entry else None,
        period=entry.period.value if entry else None,
        service_code=entry.service_code.value if entry else None,
        worker_name=(entry.worker_name or entry.worker_id) if entry else None,
        target=_entry_target(entry) if entry else None,
    )


def _audit_from_failure(
    audit_id: str,
    failure: ExportPlacementFailure,
    version: ScheduleVersion,
) -> AuditItem:
    entry = version.entry_by_id(failure.entry_id) if failure.entry_id else None
    code = (
        _validator_reason_code(failure.code)
        if failure.code.startswith("validator_")
        else ReviewReasonCode.EXPORT_PLACEMENT_FAILURE
    )
    return AuditItem(
        id=audit_id,
        kind=AuditKind.UNASSIGNED_TASK,
        severity=Severity.HIGH,
        blocking=True,
        reason=failure.message,
        reasons=[ManualReviewReason(
            code=code,
            message=failure.message,
            params={
                "entry_id": failure.entry_id or "",
                "failure_code": failure.code,
                "source_ref": failure.source_ref or "",
            },
            rule_ref="excel_io_contract.md",
        )],
        original_entry=entry,
        demand_ids=[failure.demand_id] if failure.demand_id else [],
        entry_ids=[failure.entry_id] if failure.entry_id else [],
        data_gap_ids=sorted(entry.data_gap_ids) if entry else [],
        evidence_refs=sorted(
            evidence.id for evidence in entry.source_evidence
        ) if entry else [],
        human_note=failure.proposed_disposition,
    )


def _validator_reason_code(code: str) -> ReviewReasonCode:
    raw = code.replace("validator_", "", 1)
    try:
        return ReviewReasonCode(raw)
    except ValueError:
        return ReviewReasonCode.EXPORT_PLACEMENT_FAILURE


def _unassigned_items(
    version: ScheduleVersion,
    failures: list[ExportPlacementFailure],
    reconciliation: DemandReconciliationReport,
) -> list[ExportUnassignedItem]:
    items: list[ExportUnassignedItem] = []
    audit_by_id = {audit.id: audit for audit in version.audit_items}
    disposition_by_demand = {
        item.demand_id: item for item in reconciliation.dispositions
    }
    for disposition in reconciliation.dispositions:
        if disposition.disposition != "unassigned" or not disposition.entry_id:
            continue
        entry = version.entry_by_id(disposition.entry_id)
        if entry is None:
            continue
        reason = _primary_reason(entry.review_reasons)
        audit_ids = list(disposition.audit_ids)
        terminal_audits = [
            audit_id for audit_id in audit_ids
            if audit_id in audit_by_id
            and audit_by_id[audit_id].kind in {
                AuditKind.UNASSIGNED_TASK,
                AuditKind.DUTY_UNDER_COVERAGE,
            }
            and audit_by_id[audit_id].blocking
        ]
        message = reason.message if reason else entry.explanation or "未能自動分配"
        items.append(ExportUnassignedItem(
            id=entry.id,
            demand_id=disposition.demand_id,
            entry_id=entry.id,
            audit_id=terminal_audits[0] if len(terminal_audits) == 1 else None,
            audit_ids=audit_ids,
            schedule_date=entry.schedule_date.isoformat(),
            period=entry.period.value,
            service_code=entry.service_code.value,
            target=_entry_target(entry),
            code=reason.code.value if reason else ReviewReasonCode.NO_QUALIFIED_WORKER.value,
            message=message,
            source_ref=_entry_source_ref(entry),
            worker_name=entry.worker_name,
            explanation=entry.explanation or message,
            disposition=disposition.disposition,
        ))
    for failure in failures:
        scheduler_disposition = disposition_by_demand.get(failure.demand_id or "")
        items.append(ExportUnassignedItem(
            id=failure.id,
            demand_id=failure.demand_id,
            entry_id=failure.entry_id,
            audit_id=failure.audit_id,
            audit_ids=[failure.audit_id] if failure.audit_id else [],
            schedule_date=failure.schedule_date or "",
            period=failure.period or "",
            service_code=failure.service_code or "",
            target=failure.target or "",
            code=failure.code,
            message=failure.message,
            source_ref=failure.source_ref or "",
            next_action=failure.proposed_disposition,
            is_export_failure=True,
            worker_name=failure.worker_name,
            explanation=failure.message,
            disposition=(
                scheduler_disposition.disposition
                if scheduler_disposition is not None
                else "export_failure"
            ),
        ))
    return items


def _changed_cells_for_placements(
    placements: list[PlacementTarget],
    audit_by_id: dict[str, AuditItem],
) -> dict[str, CellMarker]:
    changed: dict[str, CellMarker] = {}
    for placement in placements:
        entry = placement.entry
        note = _placement_note(
            entry,
            placement.audit_ids,
            audit_by_id,
            detail=False,
        )
        changed[placement.assignment_ref] = CellMarker(
            note=note,
            marker_kind="review" if entry.status == EntryStatus.NEEDS_REVIEW else "changed",
        )
        if placement.detail_ref:
            changed[placement.detail_ref] = CellMarker(
                note=_placement_note(
                    entry,
                    placement.audit_ids,
                    audit_by_id,
                    detail=True,
                ),
                marker_kind="review" if entry.status == EntryStatus.NEEDS_REVIEW else "changed",
            )
    return changed


def _placement_note(
    entry: ScheduleEntry,
    audit_ids: list[str],
    audit_by_id: dict[str, AuditItem],
    *,
    detail: bool,
) -> str:
    reason = _primary_reason(entry.review_reasons)
    source_ref = _entry_source_ref(entry)
    target = _entry_target(entry)
    if entry.status == EntryStatus.NEEDS_REVIEW:
        audit_reasons = [
            audit_by_id[audit_id].reason for audit_id in audit_ids
            if audit_id in audit_by_id
        ]
        reason_text = "；".join(dict.fromkeys(audit_reasons)) or (
            reason.message if reason else entry.explanation or "需人工確認"
        )
        reason_text = " ".join(reason_text.split())
        evidence_ids = sorted(evidence.id for evidence in entry.source_evidence)
        lines = [
            "RC:待審",
            f"原因: {reason_text}",
            f"審核ID: {', '.join(audit_ids) if audit_ids else '待補'}",
            f"需求ID: {entry.demand_id or '缺失'}",
            f"項目ID: {entry.id}",
            f"來源證據ID: {', '.join(evidence_ids) if evidence_ids else '缺失'}",
            f"格位: {'詳情格' if detail else '排班格'}",
        ]
    else:
        lines = [
            "RosterCopiilot: 自動生成排班草稿",
            f"項目ID: {entry.id}",
            f"服務: {entry.service_code.value} {target}".rstrip(),
        ]
    if source_ref:
        lines.append(f"來源: {source_ref}")
    return "\n".join(lines)


def _publication_label(state: PublicationState) -> str:
    return {
        "blocked": "不可發放",
        "draft": "草稿需審核",
        "ready": "可發放",
    }[state]


def _export_block_reasons(
    *,
    validator_violations: list[HardViolation],
    failures: list[ExportPlacementFailure],
    reconciliation: DemandReconciliationReport,
) -> list[str]:
    reasons: list[str] = []
    if validator_violations:
        reasons.append(f"匯出前硬性規則驗證失敗 {len(validator_violations)} 項")
    placement_failures = [
        failure for failure in failures
        if not failure.code.startswith("validator_")
    ]
    if placement_failures:
        reasons.append(f"主表落格前檢查失敗 {len(placement_failures)} 項")
    if reconciliation.errors:
        codes = sorted({error.code for error in reconciliation.errors})
        reasons.append(
            f"需求守恆或 provenance 連結失敗 {len(reconciliation.errors)} 項"
            f"（{', '.join(codes)}）"
        )
    return reasons


def _leaves_from_trigger_events(version: ScheduleVersion) -> set[tuple[str, object, str]]:
    leaves: set[tuple[str, object, str]] = set()
    for event in version.trigger_events:
        if event.worker_id is None:
            continue
        periods = [event.period.value] if event.period else [Period.AM.value, Period.PM.value]
        for period in periods:
            leaves.add((event.worker_id, event.change_date, period))
    return leaves


def _audit_ids_by_entry(version: ScheduleVersion) -> dict[str, list[str]]:
    valid_audit_ids = {audit.id for audit in version.audit_items}
    out: dict[str, set[str]] = {
        entry.id: set(entry.audit_ids) & valid_audit_ids
        for entry in version.entries
    }
    for audit in version.audit_items:
        for entry_id in audit.entry_ids:
            out.setdefault(entry_id, set()).add(audit.id)
    return {
        entry_id: sorted(audit_ids)
        for entry_id, audit_ids in out.items()
        if audit_ids
    }


def _primary_reason(reasons: list[ManualReviewReason]) -> ManualReviewReason | None:
    return reasons[0] if reasons else None


def _reason_lines(reasons: list[ManualReviewReason]) -> str:
    return "\n".join(
        f"[{reason.code.value}] {reason.message}"
        + (f"（{reason.rule_ref}）" if reason.rule_ref else "")
        for reason in reasons
    )


def _entry_target(entry: ScheduleEntry | None) -> str:
    if entry is None:
        return ""
    return entry.elder_name or entry.center or entry.route or entry.destination or ""


def _entry_source_ref(entry: ScheduleEntry | None) -> str:
    refs = _entry_source_refs(entry)
    return ", ".join(refs)


def _entry_source_refs(entry: ScheduleEntry | None) -> list[str]:
    if entry is None:
        return []
    refs: list[str] = list(entry.source_refs)
    if entry.origin_fixed_service_id:
        refs.append(f"fixed:{entry.origin_fixed_service_id}")
    if entry.origin_escort_request_id:
        refs.append(f"escort:{entry.origin_escort_request_id}")
    refs.extend(evidence.id for evidence in entry.source_evidence)
    return list(dict.fromkeys(refs))


def _cell_ref(row: int, col: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(col)}{row}"


def _add_marker_border(cell) -> None:
    border = copy(cell.border)
    sides = {
        "left": border.left,
        "right": border.right,
        "top": border.top,
        "bottom": border.bottom,
    }
    for name in ("left", "right", "top", "bottom"):
        side = sides[name]
        if side is None or side.style is None:
            setattr(border, name, MARK_SIDE)
            cell.border = border
            return
    # All four sides already carry business formatting; preserve them and rely
    # on the RC comment plus summary sheets for the marker.


def _append_rc_comment(cell, note: str) -> None:
    existing = cell.comment
    author = existing.author if existing is not None else "RosterCopiilot"
    business_text = _business_comment_text(existing.text if existing is not None else "")
    if note.startswith("RC:待審"):
        parts = [part for part in [note.strip(), business_text] if part]
    else:
        parts = [part for part in [business_text, note.strip()] if part]
    cell.comment = Comment("\n\n".join(parts), author or "RosterCopiilot")


def _business_comment_text(text: str) -> str:
    paragraphs = [part.strip() for part in text.split("\n\n") if part.strip()]
    kept = [
        part for part in paragraphs
        if not part.startswith("RosterCopiilot:") and not part.startswith("RC:")
    ]
    return "\n\n".join(kept)


# --------------------------------------------------------------------------
# Main-sheet placement helpers
# --------------------------------------------------------------------------


def _worker_column_map(
    division_layout: DivisionImportResult,
    dataset: MockDataset | None,
) -> dict[str, int]:
    out: dict[str, int] = {}
    for worker in division_layout.workers:
        out[f"name:{worker.display_name}"] = worker.column
        out[f"name:{worker.raw_header}"] = worker.column
    if dataset is not None:
        name_to_col = {
            worker.display_name: worker.column for worker in division_layout.workers
        }
        for employee in dataset.employees:
            col = name_to_col.get(employee.display_name)
            if col is not None:
                out[f"id:{employee.id}"] = col
    return out


def _slot_row_map(
    division_layout: DivisionImportResult,
) -> dict[tuple[int, str], dict[int, tuple[int, int | None]]]:
    out: dict[tuple[int, str], dict[int, tuple[int, int | None]]] = {}
    for day in division_layout.weekday_blocks:
        for period in day.periods:
            sessions: dict[int, tuple[int, int | None]] = {}
            for idx, row in enumerate(period.assignment_rows, start=1):
                detail_row = period.detail_rows[idx - 1] if idx - 1 < len(period.detail_rows) else None
                sessions[idx] = (row, detail_row)
            out[(day.weekday, period.period)] = sessions
    return out


def _clear_schedule_grid(ws, division_layout: DivisionImportResult) -> None:
    worker_cols = [w.column for w in division_layout.workers]
    rows: set[int] = set()
    for day in division_layout.weekday_blocks:
        for period in day.periods:
            rows.update(period.assignment_rows)
            rows.update(period.detail_rows)
            rows.update(period.extra_rows)
    for row in rows:
        for col in worker_cols:
            cell = ws.cell(row, col)
            cell.value = None


def _entry_worker_column(
    entry: ScheduleEntry,
    worker_columns: dict[str, int],
) -> int | None:
    if entry.worker_id and f"id:{entry.worker_id}" in worker_columns:
        return worker_columns[f"id:{entry.worker_id}"]
    if entry.worker_name and f"name:{entry.worker_name}" in worker_columns:
        return worker_columns[f"name:{entry.worker_name}"]
    return None


def _choose_session(
    entry: ScheduleEntry,
    col: int,
    occupied: set[tuple[int, int, str, int]],
    rows: dict[int, tuple[int, int | None]],
) -> int | None:
    preferred = 1 if entry.session_index is None else entry.session_index
    candidates = [preferred] + [idx for idx in sorted(rows) if idx != preferred]
    for session in candidates:
        if session not in rows:
            continue
        key = (col, entry.weekday, entry.period.value, session)
        if key not in occupied:
            return session
    return None


_FORMULA_TRIGGER_PREFIXES = ("=", "+", "-", "@")


def _formula_safe(value: str) -> str:
    """Neutralise spreadsheet formula/CSV injection in workbook-derived text.

    Text taken from uploaded workbooks (elder/destination/route names, notes)
    is written into visible cells. A value beginning with ``= + - @`` — or a
    leading tab/carriage-return — is treated as a formula by openpyxl and by
    Excel/LibreOffice on open. Prefixing a single apostrophe forces the cell to
    remain literal text without altering the visible tokens materially.
    """

    if value and (value[0] in _FORMULA_TRIGGER_PREFIXES or value[0] in ("\t", "\r")):
        return "'" + value
    return value


def _roster_cell_label(entry: ScheduleEntry) -> str:
    code = "Esc" if entry.service_code.value == "ESC" else entry.service_code.value
    if entry.service_code.value in {"AMC", "MRC", "GC"}:
        return code
    target = entry.elder_name or entry.center or entry.route or entry.destination or ""
    if entry.service_code.value == "D" and target:
        return f"D({target})"
    return f"{code}:{target}" if target else code


def _roster_detail_label(entry: ScheduleEntry) -> str:
    parts = []
    if entry.start_time or entry.end_time:
        start = entry.start_time.strftime("%H:%M") if entry.start_time else ""
        end = entry.end_time.strftime("%H:%M") if entry.end_time else ""
        parts.append(f"{start}-{end}".strip("-"))
    if entry.destination:
        parts.append(entry.destination)
    if entry.district:
        parts.append(entry.district)
    if entry.notes:
        parts.append(entry.notes)
    if entry.explanation:
        parts.append(entry.explanation)
    return " ".join(parts)[:250]


def render_export_placement_values(entry: ScheduleEntry) -> tuple[str, str]:
    """Return the exact deterministic values written for one placement.

    The parallel-run evaluator compares the generated API manifest with the
    roster owner's operational workbook.  Keeping that comparison on the
    writer's public rendering seam prevents the evaluator from copying (and
    eventually drifting from) the NGO workbook cell grammar.
    """

    return (
        _formula_safe(_roster_cell_label(entry)),
        _formula_safe(_roster_detail_label(entry)),
    )


def _export_unassigned_row(entry: ScheduleEntry, reason: str) -> list[object]:
    return [
        entry.id,
        entry.schedule_date.isoformat(),
        entry.period.value,
        entry.service_code.value,
        _formula_safe(
            entry.elder_name or entry.center or entry.route or entry.destination or ""
        ),
        _formula_safe(reason),
    ]


def _refresh_counter_columns(ws, division_layout: DivisionImportResult) -> None:
    if len(division_layout.counter_columns) < 3:
        return
    counter_cols = [ws[col + "1"].column for col in division_layout.counter_columns[:3]]
    worker_cols = [w.column for w in division_layout.workers]
    for counter in division_layout.counters:
        row = counter.row
        values = [str(ws.cell(row, col).value or "") for col in worker_cols]
        ero = sum(1 for value in values if value.upper().startswith("E+RO"))
        esc = sum(1 for value in values if value.upper().startswith(("ESC", "ESC:")))
        meal = sum(1 for value in values if value.upper().startswith("D"))
        other = meal if counter.other_label == "D" else esc
        ws.cell(row, counter_cols[0], ero)
        ws.cell(row, counter_cols[1], other)
        ws.cell(row, counter_cols[2], ero + other)
