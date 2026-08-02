"""Template-free division-grid writer for scheduler-produced assignments.

The NGO-format writer (``division_writer``) starts from the real workbook and
only *appends* review sheets — it needs the original Excel file and does not
place assignments into the staff-facing grid. The scheduler fixture, by design,
has no workbook to start from (ENGINEERING_SPEC.md §8).

This module provides the minimal missing path: it renders a ``ScheduleVersion``
into a fresh worker × (weekday / AM-PM / session) grid that mirrors the
division-sheet grammar, so each worker can see what to do when. Review-required
and unassigned work is surfaced additively (marked cells + a dedicated sheet),
never silently dropped (RB-DATA-01).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from ..domain import Employee, EntryStatus, Period, ScheduleEntry, ScheduleVersion

WEEKDAY_LABELS = {1: "一", 2: "二", 3: "三", 4: "四", 5: "五", 6: "六"}
GRID_WEEKDAYS = (1, 2, 3, 4, 5, 6)
GRID_PERIODS = (Period.AM, Period.PM)

_REVIEW_SIDE = Side(style="medium", color="C00000")
_REVIEW_BORDER = Border(left=_REVIEW_SIDE, right=_REVIEW_SIDE,
                        top=_REVIEW_SIDE, bottom=_REVIEW_SIDE)
_PLACED = {EntryStatus.SCHEDULED, EntryStatus.NEEDS_REVIEW}


def entry_label(entry: ScheduleEntry) -> str:
    """The cell grammar an NGO reviewer expects: ``service:target`` (+ marker)."""
    target = (entry.elder_name or entry.center or entry.route
              or entry.destination or "")
    session = "" if entry.session_index is None else f"[{entry.session_index}]"
    label = f"{entry.service_code.value}{session}:{target}".rstrip(":")
    if entry.status == EntryStatus.NEEDS_REVIEW:
        label += "⚠"
    return label


def grid_labels(
    version: ScheduleVersion,
    workers: list[Employee] | None = None,
) -> dict[str, dict[tuple[int, str], list[str]]]:
    """``worker_id → {(weekday, period): [labels]}`` for placed entries.

    ``workers`` may be supplied to include staff with no assignments this week;
    otherwise the grid covers only workers that appear in the roster.
    """
    grid: dict[str, dict[tuple[int, str], list[str]]] = {}
    names: dict[str, str] = {}
    for w in workers or []:
        grid[w.id] = {}
        names[w.id] = w.display_name
    for entry in version.entries:
        if entry.status not in _PLACED or not entry.worker_id:
            continue
        grid.setdefault(entry.worker_id, {})
        names.setdefault(entry.worker_id, entry.worker_name or entry.worker_id)
        key = (entry.weekday, entry.period.value)
        grid[entry.worker_id].setdefault(key, []).append(entry_label(entry))
    grid["__names__"] = names  # carried out-of-band for the workbook renderer
    return grid


def _header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="305496")


def _write_unassigned(ws: Worksheet, version: ScheduleVersion) -> None:
    headers = ["ID", "日期", "時段", "服務", "對象", "原因"]
    for col, head in enumerate(headers, start=1):
        _header(ws.cell(1, col, head))
    row = 2
    for entry in version.entries:
        if entry.status != EntryStatus.UNASSIGNED:
            continue
        ws.cell(row, 1, entry.id)
        ws.cell(row, 2, entry.schedule_date.isoformat())
        ws.cell(row, 3, entry.period.value)
        ws.cell(row, 4, entry.service_code.value)
        ws.cell(row, 5, entry.elder_name or entry.center or entry.route or "")
        ws.cell(row, 6, entry.explanation or "")
        row += 1
    ws.freeze_panes = "A2"


def build_assignment_grid_workbook(
    version: ScheduleVersion,
    workers: list[Employee] | None = None,
) -> Workbook:
    """Render the scheduler draft into a division-style assignment grid."""
    labels = grid_labels(version, workers)
    names = labels.pop("__names__")
    worker_ids = sorted(labels)

    wb = Workbook()
    ws = wb.active
    ws.title = "排班格"

    # Header: worker column, then (weekday × AM/PM).
    _header(ws.cell(1, 1, "同工"))
    columns: list[tuple[int, str]] = []
    for weekday in GRID_WEEKDAYS:
        for period in GRID_PERIODS:
            col = len(columns) + 2
            _header(ws.cell(1, col, f"週{WEEKDAY_LABELS[weekday]} {period.value}"))
            columns.append((weekday, period.value))

    review_notes = _review_targets(version)
    for r, worker_id in enumerate(worker_ids, start=2):
        ws.cell(r, 1, f"{names.get(worker_id, worker_id)}（{worker_id}）")
        cells = labels[worker_id]
        for c, key in enumerate(columns, start=2):
            entry_labels = cells.get(key, [])
            if not entry_labels:
                continue
            cell = ws.cell(r, c, "\n".join(entry_labels))
            note = review_notes.get((worker_id, key[0], key[1]))
            if note:
                cell.border = _REVIEW_BORDER
                cell.comment = Comment(f"RosterCopiilot: {note}", "RosterCopiilot")

    ws.freeze_panes = "B2"
    for col in range(1, len(columns) + 2):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 18

    _write_unassigned(wb.create_sheet("RC_未分配"), version)
    return wb


def _review_targets(version: ScheduleVersion) -> dict[tuple[str, int, str], str]:
    """Cells that need a reviewer's eye, keyed by (worker, weekday, period)."""
    out: dict[tuple[str, int, str], str] = {}
    for entry in version.entries:
        if entry.status == EntryStatus.NEEDS_REVIEW and entry.worker_id:
            out[(entry.worker_id, entry.weekday, entry.period.value)] = (
                entry.explanation or "需人工覆核")
    return out
