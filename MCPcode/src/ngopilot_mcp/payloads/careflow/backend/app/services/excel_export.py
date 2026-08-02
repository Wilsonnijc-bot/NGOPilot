"""Excel 匯出 — 用 openpyxl 寫入 NGO 模板，保留格式。

模板假設：第 1 列為標題列，從第 2 列起寫資料。每欄對應一個 schema key（順序見 HEADER_KEY_MAP）。
若 NGO 後續換真實模板，只需更換 templates/volunteer_visit_template.xlsx 並調整 HEADER_KEY_MAP。
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..config import settings
from ..llm.vision import FIELD_KEYS, VOLUNTEER_FORM_FIELDS

TEMPLATE_PATH = lambda: settings.data_path / "templates" / "volunteer_visit_template.xlsx"

# 寫入順序（模板第 1 列標題的順序）
HEADER_KEYS = [
    "batch_title", "visit_date", "volunteer_team", "volunteer_name",
    "elder_name", "elder_gender", "elder_age", "elder_phone",
    "elder_address", "living_alone", "duration_minutes",
    "mood", "health_concerns", "follow_up_needed", "follow_up_note",
    "_reviewer", "_reviewed_at",
]

HEADER_LABELS = {
    "batch_title": "批次",
    "visit_date": "探訪日期",
    "volunteer_team": "志工隊",
    "volunteer_name": "志工姓名",
    "elder_name": "長者姓名",
    "elder_gender": "性別",
    "elder_age": "年齡",
    "elder_phone": "聯絡電話",
    "elder_address": "地址",
    "living_alone": "獨居",
    "duration_minutes": "探訪時長(分)",
    "mood": "情緒狀態",
    "health_concerns": "健康關注",
    "follow_up_needed": "需要跟進",
    "follow_up_note": "跟進備註",
    "_reviewer": "審查者",
    "_reviewed_at": "審查時間",
}


def ensure_template_exists() -> Path:
    """若模板不存在，現場生成一份合理預設。"""
    path = TEMPLATE_PATH()
    if path.exists():
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "志工探訪紀錄"

    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin = Side(border_style="thin", color="BFDBFE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, key in enumerate(HEADER_KEYS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HEADER_LABELS[key])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 欄寬
    widths = {
        "batch_title": 26, "visit_date": 14, "volunteer_team": 18, "volunteer_name": 14,
        "elder_name": 14, "elder_gender": 8, "elder_age": 8, "elder_phone": 14,
        "elder_address": 32, "living_alone": 10, "duration_minutes": 14,
        "mood": 12, "health_concerns": 26, "follow_up_needed": 12,
        "follow_up_note": 28, "_reviewer": 12, "_reviewed_at": 20,
    }
    for col_idx, key in enumerate(HEADER_KEYS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = widths.get(key, 14)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def export_batch(
    *,
    batch_title: str,
    volunteer_team: str | None,
    rows: Iterable[dict],
    out_path: Path,
) -> Path:
    """rows 是 list of {final_fields..., reviewer, reviewed_at}。

    若使用者上傳了自訂模板，則使用該模板 + 它的欄位映射；否則用內建模板。
    """
    from . import template_store

    info = template_store.get_active()
    template = template_store.active_template_file()
    wb = load_workbook(template)
    ws = wb.active

    start_row = ws.max_row + 1 if ws.max_row > 1 else 2

    body_font = Font(name="Microsoft JhengHei", size=10)
    thin = Side(border_style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # mapping: column_index_str -> schema_key
    mapping: dict[int, str] = {int(k): v for k, v in (info.mapping or {}).items()}
    if not mapping:
        # 退回預設順序
        mapping = {i + 1: k for i, k in enumerate(HEADER_KEYS)}

    def _value_for(key: str, row: dict) -> object:
        ff = row.get("final_fields") or {}
        if key == "batch_title":
            return batch_title
        if key == "volunteer_team":
            return volunteer_team
        if key == "_reviewer":
            return row.get("reviewer")
        if key == "_reviewed_at":
            return row.get("reviewed_at")
        return ff.get(key)

    for r_offset, row in enumerate(rows):
        for col_idx, key in mapping.items():
            val = _value_for(key, row)
            cell = ws.cell(row=start_row + r_offset, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def make_export_path(batch_id: int) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return settings.data_path / "exports" / f"batch_{batch_id}_{ts}.xlsx"
