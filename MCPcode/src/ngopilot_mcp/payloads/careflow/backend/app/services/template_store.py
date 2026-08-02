"""使用者上傳的自訂 Excel 模板管理。

設計：
- 任何時候只有 0 或 1 個「啟用中」模板。預設使用 v0.1 內建的 NGO 模板。
- 使用者上傳 .xlsx 時：解析第 1 列為標題，與內建 schema label 做 fuzzy 對應，
  生成 `{column_index: schema_key}` 映射，存入 `templates/active.json`。
- 匯出時若有啟用模板，就 load_workbook，並按映射寫入；無映射的欄位留空。
- 若使用者上傳圖片（表格照片），目前以「待 VLM 解析」狀態存檔，UI 提示後續手動指定欄位
  （demo 版以 placeholder 處理，避免引入過多 VLM 邏輯）。
"""
from __future__ import annotations

import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from ..config import settings
from .excel_export import HEADER_LABELS

MANIFEST_NAME = "active.json"


@dataclass
class TemplateInfo:
    kind: str            # "builtin" / "user_excel" / "user_image"
    file: str            # 相對 templates/ 的檔名
    uploaded_at: Optional[str]
    original_name: Optional[str]
    headers: list[str]            # 從 xlsx 讀到的標題列
    mapping: dict[str, str]       # column_letter or index_str -> schema_key
    note: Optional[str] = None


def _manifest_path() -> Path:
    return settings.data_path / "templates" / MANIFEST_NAME


def get_active() -> TemplateInfo:
    """回傳目前啟用中的模板資訊（無設定時回傳內建預設）。"""
    mp = _manifest_path()
    if mp.exists():
        data = json.loads(mp.read_text(encoding="utf-8"))
        return TemplateInfo(**data)
    return TemplateInfo(
        kind="builtin",
        file="volunteer_visit_template.xlsx",
        uploaded_at=None,
        original_name="(內建 NGO 通用模板)",
        headers=list(HEADER_LABELS.values()),
        mapping={str(i + 1): k for i, k in enumerate(HEADER_LABELS.keys())},
        note="預設模板。上傳自訂 Excel 即可覆蓋。",
    )


def save_active(info: TemplateInfo) -> None:
    _manifest_path().write_text(
        json.dumps(info.__dict__, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def reset_to_builtin() -> TemplateInfo:
    mp = _manifest_path()
    if mp.exists():
        mp.unlink()
    return get_active()


def _normalise(s: str) -> str:
    return (s or "").strip().replace(" ", "").replace("（", "(").replace("）", ")")


def _guess_mapping(headers: list[str]) -> dict[str, str]:
    """根據 HEADER_LABELS 做模糊對應；命中即 column_index -> schema_key。"""
    label_to_key = {_normalise(label): key for key, label in HEADER_LABELS.items()}
    out: dict[str, str] = {}
    for i, h in enumerate(headers):
        n = _normalise(h)
        if not n:
            continue
        # 直接命中
        if n in label_to_key:
            out[str(i + 1)] = label_to_key[n]
            continue
        # 包含關係
        for lab_norm, key in label_to_key.items():
            if lab_norm and (lab_norm in n or n in lab_norm):
                out[str(i + 1)] = key
                break
    return out


def upload_excel_template(file_bytes: bytes, original_name: str) -> TemplateInfo:
    """接收上傳的 .xlsx，存為 templates/user_<ts>.xlsx 並建立映射。"""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = settings.data_path / "templates" / f"user_{ts}.xlsx"
    target.write_bytes(file_bytes)

    try:
        wb = load_workbook(target, read_only=True, data_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(c) if c is not None else "" for c in first_row]
        wb.close()
    except Exception as exc:  # noqa: BLE001
        target.unlink(missing_ok=True)
        raise ValueError(f"無法解析 Excel：{exc}")

    if not any(headers):
        target.unlink(missing_ok=True)
        raise ValueError("Excel 第 1 列沒有偵測到標題文字。")

    info = TemplateInfo(
        kind="user_excel",
        file=target.name,
        uploaded_at=datetime.now().isoformat(timespec="seconds"),
        original_name=original_name,
        headers=headers,
        mapping=_guess_mapping(headers),
        note="若欄位對應錯誤，請刪除後重新上傳，或回退內建模板。",
    )
    save_active(info)
    return info


def upload_image_template(file_bytes: bytes, original_name: str) -> TemplateInfo:
    """上傳表格照片：先存檔，標記為待人工確認映射。MVP 版本不會立即生效。"""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = Path(original_name).suffix.lower() or ".jpg"
    target = settings.data_path / "templates" / f"user_image_{ts}{suffix}"
    target.write_bytes(file_bytes)
    # 維持內建模板有效，但記錄已上傳的圖片
    info = get_active()
    info.note = (
        f"已收到表格照片 {original_name}（{target.name}）。"
        "圖片模板需要視覺模型解析欄位，本版本暫以內建模板匯出，"
        "下一版將支援 VLM 自動框選欄位坐標。"
    )
    save_active(info)
    return info


def update_mapping(mapping: dict[str, str]) -> TemplateInfo:
    info = get_active()
    info.mapping = mapping
    save_active(info)
    return info


def active_template_file() -> Path:
    """匯出時使用的實際 .xlsx 路徑。"""
    info = get_active()
    if info.kind == "user_excel":
        return settings.data_path / "templates" / info.file
    # 預設使用內建（不存在則由 excel_export.ensure_template_exists 生成）
    from . import excel_export
    return excel_export.ensure_template_exists()
