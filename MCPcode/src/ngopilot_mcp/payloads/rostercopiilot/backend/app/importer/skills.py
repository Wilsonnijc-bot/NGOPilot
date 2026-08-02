"""Importer for the new-staff service-shadowing matrix."""
from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from .base import ImportResult
from .models import CellRef, ImportAmbiguity, ImportBatchSummary, ParsedRecord, SourceRef

DOC_REF = "docs/spec/excel_semantics.md §1.3; docs/spec/excel_io_contract.md §1"


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def _source(path: Path | None, ws: Worksheet, row: int, col: int,
            value: object = None) -> SourceRef:
    return SourceRef(
        workbook_path=path,
        sheet_name=ws.title,
        cell=CellRef(sheet_name=ws.title, row=row, column=col),
        raw_value=value,
        doc_ref=DOC_REF,
    )


def _worker_header(raw: str) -> tuple[str, str | None]:
    match = re.match(r"^(.*?)\((.*?)\)$", raw)
    if not match:
        return raw, None
    return match.group(1).strip(), match.group(2).strip()


def parse_skills_sheet(
    worksheet: Worksheet,
    *,
    workbook_path: Path | None = None,
) -> ImportResult[dict[str, Any]]:
    workers: list[tuple[int, str, str | None]] = []
    ambiguities: list[ImportAmbiguity] = []
    records: list[ParsedRecord[dict[str, Any]]] = []

    for col in range(3, worksheet.max_column + 1):
        raw = _normalize(worksheet.cell(1, col).value)
        if not raw:
            continue
        alias, join_raw = _worker_header(raw)
        workers.append((col, alias, join_raw))

    current_category: str | None = None
    category_by_row: dict[int, str] = {}
    for row in range(2, worksheet.max_row + 1):
        category = _normalize(worksheet.cell(row, 1).value)
        if category:
            current_category = category
        item = _normalize(worksheet.cell(row, 2).value)
        if not item and not category:
            continue
        if current_category is None:
            ambiguities.append(ImportAmbiguity(
                code="SKILL_CATEGORY_MISSING",
                message="skill row has item text before any category label",
                source=_source(workbook_path, worksheet, row, 2,
                               worksheet.cell(row, 2).value),
                severity="warning",
                raw_value=item,
            ))
            continue
        category_by_row[row] = current_category

    for col, alias, join_raw in workers:
        ticks = []
        unknown_blanks = 0
        for row, category in category_by_row.items():
            item = _normalize(worksheet.cell(row, 2).value)
            value = _normalize(worksheet.cell(row, col).value)
            if value.lower() in {"v", "✓", "✔", "yes", "y"}:
                ticks.append({
                    "category": category,
                    "item": item,
                    "source_ref": f"{worksheet.title}!{worksheet.cell(row, col).coordinate}",
                })
            elif value:
                ambiguities.append(ImportAmbiguity(
                    code="SKILL_TICK_UNRECOGNISED",
                    message="non-blank skill cell is not a recognised tick",
                    source=_source(workbook_path, worksheet, row, col,
                                   worksheet.cell(row, col).value),
                    severity="warning",
                    raw_value=value,
                ))
            else:
                unknown_blanks += 1
        record = {
            "source": "division.skills",
            "source_ref": f"{worksheet.title}!{worksheet.cell(1, col).coordinate}",
            "worker_alias": alias,
            "join_date_raw": join_raw,
            "ticks": ticks,
            "routes": [t["item"] for t in ticks if t["category"] == "送飯"],
            "skills": [t["item"] for t in ticks if t["category"] != "送飯"],
            "unknown_blank_count": unknown_blanks,
            "blank_semantics": "unknown",
        }
        records.append(ParsedRecord(
            record=record,
            source=_source(workbook_path, worksheet, 1, col,
                           worksheet.cell(1, col).value),
            raw={"header": worksheet.cell(1, col).value},
            parse_confidence="high",
        ))

    summary = ImportBatchSummary(
        parser_name="division.skills",
        status="ok",
        parsed_count=len(records),
        inferred_count=sum(len(r.record["ticks"]) for r in records if r.record),
        flagged_count=len(ambiguities),
        silently_dropped_cells=0,
        notes=("blank skill cells are preserved as unknown, not negative evidence",),
        doc_ref=DOC_REF,
    )
    return ImportResult(
        summary=summary,
        records=tuple(records),
        ambiguities=tuple(ambiguities),
    )
