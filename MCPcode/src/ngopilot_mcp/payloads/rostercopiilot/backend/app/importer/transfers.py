"""Importer for case transfer logs in the division workbook."""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from .base import ImportResult
from .models import CellRef, ImportAmbiguity, ImportBatchSummary, ParsedRecord, SourceRef

DOC_REF = "docs/spec/excel_semantics.md §1.2; docs/spec/excel_io_contract.md §1"
HEADERS = (
    "序號", "個案名稱", "所屬單位", "個案經理", "更改個案經理",
    "原訂照顧員", "更改照顧員", "原訂服務日期", "更改服務日期",
    "原訂服務時間", "更改服務時間", "生效日期",
)


def _normalize(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def _source(path: Path | None, ws: Worksheet, row: int, col: int,
            value: object = None) -> SourceRef:
    return SourceRef(
        workbook_path=path,
        sheet_name=ws.title,
        cell=CellRef(sheet_name=ws.title, row=row, column=col),
        raw_value=value,
        doc_ref=DOC_REF,
    )


def parse_transfer_log(
    worksheet: Worksheet,
    *,
    workbook_path: Path | None = None,
) -> ImportResult[dict[str, Any]]:
    records: list[ParsedRecord[dict[str, Any]]] = []
    ambiguities: list[ImportAmbiguity] = []
    header = [_normalize(worksheet.cell(1, c).value) for c in range(1, 13)]
    expected_without_seq = list(HEADERS[1:])
    if header[1:12] != expected_without_seq:
        ambiguities.append(ImportAmbiguity(
            code="TRANSFER_HEADER_UNEXPECTED",
            message="transfer sheet header does not match expected columns",
            source=_source(workbook_path, worksheet, 1, 1, worksheet.cell(1, 1).value),
            severity="blocking",
            raw_value=" | ".join(header),
        ))

    for row in range(2, worksheet.max_row + 1):
        first = worksheet.cell(row, 1).value
        values = [_normalize(worksheet.cell(row, col).value) for col in range(1, 13)]
        if not any(values):
            continue
        if first in (None, "") and not values[1]:
            continue
        raw_text = " | ".join(v for v in values if v)
        record = {
            "source": "division.transfers",
            "source_ref": f"{worksheet.title}!A{row}:L{row}",
            "row": row,
            "sequence": values[0] or None,
            "elder_full_name": values[1] or None,
            "unit": values[2] or None,
            "case_manager_before": values[3] or None,
            "case_manager_after": None if values[4] == "/" else values[4] or None,
            "worker_before": values[5] or None,
            "worker_after_raw": None if values[6] == "/" else values[6] or None,
            "service_day_before": values[7] or None,
            "service_day_after": None if values[8] == "/" else values[8] or None,
            "service_time_before": values[9] or None,
            "service_time_after": None if values[10] == "/" else values[10] or None,
            "effective_date_raw": values[11] or None,
            "status": "pending_review",
        }
        records.append(ParsedRecord(
            record=record,
            source=_source(workbook_path, worksheet, row, 1, first),
            raw={str(i): worksheet.cell(row, i).value for i in range(1, 13)},
            parse_confidence="medium" if "待定" in raw_text or "TBC" in raw_text.upper()
            else "high",
        ))
        if "TBC" in raw_text.upper():
            ambiguities.append(ImportAmbiguity(
                code="TRANSFER_TBC",
                message="transfer row contains TBC worker assignment",
                source=_source(workbook_path, worksheet, row, 7,
                               worksheet.cell(row, 7).value),
                severity="blocking",
                raw_value=raw_text,
                resolution_hint="confirm the actual replacement worker",
            ))
        if "待定" in raw_text or "?" in values[11]:
            ambiguities.append(ImportAmbiguity(
                code="TRANSFER_EFFECTIVE_DATE_UNCLEAR",
                message="transfer effective date is fuzzy or pending",
                source=_source(workbook_path, worksheet, row, 12,
                               worksheet.cell(row, 12).value),
                severity="blocking",
                raw_value=values[11],
                resolution_hint="confirm exact effective date before promotion",
            ))
        if values[1] and len(values[1]) >= 3:
            ambiguities.append(ImportAmbiguity(
                code="FULL_NAME_LEAK",
                message="transfer log uses full elder name; link to masked alias manually",
                source=_source(workbook_path, worksheet, row, 2,
                               worksheet.cell(row, 2).value),
                severity="warning",
                raw_value=values[1],
                resolution_hint="map full name to the workbook alias used in schedules",
            ))

    summary = ImportBatchSummary(
        parser_name="division.transfers",
        status="ok",
        parsed_count=len(records),
        inferred_count=0,
        flagged_count=len(ambiguities),
        silently_dropped_cells=0,
        notes=("TBC/effective-date-unclear rows remain review items",),
        doc_ref=DOC_REF,
    )
    return ImportResult(
        summary=summary,
        records=tuple(records),
        ambiguities=tuple(ambiguities),
    )
