"""Excel export: demo workbook with five sheets.

Not a pixel-perfect clone of the NGO layout (that is Phase 1's exporter);
this workbook proves the end-to-end flow and keeps the *semantics* of the
real file: worker columns, weekday×period rows, and the NGO's colour language
(yellow escorts, per-centre duty colours, pink home visits — see
docs/spec/excel_semantics.md §fill-colour census).
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..domain import (
    AuditStatus,
    EntryStatus,
    ImpactReport,
    MockDataset,
    Period,
    ScheduleEntry,
    ScheduleVersion,
    SERVICE_CATEGORY,
    ServiceCategory,
    ServiceCode,
)
from ..engine import week_dates

PERIOD_LABEL = {Period.AM: "上午", Period.PM: "下午"}
WEEKDAY_LABEL = ["一", "二", "三", "四", "五", "六", "日"]

# Colour language mirrors the real workbook (excel_semantics.md).
CATEGORY_FILL = {
    ServiceCode.EXERCISE: "F4CCCC",
    ServiceCode.HOME_CLEAN: "F9A825",
    ServiceCode.PERSONAL_CARE: "F4CCCC",
    ServiceCode.BATH: "F4CCCC",
    ServiceCode.ESCORT: "FFFF00",
    ServiceCode.MEAL: "F9CB9C",
    ServiceCode.DUTY_AMC: "FFE599",
    ServiceCode.DUTY_MRC: "D9EAD3",
    ServiceCode.DUTY_GC: "A4C2F4",
    ServiceCode.KITCHEN: "9FC5E8",
}
STATUS_FILL = {
    EntryStatus.NEEDS_REVIEW: "FDE9A9",   # amber: awaiting human review
    EntryStatus.AFFECTED: "FCE4D6",       # light orange: touched by a change
    EntryStatus.UNASSIGNED: "F8CBAD",     # red-ish: nobody assigned
    EntryStatus.CANCELLED: "D9D9D9",      # gray
}
SEVERITY_FILL = {"high": "F8CBAD", "warning": "FDE9A9", "info": "DEEBF7"}

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="305496")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER


def _body(cell, fill: str | None = None, strike: bool = False) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = _BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if strike:
        cell.font = Font(strike=True, color="808080")


def _entry_label(e: ScheduleEntry) -> str:
    target = e.elder_name or e.center or e.route or e.destination or ""
    line = f"{e.service_code.value}:{target}" if target else e.service_code.value
    if e.start_time:
        line += f"\n{e.start_time.strftime('%H:%M')}"
        if e.end_time:
            line += f"-{e.end_time.strftime('%H:%M')}"
    marker = {
        EntryStatus.NEEDS_REVIEW: "⚠待審",
        EntryStatus.AFFECTED: "⚠受影響",
        EntryStatus.CANCELLED: "✖取消",
        EntryStatus.UNASSIGNED: "❗待分配",
    }.get(e.status)
    if marker:
        line += f"\n[{marker}]"
    return line


def _cell_fill(entries: list[ScheduleEntry]) -> str | None:
    # status colour wins over category colour (problems must be visible)
    for e in entries:
        if e.status in STATUS_FILL and e.status != EntryStatus.CANCELLED:
            return STATUS_FILL[e.status]
    if all(e.status == EntryStatus.CANCELLED for e in entries):
        return STATUS_FILL[EntryStatus.CANCELLED]
    return CATEGORY_FILL.get(entries[0].service_code)


# ---------------------------------------------------------------- sheet 1
def _sheet_master(wb: Workbook, dataset: MockDataset, version: ScheduleVersion) -> None:
    ws = wb.active
    ws.title = "總排班"
    ws.freeze_panes = "C3"

    ws.cell(1, 1, f"RosterCopiilot 週排班 {version.week_start.isoformat()} "
                  f"(版本 {version.id} · {version.kind.value})").font = Font(bold=True, size=14)

    workers = sorted(dataset.employees, key=lambda w: w.id)
    ws.cell(2, 1, "日期")
    ws.cell(2, 2, "時段")
    _header(ws.cell(2, 1)); _header(ws.cell(2, 2))
    for col, w in enumerate(workers, start=3):
        c = ws.cell(2, col, f"{w.display_name}\n({w.id})")
        _header(c)
        ws.column_dimensions[get_column_letter(col)].width = 16

    by_slot: dict[tuple, list[ScheduleEntry]] = defaultdict(list)
    for e in version.entries:
        if e.worker_id:
            by_slot[(e.schedule_date, e.period, e.worker_id)].append(e)

    row = 3
    for on in week_dates(version.week_start):
        for period in (Period.AM, Period.PM):
            ws.cell(row, 1, f"{on.strftime('%m-%d')} 週{WEEKDAY_LABEL[on.isoweekday() - 1]}")
            ws.cell(row, 2, PERIOD_LABEL[period])
            _body(ws.cell(row, 1)); _body(ws.cell(row, 2))
            for col, w in enumerate(workers, start=3):
                entries = sorted(by_slot.get((on, period, w.id), []),
                                 key=lambda e: (e.session_index or 0, e.id))
                cell = ws.cell(row, col)
                if entries:
                    cell.value = "\n———\n".join(_entry_label(e) for e in entries)
                    _body(cell, _cell_fill(entries),
                          strike=all(e.status == EntryStatus.CANCELLED
                                     for e in entries))
                else:
                    _body(cell)
            ws.row_dimensions[row].height = 64
            row += 1
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 7


# ---------------------------------------------------------------- sheet 2
def _sheet_escort(wb: Workbook, version: ScheduleVersion) -> None:
    ws = wb.create_sheet("護送時間表")
    ws.freeze_panes = "A2"
    headers = ["日期", "上/下午", "長者", "應診時間", "目的地", "交通", "護送同工",
               "狀態", "說明"]
    for col, h in enumerate(headers, start=1):
        _header(ws.cell(1, col, h))
    escorts = sorted(
        (e for e in version.entries if e.service_code == ServiceCode.ESCORT),
        key=lambda e: (e.schedule_date, e.period.value,
                       e.start_time or datetime.min.time()))
    for r, e in enumerate(escorts, start=2):
        vals = [e.schedule_date.isoformat(), PERIOD_LABEL[e.period],
                e.elder_name or "", e.start_time.strftime("%H:%M") if e.start_time else "",
                e.destination or "", e.notes or "", e.worker_name or "",
                e.status.value, e.explanation or ""]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(r, col, v)
            _body(cell, STATUS_FILL.get(e.status),
                  strike=e.status == EntryStatus.CANCELLED)
    for col, w in enumerate([11, 8, 10, 9, 18, 12, 10, 12, 60], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------- sheet 3
def _sheet_review(wb: Workbook, version: ScheduleVersion) -> None:
    ws = wb.create_sheet("人工審核")
    ws.freeze_panes = "A2"
    headers = ["編號", "狀態", "須批准", "風險", "類型", "原因", "結構化理由",
               "原安排", "建議安排", "其他選項", "人工備註"]
    for col, h in enumerate(headers, start=1):
        _header(ws.cell(1, col, h))
    order = {"high": 0, "warning": 1, "info": 2}
    items = sorted(version.audit_items,
                   key=lambda a: (a.status != AuditStatus.PENDING, not a.blocking,
                                  order[a.severity.value], a.id))
    for r, a in enumerate(items, start=2):
        vals = [
            a.id, a.status.value, "是" if a.blocking else "",
            a.severity.value, a.kind.value, a.reason,
            "\n".join(f"[{x.code.value}] {x.message}"
                      + (f"（{x.rule_ref}）" if x.rule_ref else "")
                      for x in a.reasons),
            _entry_label(a.original_entry) if a.original_entry else "",
            _entry_label(a.suggested_entry) if a.suggested_entry else "",
            "\n".join(f"{alt.worker_name}: {alt.explanation or ''}"
                      for alt in a.alternatives),
            a.human_note or "",
        ]
        for col, v in enumerate(vals, start=1):
            _body(ws.cell(r, col, v), SEVERITY_FILL.get(a.severity.value)
                  if col <= 5 else None)
    for col, w in enumerate([12, 9, 7, 8, 20, 46, 46, 22, 22, 34, 20], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------- sheet 4
def _sheet_unassigned(wb: Workbook, version: ScheduleVersion) -> None:
    ws = wb.create_sheet("未分配")
    ws.freeze_panes = "A2"
    headers = ["日期", "時段", "服務", "對象", "地區", "原因（結構化）", "說明"]
    for col, h in enumerate(headers, start=1):
        _header(ws.cell(1, col, h))
    rows = [e for e in version.entries if e.status == EntryStatus.UNASSIGNED]
    rows.sort(key=lambda e: (e.schedule_date, e.period.value, e.id))
    for r, e in enumerate(rows, start=2):
        vals = [e.schedule_date.isoformat(), PERIOD_LABEL[e.period],
                e.service_code.value,
                e.elder_name or e.center or e.route or "",
                e.district or "",
                "\n".join(f"[{x.code.value}] {x.message}" for x in e.review_reasons),
                e.explanation or ""]
        for col, v in enumerate(vals, start=1):
            _body(ws.cell(r, col, v), "F8CBAD" if col == 1 else None)
    for col, w in enumerate([11, 8, 8, 14, 10, 52, 52], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------- sheet 5
def _sheet_impact(wb: Workbook, version: ScheduleVersion,
                  reports: list[ImpactReport] | None) -> None:
    ws = wb.create_sheet("變更影響")
    ws.freeze_panes = "A2"
    headers = ["變更事件", "日期", "風險", "須審批", "影響", "說明"]
    for col, h in enumerate(headers, start=1):
        _header(ws.cell(1, col, h))
    r = 2
    if not version.trigger_events:
        _body(ws.cell(r, 1, "基準排班——本版本沒有變更事件"))
    for report in reports or []:
        ev = report.event
        head = ws.cell(r, 1, f"{ev.id or ''} {ev.type.value}")
        _body(head, SEVERITY_FILL.get(report.risk_level.value))
        _body(ws.cell(r, 2, ev.change_date.isoformat()))
        _body(ws.cell(r, 3, report.risk_level.value))
        _body(ws.cell(r, 4, "是" if report.requires_review else ""))
        _body(ws.cell(r, 5, report.summary))
        _body(ws.cell(r, 6, ev.reason or ""))
        r += 1
        for imp in report.impacts:
            _body(ws.cell(r, 2, imp.title), SEVERITY_FILL.get(imp.severity.value))
            _body(ws.cell(r, 3, imp.severity.value))
            _body(ws.cell(r, 4, "是" if imp.requires_review else ""))
            _body(ws.cell(r, 5, imp.description))
            _body(ws.cell(r, 6, "影響同工: " + ",".join(imp.affected_worker_ids)))
            r += 1
    r += 1
    _body(ws.cell(r, 1, "指標摘要"), "DEEBF7")
    r += 1
    for k, v in sorted(version.summary.items()):
        _body(ws.cell(r, 1, k))
        _body(ws.cell(r, 2, v))
        r += 1
    for col, w in enumerate([26, 22, 9, 9, 60, 40], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------- public
def build_workbook(dataset: MockDataset, version: ScheduleVersion,
                   reports: list[ImpactReport] | None = None) -> Workbook:
    wb = Workbook()
    _sheet_master(wb, dataset, version)
    _sheet_escort(wb, version)
    _sheet_review(wb, version)
    _sheet_unassigned(wb, version)
    _sheet_impact(wb, version, reports)
    return wb


def save_workbook(dataset: MockDataset, version: ScheduleVersion,
                  reports: list[ImpactReport] | None = None,
                  output_dir: Path | None = None) -> Path:
    output_dir = output_dir or Path(__file__).resolve().parents[3] / "data" / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"rostercopiilot_{version.week_start.isoformat()}_{version.id}_{ts}.xlsx"
    build_workbook(dataset, version, reports).save(path)
    return path
